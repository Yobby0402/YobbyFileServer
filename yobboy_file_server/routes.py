# routes.py
import os
import sys
import re
import configparser
import json
import logging
import math
import tempfile
import unicodedata
from datetime import datetime
from typing import Any, Dict, Optional
# 确保在文件顶部添加必要的导入
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_from_directory, send_file, make_response, abort, current_app, Response
from werkzeug.exceptions import RequestEntityTooLarge
import markdown
from markdown_it import MarkdownIt
from mdit_py_plugins import tasklists, deflist, footnote
from urllib.parse import quote # 导入 quote 用于编码文件名
import posixpath # 用于处理 URL 路径
from .share_links import ShareLinkManager  # 导入分享链接管理器
from .todo_manager import todo_manager
from .todo_extended_manager import todo_extended_manager
from . import todo_report_builder
from .serial_manager import serial_manager
from .shared_serial_hub import shared_serial_hub
from .product_compare_manager import product_compare_manager
from .logging_setup import parse_log_level
from .paths import project_base_dir, project_path
from . import todo_kb_store
from . import product_compare_kb_store

_routes_logger = logging.getLogger('yobboy_file_server.routes')


def log_message(message, level='INFO'):
    """路由层 Git 相关日志（委托给统一 logging 树）。"""
    _routes_logger.log(parse_log_level(level, logging.INFO), '%s', message)


# 检查用户是否已登录的函数
def is_logged_in():
    """检查用户是否已登录"""
    return 'logged_in' in session


WINDOWS_RESERVED_FILENAMES = {
    "CON", "PRN", "AUX", "NUL",
    "COM1", "COM2", "COM3", "COM4", "COM5", "COM6", "COM7", "COM8", "COM9",
    "LPT1", "LPT2", "LPT3", "LPT4", "LPT5", "LPT6", "LPT7", "LPT8", "LPT9",
}


def _is_path_under_root(root_path: str, candidate_path: str) -> bool:
    """Return True when candidate_path is inside root_path, using path boundaries."""
    try:
        root_abs = os.path.abspath(root_path)
        candidate_abs = os.path.abspath(candidate_path)
        common = os.path.commonpath([root_abs, candidate_abs])
    except (OSError, ValueError):
        return False
    return os.path.normcase(common) == os.path.normcase(root_abs)


def _resolve_upload_target_dir(root_path: str, browser_path: str) -> str:
    """Resolve a browser-relative path into an existing upload target directory."""
    if not root_path:
        raise ValueError("文件根目录未配置")

    root_abs = os.path.abspath(root_path)
    raw_path = (browser_path or "").replace("\\", "/").strip()
    if raw_path.startswith("//") or re.match(r"^[A-Za-z]:", raw_path):
        raise PermissionError("无权访问该路径")

    raw_path = raw_path.lstrip("/")
    normalized = posixpath.normpath(raw_path) if raw_path else ""
    if normalized in ("", "."):
        target_dir = root_abs
    elif normalized == ".." or normalized.startswith("../"):
        raise PermissionError("无权访问该路径")
    else:
        target_dir = os.path.abspath(os.path.join(root_abs, *normalized.split("/")))

    if not _is_path_under_root(root_abs, target_dir):
        raise PermissionError("无权访问该路径")
    if not os.path.exists(target_dir):
        raise FileNotFoundError("上传目录不存在")
    if not os.path.isdir(target_dir):
        raise NotADirectoryError("上传目标不是目录")
    return target_dir


def _clean_upload_filename(filename: str) -> str:
    """Normalize a browser supplied filename while preserving non-ASCII names."""
    cleaned = unicodedata.normalize("NFC", filename or "")
    cleaned = cleaned.replace("\x00", "").replace("\\", "/")
    cleaned = posixpath.basename(cleaned).strip().rstrip(" .")

    if not cleaned or cleaned in (".", ".."):
        raise ValueError("文件名不能为空")
    if any(ord(ch) < 32 for ch in cleaned):
        raise ValueError(f"文件名包含非法控制字符: {cleaned}")
    if any(ch in cleaned for ch in '<>:"|?*/\\'):
        raise ValueError(f"文件名包含非法字符: {cleaned}")

    stem = cleaned.split(".", 1)[0].upper()
    if stem in WINDOWS_RESERVED_FILENAMES:
        raise ValueError(f"文件名为系统保留名称: {cleaned}")
    return cleaned


def _save_upload_atomically(file_storage, target_dir: str, filename: str) -> str:
    """Save an uploaded file through a temporary file, then atomically replace target."""
    fd, temp_path = tempfile.mkstemp(prefix=f".{filename}.", suffix=".uploading", dir=target_dir)
    os.close(fd)
    final_path = os.path.join(target_dir, filename)

    try:
        file_storage.save(temp_path)
        if not _is_path_under_root(target_dir, final_path):
            raise PermissionError("无权访问该路径")
        os.replace(temp_path, final_path)
    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError:
                _routes_logger.warning("清理上传临时文件失败: %s", temp_path)
    return final_path


def _upload_size_error_response():
    max_bytes = current_app.config.get("MAX_CONTENT_LENGTH")
    max_mb = int(max_bytes / (1024 * 1024)) if max_bytes else None
    if max_mb:
        message = f"上传文件过大，当前限制为 {max_mb} MB"
    else:
        message = "上传文件过大"
    return jsonify({'success': False, 'error': message}), 413


def handle_upload_files_request():
    """Handle file uploads for the browser file manager."""
    if not is_logged_in():
        return jsonify({'success': False, 'error': '未登录'}), 401

    try:
        files = request.files.getlist('files')
        if not files:
            return jsonify({'success': False, 'error': '没有文件'}), 400

        path = request.form.get('path', '')
        root_path = current_app.config['ROOT_DIR']
        target_dir = _resolve_upload_target_dir(root_path, path)

        uploaded = []
        failed = []
        for file in files:
            original_name = file.filename or ''
            try:
                filename = _clean_upload_filename(original_name)
                file_path = _save_upload_atomically(file, target_dir, filename)
                uploaded.append({
                    'name': filename,
                    'size': os.path.getsize(file_path),
                })
            except Exception as file_error:
                failed.append({
                    'name': original_name or '(未命名文件)',
                    'error': str(file_error),
                })
                _routes_logger.warning("上传文件失败: %s -> %s", original_name, file_error)

        count = len(uploaded)
        failed_count = len(failed)
        if count == 0:
            return jsonify({
                'success': False,
                'error': failed[0]['error'] if failed else '没有可上传的有效文件',
                'count': 0,
                'failed': failed_count,
                'errors': failed,
            }), 400

        if failed_count:
            return jsonify({
                'success': False,
                'error': f'已上传 {count} 个文件，{failed_count} 个文件失败',
                'count': count,
                'failed': failed_count,
                'files': uploaded,
                'errors': failed,
            }), 207

        return jsonify({
            'success': True,
            'message': f'成功上传{count}个文件',
            'count': count,
            'failed': 0,
            'files': uploaded,
        })
    except PermissionError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except (FileNotFoundError, NotADirectoryError, ValueError) as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except RequestEntityTooLarge:
        return _upload_size_error_response()
    except Exception as e:
        _routes_logger.exception("上传失败")
        return jsonify({'success': False, 'error': f'上传失败: {str(e)}'}), 500


REMOTE_SERIAL_HTTPS_MODE_FULL = 'full'
REMOTE_SERIAL_HTTPS_MODE_COMPAT = 'compat'


def normalize_remote_serial_https_mode(value):
    """Normalize the stored remote-serial HTTPS mode."""
    mode = (value or '').strip().lower()
    if mode == REMOTE_SERIAL_HTTPS_MODE_COMPAT:
        return REMOTE_SERIAL_HTTPS_MODE_COMPAT
    return REMOTE_SERIAL_HTTPS_MODE_FULL


def get_serial_https_port():
    """Return the configured HTTPS port for the serial companion listener."""
    try:
        main_port = int(current_app.config.get('PORT', 5000))
    except (TypeError, ValueError):
        main_port = 5000
    default_port = 5444 if main_port == 5443 else 5443
    try:
        port = int(current_app.config.get('SERIAL_HTTPS_PORT', default_port))
        if 1 <= port <= 65535:
            return port
    except (TypeError, ValueError):
        pass
    return default_port


def build_serial_https_url(path=None):
    """Build the HTTPS URL used by the serial tool in compatibility mode."""
    host = request.host.split(':', 1)[0]
    target_path = path or request.path
    query_string = request.query_string.decode('utf-8', errors='ignore')
    if query_string:
        target_path = f"{target_path}?{query_string}"
    return f"https://{host}:{get_serial_https_port()}{target_path}"


def get_serial_server_config(page_path=None):
    """Build serial-tool runtime config shared with templates."""
    remote_serial_enabled = bool(current_app.config.get('REMOTE_SERIAL_ENABLED', False))
    remote_serial_https_mode = normalize_remote_serial_https_mode(
        current_app.config.get('REMOTE_SERIAL_HTTPS_MODE', REMOTE_SERIAL_HTTPS_MODE_FULL)
    )
    serial_https_active = bool(current_app.config.get('SERIAL_HTTPS_ACTIVE', False))
    cert_file = (current_app.config.get('HTTPS_CERT_FILE') or '').strip()
    key_file = (current_app.config.get('HTTPS_KEY_FILE') or '').strip()
    serial_https_url = build_serial_https_url(page_path or request.path)
    return {
        'remote_serial_enabled': remote_serial_enabled,
        'https_expected': remote_serial_enabled,
        'https_cert_configured': bool(cert_file and key_file),
        'https_mode': remote_serial_https_mode,
        'compat_mode': remote_serial_https_mode == REMOTE_SERIAL_HTTPS_MODE_COMPAT,
        'serial_https_port': get_serial_https_port(),
        'serial_https_active': serial_https_active,
        'serial_https_url': serial_https_url,
    }

# 创建markdown-it实例，支持多种扩展
def create_markdown_parser():
    """创建配置好的markdown-it解析器"""
    md = MarkdownIt("default", {"breaks": True, "html": True})
    
    # 启用内建规则以支持表格与删除线
    md.enable(["table", "strikethrough"]) 
    
    # 启用插件
    md.use(tasklists.tasklists_plugin)
    md.use(deflist.deflist_plugin)
    md.use(footnote.footnote_plugin)
    
    return md

# 全局markdown解析器实例
markdown_parser = create_markdown_parser()

# 处理图片路径的函数
def process_image_paths(content, current_file_path):
    """处理Markdown内容中的图片路径"""
    # 定义正则表达式匹配Markdown图片语法
    img_pattern = re.compile(r'!\[(.*?)\]\(([^\s\)]+)(?:\s+"([^"]*)")?\)')
    
    def replace_img_path(match):
        alt_text = match.group(1)
        img_path = match.group(2)
        title = match.group(3)
        
        # 跳过已经是绝对URL或以/download或/preview开头的路径
        if img_path.startswith(('http://', 'https://', '/download', '/preview')):
            return match.group(0)
        
        # 处理相对路径
        parent_dir = posixpath.dirname(current_file_path)
        if parent_dir:
            # 如果是相对于当前文件的路径
            new_path = posixpath.join(parent_dir, img_path)
        else:
            # 如果当前在根目录
            new_path = img_path
        
        # 生成预览链接
        preview_url = f'/preview/{new_path}'
        
        # 重新组合图片标记
        if title:
            return f'![{alt_text}]({preview_url} "{title}")'
        else:
            return f'![{alt_text}]({preview_url})'
    
    # 执行替换
    return img_pattern.sub(replace_img_path, content)

# 渲染Markdown内容
def render_markdown_content(content, filepath):
    """使用markdown-it-py渲染Markdown内容"""
    try:
        # 先处理图片路径
        processed_content = process_image_paths(content, filepath)
        
        # 使用markdown-it-py渲染
        html_content = markdown_parser.render(processed_content)
        
        return html_content
    except Exception as e:
        return f"<p>渲染Markdown时出错: {e}</p>"

# 文件类型常量定义
OFFICE_EXTENSIONS = ['.docx', '.xlsx', '.pptx']
IMAGE_EXTENSIONS = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.svg', '.webp', '.avif', '.ico']
MARKDOWN_EXTENSIONS = ['.md', '.markdown']
PDF_EXTENSIONS = ['.pdf']
VIDEO_EXTENSIONS = ['.mp4', '.avi', '.mov', '.wmv', '.webm', '.m4v']
AUDIO_EXTENSIONS = ['.mp3', '.wav', '.flac', '.ogg', '.wma', '.m4a', '.aac', '.opus', '.weba', '.aiff', '.aif', '.amr']  # 音频文件格式
DRAWIO_EXTENSIONS = ['.drawio', '.diagram', '.dio', '.xml']  # 添加.xml作为draw.io格式
MODEL_3D_EXTENSIONS = ['.gltf', '.glb', '.obj', '.stl', '.fbx', '.step', '.stp']  # 3D模型文件格式

# 获取data文件夹路径
def get_data_path(relative_path=''):
    """获取data文件夹路径"""
    if getattr(sys, 'frozen', False):
        # 打包环境：exe 所在目录
        base_path = project_base_dir()
    else:
        # 开发环境：.py 文件所在目录
        base_path = project_base_dir()
    data_dir = os.path.join(base_path, 'data')
    if relative_path:
        return os.path.join(data_dir, relative_path)
    return data_dir

# 获取收藏文件路径
def get_favorites_path():
    """获取收藏文件路径"""
    return get_data_path('file_server/favorites.json')

# 加载收藏数据
def load_favorites():
    """加载收藏数据"""
    favorites_path = get_favorites_path()
    try:
        os.makedirs(os.path.dirname(favorites_path), exist_ok=True)
        if os.path.exists(favorites_path):
            with open(favorites_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        else:
            return {'groups': [], 'items': []}
    except Exception as e:
        _routes_logger.warning("加载收藏数据失败: %s", e)
        return {'groups': [], 'items': []}

# 保存收藏数据
def save_favorites(data):
    """保存收藏数据"""
    favorites_path = get_favorites_path()
    try:
        os.makedirs(os.path.dirname(favorites_path), exist_ok=True)
        with open(favorites_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        return True
    except Exception as e:
        _routes_logger.warning("保存收藏数据失败: %s", e)
        return False

def _excel_display_width(value):
    """估算单元格内容显示宽度，兼顾中文和换行。"""
    if value is None:
        return 0.0

    text = str(value)
    if not text:
        return 0.0

    max_line_width = 0.0
    for line in text.split('\n'):
        line_width = 0.0
        for char in line:
            if char == '\t':
                line_width += 4
            elif unicodedata.east_asian_width(char) in ('F', 'W'):
                line_width += 2
            elif ord(char) < 128:
                line_width += 1
            else:
                line_width += 1.5
        max_line_width = max(max_line_width, line_width)

    return max_line_width


def _excel_estimate_wrapped_lines(value, column_width):
    """根据列宽估算换行后的行数。"""
    if value is None:
        return 1

    text = str(value)
    if not text:
        return 1

    available_width = max(float(column_width or 0) - 1.5, 1)
    total_lines = 0

    for line in text.split('\n'):
        display_width = _excel_display_width(line)
        total_lines += max(1, math.ceil(display_width / available_width))

    return max(total_lines, 1)


def _excel_apply_auto_dimensions(ws, min_width=8, max_width=60, header_row=1, width_hints=None):
    """为工作表自动调整列宽和行高，避免内容被截断。"""
    if ws.max_row < header_row or ws.max_column < 1:
        return

    from openpyxl.utils import get_column_letter

    normalized_hints = {}
    if width_hints:
        for col_letter, hint in width_hints.items():
            try:
                normalized_hints[str(col_letter).upper()] = float(hint)
            except (TypeError, ValueError):
                continue

    final_widths = {}

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_content_width = 0.0

        for row_idx in range(header_row, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value in (None, ''):
                continue
            max_content_width = max(max_content_width, _excel_display_width(value))

        existing_width = ws.column_dimensions[col_letter].width or 0
        desired_width = max(
            float(min_width),
            max_content_width + 2.5,
            float(existing_width),
            normalized_hints.get(col_letter, 0.0),
        )
        final_width = min(desired_width, float(max_width))
        ws.column_dimensions[col_letter].width = final_width
        final_widths[col_letter] = final_width

    for row_idx in range(header_row, ws.max_row + 1):
        max_lines = 1

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value in (None, ''):
                continue

            col_letter = get_column_letter(col_idx)
            column_width = final_widths.get(col_letter, min_width)
            alignment = getattr(cell, 'alignment', None)
            wrap_text = bool(alignment and alignment.wrap_text)

            if wrap_text:
                max_lines = max(max_lines, _excel_estimate_wrapped_lines(value, column_width))
            else:
                max_lines = max(max_lines, max(1, str(value).count('\n') + 1))

        base_height = 24 if row_idx == header_row else 20
        desired_height = base_height + max(0, max_lines - 1) * 16
        existing_height = ws.row_dimensions[row_idx].height or 0
        ws.row_dimensions[row_idx].height = min(max(desired_height, existing_height, base_height), 240)

# 修复init_app函数内部的Draw.io路由

def init_app(app):
    """初始化路由"""
    global current_app
    current_app = app

    @app.errorhandler(RequestEntityTooLarge)
    def handle_request_entity_too_large(error):
        return _upload_size_error_response()
    
    # 初始化分享链接管理器
    share_manager = ShareLinkManager()
    app.config['SHARE_MANAGER'] = share_manager

    # 初始化 ToDo 管理器
    app.config['TODO_MANAGER'] = todo_manager
    
    # 初始化 ToDo 扩展数据管理器
    app.config['TODO_EXTENDED_MANAGER'] = todo_extended_manager
    
    # 初始化产品对比管理器
    app.config['PRODUCT_COMPARE_MANAGER'] = product_compare_manager

    from .local_ai_paths import ensure_ai_layout
    ensure_ai_layout()

    from .local_ai_routes import register_local_ai_routes
    register_local_ai_routes(app)

    from .local_erp_routes import register_local_erp_routes
    register_local_erp_routes(app)

    def _todo_kb_root_dir() -> str:
        return str(app.config.get('ROOT_DIR') or project_base_dir())

    def _todo_kb_app_config() -> Dict[str, Any]:
        return {
            "LOCAL_AI_API_BASE_URL": app.config.get("LOCAL_AI_API_BASE_URL", ""),
            "LOCAL_AI_EMBED_API_BASE_URL": app.config.get(
                "LOCAL_AI_EMBED_API_BASE_URL",
                app.config.get("LOCAL_AI_API_BASE_URL", ""),
            ),
            "LOCAL_AI_EMBED_MODEL": app.config.get("LOCAL_AI_EMBED_MODEL", ""),
            "LOCAL_AI_EMBED_API_KEY": app.config.get("LOCAL_AI_EMBED_API_KEY", "lm-studio"),
            "LOCAL_AI_EMBED_QUERY_INSTRUCTION": app.config.get(
                "LOCAL_AI_EMBED_QUERY_INSTRUCTION",
                "Represent this query for retrieving relevant passages from the local knowledge base.",
            ),
            "LOCAL_AI_EMBED_BATCH_SIZE": app.config.get("LOCAL_AI_EMBED_BATCH_SIZE", 16),
        }

    def _todo_kb_rebuild_all_async() -> None:
        return None

    def _product_compare_kb_rebuild_all_async() -> None:
        return None

    # 初始化Draw.io静态文件目录（静默检查，不影响运行）
    # Draw.io是可选功能，不存在也不影响文件浏览器功能
    
    # 检查Git功能是否启用
    git_enabled = app.config.get('GIT_ENABLED', False)
    
    if git_enabled:
        # 尝试加载Git模块
        try:
            from .git_manager import GitManager
            from .git_config_manager import GitConfigManager
            from .git_routes import register_git_routes
            
            git_manager = GitManager()
            git_config_manager = GitConfigManager()
            
            app.config['GIT_MANAGER'] = git_manager
            app.config['GIT_CONFIG_MANAGER'] = git_config_manager
            
            # 注册Git路由
            register_git_routes(app)
            
            _routes_logger.info("Git功能已启用")
        except ImportError as e:
            _routes_logger.warning("Git功能启用失败: %s", e)
            _routes_logger.warning("请安装 GitPython: pip install GitPython")
            app.config['GIT_ENABLED'] = False
        except Exception as e:
            _routes_logger.error("Git功能初始化失败: %s", e)
            app.config['GIT_ENABLED'] = False
    else:
        _routes_logger.info("Git功能已禁用（可在设置中启用）")
        # 即使Git功能未启用，也注册所有Git路由返回403，避免404错误
        # 注意：/api/git/status 不在这个列表中，因为它需要总是可用来检查Git状态
        git_routes = [
            '/api/git/check',
            '/api/git/init',
            '/api/git/clone',
            '/api/git/pull',
            '/api/git/push',
            '/api/git/commit',
            '/api/git/tag/create',
            '/api/git/tag/list',
            '/api/git/branch/list',
            '/api/git/branch/checkout',
            '/api/git/config/list',
            '/api/git/config/save',
            '/api/git/config/delete',
            '/api/git/config/set_default',
            '/api/git/config/test',
            '/api/git/repos/list'
        ]
        
        def create_disabled_route(route_path, methods=['GET', 'POST']):
            def disabled_handler():
                return jsonify({'success': False, 'error': 'Git功能未启用', 'git_enabled': False}), 403
            disabled_handler.__name__ = f"git_disabled_{route_path.replace('/', '_').replace('-', '_')}"
            return disabled_handler
        
        for route_path in git_routes:
            # 根据路由确定方法
            if route_path in ['/api/git/check', '/api/git/tag/list', 
                            '/api/git/branch/list', '/api/git/config/list', '/api/git/repos/list']:
                methods_list = ['GET']
            else:
                methods_list = ['POST']
            
            handler = create_disabled_route(route_path, methods_list)
            app.route(route_path, methods=methods_list)(handler)
    
    # 添加一个简单的API端点来检查Git启用状态（这个端点总是可用的，无论Git是否启用）
    # 注意：路径使用 /api/git/config/status 避免与 git_routes.py 中的 /api/git/status 冲突
    @app.route('/api/git/config/status', methods=['GET'])
    def git_config_status():
        """检查Git功能启用状态"""
        git_enabled = app.config.get('GIT_ENABLED', False)
        has_manager = 'GIT_MANAGER' in app.config
        git_workdir = app.config.get('GIT_WORKDIR', '')
        
        return jsonify({
            'git_enabled': git_enabled,
            'has_manager': has_manager,
            'git_workdir': git_workdir,
            'workdir_exists': os.path.exists(git_workdir) if git_workdir else False
        })
    
    def has_todo_access():
        return is_logged_in() or session.get('todo_direct_access')

    @app.route('/')
    def index():
        """首页，显示操作选择界面"""
        # 添加登录验证
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        
        return render_template('choice.html')

    @app.route('/todo/v2')
    def todo_v2_page():
        """新版ToDo界面（需登录）"""
        if not is_logged_in():
            return redirect(url_for('login'))
        session.pop('todo_direct_access', None)
        return render_template('todo_v2.html', direct_access=False)

    @app.route('/todo/v2/direct')
    def todo_v2_direct():
        """新版ToDo后门入口（免登录）"""
        session['todo_direct_access'] = True
        return render_template('todo_v2.html', direct_access=True)

    @app.route('/todo/v2/preview')
    def todo_v2_preview():
        """ToDo表格预览界面（需登录）"""
        if not is_logged_in():
            return redirect(url_for('login'))
        session.pop('todo_direct_access', None)
        return render_template('todo_v2_preview.html', direct_access=False)

    @app.route('/todo/v2/preview/direct')
    def todo_v2_preview_direct():
        """ToDo表格预览界面后门入口（免登录）"""
        session['todo_direct_access'] = True
        return render_template('todo_v2_preview.html', direct_access=True)

    @app.route('/api/todo/items', methods=['GET', 'POST'])
    def todo_collection():
        """ToDo 列表查询与创建"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']

        if request.method == 'GET':
            state = manager.list_todos()
            timeline = manager.get_timeline()
            return jsonify({
                'success': True,
                'todos': state['todos'],
                'projects': state.get('projects', {}),
                'timeline': timeline
            })

        payload = request.json or {}
        try:
            todo, event = manager.create_todo(payload)
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

        timeline = manager.get_timeline()
        return jsonify({
            'success': True,
            'todo': todo,
            'event': event,
            'timeline': timeline,
            'projects': manager.get_projects()
        })

    @app.route('/api/todo/items/<todo_id>', methods=['GET', 'PUT', 'DELETE'])
    def todo_item(todo_id):
        """ToDo 单项查询、更新与删除"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']

        if request.method == 'GET':
            try:
                todo = manager.get_todo(todo_id)
            except ValueError:
                return jsonify({'success': False, 'error': '未找到 ToDo'}), 404
            return jsonify({'success': True, 'todo': todo})

        if request.method == 'PUT':
            payload = request.json or {}
            try:
                todo, event = manager.update_todo(todo_id, payload)
            except ValueError as exc:
                return jsonify({'success': False, 'error': str(exc)}), 404
            except Exception as exc:
                return jsonify({'success': False, 'error': str(exc)}), 400

            timeline = manager.get_timeline()
            response = {
                'success': True,
                'todo': todo,
                'event': event,
                'timeline': timeline,
                'projects': manager.get_projects()
            }
            return jsonify(response)

        # DELETE
        try:
            manager.delete_todo(todo_id)
        except ValueError:
            return jsonify({'success': False, 'error': '未找到 ToDo'}), 404
        timeline = manager.get_timeline()
        return jsonify({
            'success': True,
            'todo_id': todo_id,
            'timeline': timeline,
            'projects': manager.get_projects()
        })

    @app.route('/api/todo/items/<todo_id>/comments', methods=['POST'])
    def todo_comment(todo_id):
        """为 ToDo 添加评论"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']
        data = request.json or {}
        comment_text = data.get('content')

        try:
            todo, event = manager.add_comment_legacy(todo_id, comment_text)
        except ValueError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 500

        timeline = manager.get_timeline()
        return jsonify({
            'success': True,
            'todo': todo,
            'event': event,
            'timeline': timeline,
            'projects': manager.get_projects()
        })
    
    @app.route('/api/todo/events/<event_id>', methods=['DELETE'])
    def todo_event(event_id):
        """删除单个时间轴事件"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']
        try:
            result = manager.delete_event(event_id)
        except ValueError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 404

        timeline = manager.get_timeline()
        return jsonify({
            'success': True,
            'event_id': event_id,
            'event_type': result.get('event_type'),
            'todo': result.get('todo'),
            'timeline': timeline,
            'projects': manager.get_projects()
        })

    # ===== 新版ToDo API（基于项目和任务的层级结构） =====

    @app.route('/api/todo/v2/data', methods=['GET'])
    def todo_v2_data():
        """获取所有项目和任务数据"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']
        try:
            data = manager.list_all()
            overview = manager.get_pending_overview()
            return jsonify({
                'success': True,
                'data': data,
                'overview': overview
            })
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 500

    @app.route('/api/todo/v2/projects', methods=['POST'])
    def todo_v2_create_project():
        """创建新项目"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']
        payload = request.json or {}
        try:
            project = manager.create_project(payload)
            _todo_kb_rebuild_all_async()
            return jsonify({'success': True, 'project': project})
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

    @app.route('/api/todo/v2/projects/<project_id>', methods=['GET', 'PUT', 'DELETE'])
    def todo_v2_project(project_id):
        """项目操作：查询、更新、删除"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']

        if request.method == 'GET':
            try:
                project = manager.get_project(project_id)
                return jsonify({'success': True, 'project': project})
            except ValueError:
                return jsonify({'success': False, 'error': '项目不存在'}), 404

        if request.method == 'PUT':
            payload = request.json or {}
            try:
                project = manager.update_project(project_id, payload)
                _todo_kb_rebuild_all_async()
                return jsonify({'success': True, 'project': project})
            except ValueError:
                return jsonify({'success': False, 'error': '项目不存在'}), 404
            except Exception as exc:
                return jsonify({'success': False, 'error': str(exc)}), 400

        # DELETE
        try:
            project = manager.delete_project(project_id)
            _todo_kb_rebuild_all_async()
            return jsonify({'success': True, 'project': project})
        except ValueError:
            return jsonify({'success': False, 'error': '项目不存在'}), 404

    @app.route('/api/todo/v2/projects/<project_id>/tasks', methods=['POST'])
    def todo_v2_create_task(project_id):
        """在项目中创建新任务"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']
        payload = request.json or {}
        try:
            task, update_record = manager.create_task(project_id, payload)
            _todo_kb_rebuild_all_async()
            return jsonify({'success': True, 'task': task, 'update_record': update_record})
        except ValueError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 404
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

    @app.route('/api/todo/v2/projects/<project_id>/tasks/<task_id>', methods=['GET', 'PUT', 'DELETE'])
    def todo_v2_task(project_id, task_id):
        """任务操作：查询、更新、删除"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']

        if request.method == 'GET':
            try:
                project = manager.get_project(project_id)
                task = next((t for t in project.get('tasks', []) if t.get('id') == task_id), None)
                if not task:
                    return jsonify({'success': False, 'error': '任务不存在'}), 404
                return jsonify({'success': True, 'task': task})
            except ValueError:
                return jsonify({'success': False, 'error': '项目不存在'}), 404

        if request.method == 'PUT':
            payload = request.json or {}
            try:
                task, update_records = manager.update_task(project_id, task_id, payload)
                _todo_kb_rebuild_all_async()
                return jsonify({'success': True, 'task': task, 'update_records': update_records})
            except ValueError as exc:
                return jsonify({'success': False, 'error': str(exc)}), 404
            except Exception as exc:
                return jsonify({'success': False, 'error': str(exc)}), 400

        # DELETE
        try:
            task = manager.delete_task(project_id, task_id)
            _todo_kb_rebuild_all_async()
            return jsonify({'success': True, 'task': task})
        except ValueError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 404

    @app.route('/api/todo/v2/projects/<project_id>/tasks/reorder', methods=['POST'])
    def todo_v2_reorder_tasks(project_id):
        """重新排序任务"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']
        payload = request.json or {}
        task_ids = payload.get('task_ids', [])

        if not isinstance(task_ids, list):
            return jsonify({'success': False, 'error': 'task_ids必须是数组'}), 400

        try:
            project = manager.reorder_tasks(project_id, task_ids)
            return jsonify({'success': True, 'project': project})
        except ValueError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 404

    @app.route('/api/todo/v2/projects/<project_id>/tasks/<task_id>/comments', methods=['POST'])
    def todo_v2_add_comment(project_id, task_id):
        """为任务添加评论"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']
        payload = request.json or {}
        comment_text = payload.get('content', '')

        try:
            task, update_record = manager.add_comment(project_id, task_id, comment_text)
            _todo_kb_rebuild_all_async()
            return jsonify({'success': True, 'task': task, 'update_record': update_record})
        except ValueError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 500

    @app.route('/api/todo/v2/projects/<project_id>/tasks/<task_id>/comments/<comment_id>', methods=['DELETE'])
    def todo_v2_delete_comment(project_id, task_id, comment_id):
        """删除评论"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']
        try:
            task = manager.delete_comment(project_id, task_id, comment_id)
            _todo_kb_rebuild_all_async()
            return jsonify({'success': True, 'task': task})
        except ValueError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 404

    @app.route('/api/todo/v2/overview', methods=['GET'])
    def todo_v2_overview():
        """获取待完成概览"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']
        try:
            overview = manager.get_pending_overview()
            return jsonify({'success': True, 'overview': overview})
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 500

    # ===== ToDo扩展功能API（项目描述和链接） =====

    @app.route('/api/todo/v2/projects/<project_id>/description', methods=['GET', 'PUT', 'DELETE'])
    def todo_v2_project_description(project_id):
        """项目描述操作：获取、更新、删除"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        extended_manager = current_app.config['TODO_EXTENDED_MANAGER']

        if request.method == 'GET':
            description = extended_manager.get_project_description(project_id)
            return jsonify({'success': True, 'description': description})

        if request.method == 'PUT':
            payload = request.json or {}
            description_md = payload.get('description', '')
            extended_manager.set_project_description(project_id, description_md)
            _todo_kb_rebuild_all_async()
            return jsonify({'success': True, 'description': description_md})

        # DELETE
        extended_manager.delete_project_description(project_id)
        _todo_kb_rebuild_all_async()
        return jsonify({'success': True})

    @app.route('/api/todo/v2/projects/<project_id>/description/preview', methods=['POST'])
    def todo_v2_project_description_preview(project_id):
        """预览项目描述的Markdown渲染结果"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        payload = request.json or {}
        description_md = payload.get('description', '')
        
        try:
            # 使用现有的markdown渲染功能
            html_content = markdown_parser.render(description_md)
            return jsonify({'success': True, 'html': html_content})
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 500

    @app.route('/api/todo/v2/projects/<project_id>/links', methods=['GET', 'POST'])
    def todo_v2_project_links(project_id):
        """项目链接操作：获取、添加"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        extended_manager = current_app.config['TODO_EXTENDED_MANAGER']

        if request.method == 'GET':
            links = extended_manager.get_project_links(project_id)
            return jsonify({'success': True, 'links': links})

        # POST - 添加链接
        payload = request.json or {}
        name = payload.get('name', '').strip()
        url = payload.get('url', '').strip()
        
        if not name or not url:
            return jsonify({'success': False, 'error': '链接名称和地址不能为空'}), 400

        link = extended_manager.add_project_link(project_id, name, url)
        _todo_kb_rebuild_all_async()
        return jsonify({'success': True, 'link': link})

    @app.route('/api/todo/v2/projects/<project_id>/links/<link_id>', methods=['PUT', 'DELETE'])
    def todo_v2_project_link(project_id, link_id):
        """项目链接操作：更新、删除"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        extended_manager = current_app.config['TODO_EXTENDED_MANAGER']

        if request.method == 'PUT':
            payload = request.json or {}
            name = payload.get('name', '').strip()
            url = payload.get('url', '').strip()
            
            if not name or not url:
                return jsonify({'success': False, 'error': '链接名称和地址不能为空'}), 400

            link = extended_manager.update_project_link(project_id, link_id, name, url)
            if link:
                _todo_kb_rebuild_all_async()
                return jsonify({'success': True, 'link': link})
            return jsonify({'success': False, 'error': '链接不存在'}), 404

        # DELETE
        success = extended_manager.delete_project_link(project_id, link_id)
        if success:
            _todo_kb_rebuild_all_async()
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': '链接不存在'}), 404

    @app.route('/api/todo/v2/tasks/<task_id>/links', methods=['GET', 'POST'])
    def todo_v2_task_links(task_id):
        """任务链接操作：获取、添加"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        extended_manager = current_app.config['TODO_EXTENDED_MANAGER']

        if request.method == 'GET':
            links = extended_manager.get_task_links(task_id)
            return jsonify({'success': True, 'links': links})

        # POST - 添加链接
        payload = request.json or {}
        name = payload.get('name', '').strip()
        url = payload.get('url', '').strip()
        
        if not name or not url:
            return jsonify({'success': False, 'error': '链接名称和地址不能为空'}), 400

        link = extended_manager.add_task_link(task_id, name, url)
        _todo_kb_rebuild_all_async()
        return jsonify({'success': True, 'link': link})

    @app.route('/api/todo/v2/tasks/<task_id>/links/<link_id>', methods=['PUT', 'DELETE'])
    def todo_v2_task_link(task_id, link_id):
        """任务链接操作：更新、删除"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        extended_manager = current_app.config['TODO_EXTENDED_MANAGER']

        if request.method == 'PUT':
            payload = request.json or {}
            name = payload.get('name', '').strip()
            url = payload.get('url', '').strip()
            
            if not name or not url:
                return jsonify({'success': False, 'error': '链接名称和地址不能为空'}), 400

            link = extended_manager.update_task_link(task_id, link_id, name, url)
            if link:
                _todo_kb_rebuild_all_async()
                return jsonify({'success': True, 'link': link})
            return jsonify({'success': False, 'error': '链接不存在'}), 404

        # DELETE
        success = extended_manager.delete_task_link(task_id, link_id)
        if success:
            _todo_kb_rebuild_all_async()
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': '链接不存在'}), 404

    @app.route('/api/todo/v2/meeting_notes', methods=['GET', 'POST'])
    def todo_v2_meeting_notes():
        """会议记录操作：获取列表、创建/更新"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        extended_manager = current_app.config['TODO_EXTENDED_MANAGER']

        if request.method == 'GET':
            # 获取所有笔记列表
            notes = extended_manager.list_meeting_notes()
            return jsonify({'success': True, 'notes': notes})

        # POST - 创建或更新笔记
        payload = request.json or {}
        date_str = payload.get('date', '').strip()
        content = payload.get('content', '')

        if not date_str:
            return jsonify({'success': False, 'error': '日期不能为空'}), 400

        extended_manager.set_meeting_note(date_str, content)
        _todo_kb_rebuild_all_async()
        return jsonify({'success': True})

    @app.route('/api/todo/v2/meeting_notes/<date_str>', methods=['GET', 'PUT', 'DELETE'])
    def todo_v2_meeting_note(date_str):
        """会议记录操作：获取、更新、删除指定日期的笔记"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        extended_manager = current_app.config['TODO_EXTENDED_MANAGER']

        if request.method == 'GET':
            content = extended_manager.get_meeting_note(date_str)
            if content is not None:
                return jsonify({'success': True, 'content': content})
            return jsonify({'success': False, 'error': '笔记不存在'}), 404

        if request.method == 'PUT':
            payload = request.json or {}
            content = payload.get('content', '')
            extended_manager.set_meeting_note(date_str, content)
            _todo_kb_rebuild_all_async()
            return jsonify({'success': True})

        # DELETE
        success = extended_manager.delete_meeting_note(date_str)
        if success:
            _todo_kb_rebuild_all_async()
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': '笔记不存在'}), 404

    @app.route('/api/todo/v2/meeting_notes/<date_str>/preview', methods=['POST'])
    def todo_v2_meeting_note_preview(date_str):
        """预览会议记录的Markdown渲染结果"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        payload = request.json or {}
        content = payload.get('content', '')
        
        try:
            html_content = markdown_parser.render(content)
            return jsonify({'success': True, 'html': html_content})
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 500

    @app.route('/api/todo/v2/reports/<report_type>', methods=['GET'])
    def todo_v2_report_list(report_type):
        """获取日报/周报列表"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        extended_manager = current_app.config['TODO_EXTENDED_MANAGER']
        try:
            items = extended_manager.list_reports(report_type)
            return jsonify({'success': True, 'items': items})
        except ValueError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

    @app.route('/api/todo/v2/reports/<report_type>/generate', methods=['POST'])
    def todo_v2_report_generate(report_type):
        """生成并可选保存日报/周报"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']
        extended_manager = current_app.config['TODO_EXTENDED_MANAGER']
        payload = request.json or {}
        save = bool(payload.get('save', True))

        try:
            if report_type == 'daily':
                report = todo_report_builder.build_daily_report(
                    manager,
                    extended_manager,
                    date_str=(payload.get('date') or '').strip() or None,
                )
            elif report_type == 'weekly':
                report = todo_report_builder.build_weekly_report(
                    manager,
                    extended_manager,
                    week_key=(payload.get('week_key') or '').strip() or None,
                    ref_date=(payload.get('date') or '').strip() or None,
                )
            else:
                return jsonify({'success': False, 'error': '不支持的 report_type'}), 400
        except ValueError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

        record = report.to_record()
        if save:
            record = extended_manager.set_report(report_type, report.key, record)
            _todo_kb_rebuild_all_async()
        return jsonify({'success': True, 'report': record, 'saved': save})

    @app.route('/api/todo/v2/reports/<report_type>/<report_key>', methods=['GET', 'PUT', 'DELETE'])
    def todo_v2_report_detail(report_type, report_key):
        """获取、更新、删除日报/周报"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        extended_manager = current_app.config['TODO_EXTENDED_MANAGER']

        if request.method == 'GET':
            try:
                item = extended_manager.get_report(report_type, report_key)
            except ValueError as exc:
                return jsonify({'success': False, 'error': str(exc)}), 400
            if not item:
                return jsonify({'success': False, 'error': '报表不存在'}), 404
            return jsonify({'success': True, 'report': item})

        if request.method == 'PUT':
            payload = request.json or {}
            try:
                existing = extended_manager.get_report(report_type, report_key)
                if not existing:
                    return jsonify({'success': False, 'error': '报表不存在'}), 404
                merged = {
                    **existing,
                    'title': payload.get('title', existing.get('title', '')),
                    'content': payload.get('content', existing.get('content', '')),
                    'period_start': payload.get('period_start', existing.get('period_start', '')),
                    'period_end': payload.get('period_end', existing.get('period_end', '')),
                    'source_tasks': payload.get('source_tasks', existing.get('source_tasks', [])),
                    'source_notes': payload.get('source_notes', existing.get('source_notes', [])),
                    'generated_at': payload.get('generated_at', existing.get('generated_at', '')),
                    'created_at': existing.get('created_at', ''),
                }
                item = extended_manager.set_report(report_type, report_key, merged)
                _todo_kb_rebuild_all_async()
                return jsonify({'success': True, 'report': item})
            except ValueError as exc:
                return jsonify({'success': False, 'error': str(exc)}), 400

        try:
            success = extended_manager.delete_report(report_type, report_key)
        except ValueError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400
        if not success:
            return jsonify({'success': False, 'error': '报表不存在'}), 404
        _todo_kb_rebuild_all_async()
        return jsonify({'success': True})

    @app.route('/api/todo/v2/file_tree', methods=['GET'])
    def todo_v2_file_tree():
        """获取文件浏览器文件树（用于文件选择器）- 仅加载第一层，支持懒加载"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        root_dir = current_app.config.get('ROOT_DIR')
        if not root_dir or not os.path.exists(root_dir) or not os.path.isdir(root_dir):
            return jsonify({'success': False, 'error': '共享文件夹未设置'}), 404

        # 只加载第一层，不递归加载子目录（支持懒加载）
        def build_tree_first_level(dir_path, rel_path=''):
            """只构建第一层文件树"""
            result = []
            try:
                items = os.listdir(dir_path)
                items.sort(key=lambda x: (not os.path.isdir(os.path.join(dir_path, x)), x.lower()))
                
                for item in items:
                    try:
                        item_path = os.path.join(dir_path, item)
                        item_rel_path = posixpath.join(rel_path, item) if rel_path else item
                        
                        node = {
                            'name': item,
                            'path': item_rel_path,
                            'is_dir': os.path.isdir(item_path)
                        }
                        
                        # 如果是目录，标记有子节点，但不加载内容（懒加载）
                        if node['is_dir']:
                            # 检查是否有子项
                            try:
                                sub_items = os.listdir(item_path)
                                node['has_children'] = len(sub_items) > 0
                                node['children'] = []  # 初始为空，点击展开时再加载
                            except:
                                node['has_children'] = False
                                node['children'] = []
                        
                        result.append(node)
                    except Exception as e:
                        _routes_logger.warning("Error processing item %s in %s: %s", item, dir_path, e)
                        continue
            except Exception as e:
                _routes_logger.warning("Error building tree for %s: %s", dir_path, e)
            
            return result

        try:
            tree = build_tree_first_level(root_dir)
            return jsonify({'success': True, 'tree': tree})
        except Exception as exc:
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': str(exc)}), 500
    
    @app.route('/api/todo/v2/file_tree/<path:folder_path>', methods=['GET'])
    def todo_v2_file_tree_subfolder(folder_path):
        """获取指定文件夹的子文件树（用于懒加载）"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        root_dir = current_app.config.get('ROOT_DIR')
        if not root_dir or not os.path.exists(root_dir) or not os.path.isdir(root_dir):
            return jsonify({'success': False, 'error': '共享文件夹未设置'}), 404

        # 构建完整路径
        folder_full_path = os.path.join(root_dir, folder_path)
        
        # 安全检查：确保路径在root_dir内
        if not os.path.abspath(folder_full_path).startswith(os.path.abspath(root_dir)):
            return jsonify({'success': False, 'error': '路径不安全'}), 400
        
        if not os.path.exists(folder_full_path) or not os.path.isdir(folder_full_path):
            return jsonify({'success': False, 'error': '文件夹不存在'}), 404

        def build_tree_first_level(dir_path, rel_path=''):
            """只构建第一层文件树"""
            result = []
            try:
                items = os.listdir(dir_path)
                items.sort(key=lambda x: (not os.path.isdir(os.path.join(dir_path, x)), x.lower()))
                
                for item in items:
                    try:
                        item_path = os.path.join(dir_path, item)
                        item_rel_path = posixpath.join(rel_path, item) if rel_path else item
                        
                        node = {
                            'name': item,
                            'path': item_rel_path,
                            'is_dir': os.path.isdir(item_path)
                        }
                        
                        if node['is_dir']:
                            try:
                                sub_items = os.listdir(item_path)
                                node['has_children'] = len(sub_items) > 0
                                node['children'] = []
                            except:
                                node['has_children'] = False
                                node['children'] = []
                        
                        result.append(node)
                    except Exception as e:
                        _routes_logger.warning("Error processing item %s in %s: %s", item, dir_path, e)
                        continue
            except Exception as e:
                _routes_logger.warning("Error building tree for %s: %s", dir_path, e)
            
            return result

        try:
            tree = build_tree_first_level(folder_full_path, folder_path)
            return jsonify({'success': True, 'tree': tree})
        except Exception as exc:
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': str(exc)}), 500

    @app.route('/api/todo/v2/export/report', methods=['POST'])
    def todo_v2_export_report():
        """导出汇报表格为Excel文件"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        try:
            from flask import send_file
            from io import BytesIO
            from datetime import datetime
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            payload = request.json or {}
            tasks = payload.get('tasks', [])
            custom_file_name = payload.get('fileName', f'汇报表格_{datetime.now().strftime("%Y-%m-%d")}.xlsx')
            column_widths_config = payload.get('columnWidths', {})  # 前端传递的列宽配置

            if not tasks:
                return jsonify({'success': False, 'error': '汇报数据为空'}), 400

            # 创建Excel工作簿
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            ws = wb.create_sheet('汇报表格', 0)

            # 设置表头
            headers = ['项目名称', '序号', '任务简述', '详细描述', '状态', '当前进展', '本周计划']
            ws.append(headers)
            
            # 启用筛选
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

            # 设置表头样式
            header_fill = PatternFill(start_color='D6F2EA', end_color='C8ECE0', fill_type='solid')
            header_font = Font(bold=True, color='0F433B', size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # 填充数据
            current_project_name = None
            project_name_start_row = None
            project_name_col_idx = 1

            for row_idx, task in enumerate(tasks, start=2):
                project_name = task.get('project_name', '未命名项目')
                project_phase = task.get('project_phase', '')
                if project_phase:
                    project_name_display = f"{project_name}\n({project_phase})"
                else:
                    project_name_display = project_name

                # 处理项目名称合并
                if project_name != current_project_name:
                    if current_project_name is not None and project_name_start_row is not None:
                        # 合并上一个项目的名称单元格
                        if row_idx - 1 > project_name_start_row:
                            ws.merge_cells(
                                start_row=project_name_start_row,
                                start_column=project_name_col_idx,
                                end_row=row_idx - 1,
                                end_column=project_name_col_idx
                            )
                    current_project_name = project_name
                    project_name_start_row = row_idx

                # 构建当前进展内容
                current_progress = task.get('current_progress', '')
                if not current_progress:
                    is_completed = (task.get('progress') or 0) >= 100
                    if is_completed:
                        current_progress = '已完成'
                    else:
                        current_progress = '进行中'

                # 填充行数据
                row_data = [
                    project_name_display if row_idx == project_name_start_row else '',  # 项目名称（只在第一行显示）
                    task.get('index', ''),
                    task.get('summary', ''),
                    task.get('description', ''),
                    task.get('status', ''),
                    current_progress,
                    task.get('weekly_plan', '')
                ]

                # 计算该行需要的最大行数（用于设置行高）
                max_lines = 1
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = border
                    
                    # 所有单元格：垂直居中，自动换行
                    if col_idx == 1:  # 项目名称列
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    elif col_idx == 2:  # 序号列
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    elif col_idx == 5:  # 状态列
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    else:  # 其他列（简述、描述、进展、计划）
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    
                    # 更准确地估算该单元格需要的行数
                    if value:
                        col_letter = get_column_letter(col_idx)
                        col_width = column_widths_config.get(col_letter, 20)  # Excel列宽单位
                        text = str(value)
                        text_length = len(text)
                        
                        # 计算换行符数量
                        newline_count = text.count('\n')
                        
                        # Excel列宽单位转换为字符数：1单位 ≈ 7像素，假设字体大小为11，每个字符约6-7像素宽
                        # 更准确的估算：列宽单位 * 1.2 ≈ 字符数（因为Excel列宽单位比实际字符宽度大）
                        chars_per_line = max(1, int(col_width * 1.2))
                        
                        # 计算需要的行数：考虑换行符和文本长度
                        if newline_count > 0:
                            # 如果有换行符，每行单独计算
                            lines = text.split('\n')
                            for line in lines:
                                line_chars = len(line)
                                line_count = max(1, (line_chars // chars_per_line) + 1)
                                max_lines = max(max_lines, line_count)
                            max_lines += newline_count  # 加上换行符本身
                        else:
                            # 没有换行符，根据文本长度计算
                            estimated_lines = max(1, (text_length // chars_per_line) + 1)
                            max_lines = max(max_lines, estimated_lines)
                
                # 设置行高：基础高度15 + 每行额外15（确保内容完整显示）
                row_height = 15 + (max_lines - 1) * 15
                ws.row_dimensions[row_idx].height = min(row_height, 150)  # 最大150，确保内容完整

            # 合并最后一个项目的名称单元格
            if current_project_name is not None and project_name_start_row is not None:
                if len(tasks) + 1 > project_name_start_row:
                    ws.merge_cells(
                        start_row=project_name_start_row,
                        start_column=project_name_col_idx,
                        end_row=len(tasks) + 1,
                        end_column=project_name_col_idx
                    )

            # 设置列宽（使用前端传递的列宽，如果没有则根据内容计算）
            default_column_widths = {
                'A': 20,  # 项目名称
                'B': 8,   # 序号
                'C': 25,  # 任务简述
                'D': 40,  # 详细描述
                'E': 12,  # 状态
                'F': 30,  # 当前进展
                'G': 30   # 本周计划
            }
            
            # 如果前端没有传递列宽，根据内容计算合适的列宽
            if not column_widths_config:
                # 计算每列的最大内容长度
                max_lengths = {}
                for task in tasks:
                    values = [
                        task.get('project_name', '') + ('\n' + task.get('project_phase', '') if task.get('project_phase') else ''),
                        str(task.get('index', '')),
                        task.get('summary', ''),
                        task.get('description', ''),
                        task.get('status', ''),
                        task.get('current_progress', ''),
                        task.get('weekly_plan', '')
                    ]
                    for idx, val in enumerate(values):
                        col_letter = get_column_letter(idx + 1)
                        length = len(str(val))
                        if col_letter not in max_lengths or length > max_lengths[col_letter]:
                            max_lengths[col_letter] = length
                
                # 根据内容长度设置列宽（稍微宽一些）
                for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                    if col_letter in max_lengths:
                        # 列宽 = 内容长度 * 1.2 + 2（稍微宽一些）
                        content_width = max_lengths[col_letter] * 1.2 + 2
                        default_width = default_column_widths.get(col_letter, 20)
                        column_widths_config[col_letter] = max(default_width, min(content_width, 60))  # 最大60
            
            for col_letter, default_width in default_column_widths.items():
                # 优先使用前端传递的列宽，否则使用计算或默认值
                width = column_widths_config.get(col_letter, default_width)
                ws.column_dimensions[col_letter].width = width

            # 设置表头行高
            ws.row_dimensions[1].height = 25
            _excel_apply_auto_dimensions(ws, min_width=8, max_width=60, header_row=1)

            # 保存到内存
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=custom_file_name
            )

        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': str(e)}), 500

    @app.route('/api/todo/v2/export/excel', methods=['POST'])
    def todo_v2_export_excel():
        """导出TODO数据为Excel文件"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        try:
            from flask import send_file
            from io import BytesIO
            from datetime import datetime
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            config = request.json or {}
            manager = current_app.config['TODO_MANAGER']
            
            # 获取自定义文件名
            custom_file_name = config.get('fileName')
            
            # 检查是否是预览模式
            time_range = config.get('timeRange', 'all')
            if time_range == 'preview':
                # 预览模式：直接使用传入的预览数据
                preview_data = config.get('previewData', [])
                if not preview_data:
                    return jsonify({'success': False, 'error': '预览数据为空'}), 400
                all_tasks = preview_data
            else:
                # 正常模式：从数据库获取数据
                data = manager.list_all()
                
                # 获取所有任务并扁平化
                all_tasks = []
                for project in data.get('projects', []):
                    for task in project.get('tasks', []):
                        task_with_project = {
                            **task,
                            'project_id': project['id'],
                            'project_name': project['name'],
                            'project_color': project.get('color', '#4facfe'),
                        }
                        all_tasks.append(task_with_project)

                # 应用时间范围筛选
            if time_range == 'completed':
                all_tasks = [t for t in all_tasks if (t.get('progress') or 0) >= 100]
            elif time_range == 'pending':
                all_tasks = [t for t in all_tasks if (t.get('progress') or 0) < 100]
            elif time_range == 'custom':
                custom_range = config.get('customTimeRange')
                if custom_range:
                    start_date = datetime.strptime(custom_range['start'], '%Y-%m-%d')
                    end_date = datetime.strptime(custom_range['end'], '%Y-%m-%d')
                    field = custom_range.get('field', 'created')
                    
                    filtered_tasks = []
                    for task in all_tasks:
                        date_str = None
                        if field == 'created':
                            date_str = task.get('created_at')
                        elif field == 'updated':
                            date_str = task.get('updated_at')
                        elif field == 'due':
                            date_str = task.get('due_date')
                        
                        if date_str:
                            try:
                                task_date = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                                if start_date <= task_date <= end_date:
                                    filtered_tasks.append(task)
                            except:
                                pass
                    all_tasks = filtered_tasks

            # 获取列配置
            columns = config.get('columns', [])
            column_order = config.get('columnOrder', columns)
            project_mode = config.get('projectMode', 'single')
            comment_mode = config.get('commentMode', 'inline')
            merge_project_name = config.get('mergeProjectName', False)
            add_global_index = config.get('addGlobalIndex', False)

            # 列定义映射
            column_definitions = {
                'project_name': '项目名称',
                'index': '序号',
                'summary': '任务简述',
                'description': '详细描述',
                'priority': '优先级',
                'progress': '进度',
                'due_date': '预计完成时间',
                'created_at': '创建时间',
                'updated_at': '最后更新时间',
                'change_count': '改动数量',
            }

            # 创建Excel工作簿
            wb = openpyxl.Workbook()
            
            # 删除默认sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            if project_mode == 'single':
                # 方式一：同一Sheet
                ws = wb.create_sheet('任务列表', 0)
                
                # 设置表头
                headers = []
                header_col_map = {}  # 记录表头列索引到column_order的映射
                col_idx = 0
                
                # 如果启用总序号，在第一列添加总序号
                if add_global_index:
                    headers.append('总序号')
                    header_col_map[col_idx] = 'global_index'
                    col_idx += 1
                
                if 'project_name' in column_order:
                    headers.append('项目名称')
                    header_col_map[col_idx] = 'project_name'
                    col_idx += 1
                
                for col in column_order:
                    if col == 'project_name':
                        continue
                    if col in column_definitions:
                        # 如果是评论数且使用内联模式，表头显示"评论"而不是"评论数"
                        if col == 'change_count' and comment_mode == 'inline':
                            headers.append('评论')
                        else:
                            headers.append(column_definitions[col])
                        header_col_map[col_idx] = col
                        col_idx += 1
                
                # 如果使用sheet模式且没有change_count列，添加评论列用于超链接
                if comment_mode == 'sheet' and 'change_count' not in column_order:
                    headers.append('评论')
                    header_col_map[col_idx] = 'comment'
                    col_idx += 1
                elif comment_mode == 'inline' and 'change_count' not in column_order:
                    headers.append('评论')
                    header_col_map[col_idx] = 'comment'
                    col_idx += 1
                
                ws.append(headers)
                
                # 设置表头样式
                header_fill = PatternFill(start_color='D6F2EA', end_color='C8ECE0', fill_type='solid')
                header_font = Font(bold=True, color='1E5F4F', size=11)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border

                # 添加数据行
                row_num = 2
                comment_sheet_rows = []  # 用于方式二的评论数据
                comment_row_counter = 2  # 评论Sheet的行号计数器（从2开始，因为第1行是表头）
                
                # 定义交替行颜色
                even_row_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                odd_row_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                cell_border = Border(
                    left=Side(style='thin', color='E0E0E0'),
                    right=Side(style='thin', color='E0E0E0'),
                    top=Side(style='thin', color='E0E0E0'),
                    bottom=Side(style='thin', color='E0E0E0')
                )
                
                # 用于记录项目名称合并信息
                project_name_merge_ranges = []  # 存储需要合并的范围 [(start_row, end_row, col_idx)]
                current_project_name = None
                project_name_start_row = None
                project_name_col_idx = None
                
                # 找到项目名称列的索引
                if 'project_name' in column_order:
                    for header_idx, col_key in header_col_map.items():
                        if col_key == 'project_name':
                            project_name_col_idx = header_idx + 1  # Excel列索引从1开始
                            break
                
                # 计算每个任务在其项目中的序号
                project_task_counters = {}  # {project_id: counter}
                
                for global_idx, task in enumerate(all_tasks, 1):
                    row_data = []
                    comments = task.get('comments', [])
                    is_even = (row_num % 2 == 0)
                    row_fill = even_row_fill if is_even else odd_row_fill
                    
                    task_project_name = task.get('project_name', '--')
                    
                    # 处理项目名称合并逻辑
                    if merge_project_name and project_name_col_idx:
                        if task_project_name != current_project_name:
                            # 项目名称变化，如果之前有连续的项目名称，记录合并范围
                            if current_project_name is not None and project_name_start_row is not None:
                                if row_num - 1 > project_name_start_row:
                                    # 有多个连续行，需要合并
                                    project_name_merge_ranges.append((project_name_start_row, row_num - 1, project_name_col_idx))
                                # 重置
                            current_project_name = task_project_name
                            project_name_start_row = row_num
                    
                    # 按照表头顺序构建数据行
                    for header_idx in range(len(headers)):
                        col_key = header_col_map.get(header_idx)
                        
                        if col_key == 'global_index':
                            row_data.append(global_idx)
                        elif col_key == 'project_name':
                            row_data.append(task_project_name)
                        elif col_key == 'index':
                            # 使用任务在其项目中的序号
                            project_id = task.get('project_id', '')
                            if project_id not in project_task_counters:
                                project_task_counters[project_id] = 0
                            project_task_counters[project_id] += 1
                            row_data.append(project_task_counters[project_id])
                        elif col_key == 'summary':
                            row_data.append(task.get('summary', '--'))
                        elif col_key == 'description':
                            row_data.append(task.get('description', '--'))
                        elif col_key == 'priority':
                            row_data.append(task.get('priority', 3))
                        elif col_key == 'progress':
                            row_data.append(f"{task.get('progress', 0)}%")
                        elif col_key == 'due_date':
                            due_date = task.get('due_date')
                            row_data.append(due_date if due_date else '--')
                        elif col_key == 'created_at':
                            created = task.get('created_at', '')
                            if created:
                                try:
                                    dt = datetime.fromisoformat(created.replace('Z', '+00:00'))
                                    row_data.append(dt.strftime('%Y-%m-%d %H:%M'))
                                except:
                                    row_data.append(created)
                            else:
                                row_data.append('--')
                        elif col_key == 'updated_at':
                            updated = task.get('updated_at', '')
                            if updated:
                                try:
                                    dt = datetime.fromisoformat(updated.replace('Z', '+00:00'))
                                    row_data.append(dt.strftime('%Y-%m-%d %H:%M'))
                                except:
                                    row_data.append(updated)
                            else:
                                row_data.append('--')
                        elif col_key == 'change_count':
                            # 改动数量：显示修改次数和评论数
                            update_history = task.get('update_history', [])
                            update_count = len([r for r in update_history if r.get('field') != 'comment'])
                            comment_count = len(comments)
                            row_data.append(f"{update_count}次修改 {comment_count}条评论")
                        elif col_key == 'comment':
                            # 额外的评论列
                            if comment_mode == 'inline':
                                # 内联模式：显示评论内容
                                if comments:
                                    comment_text = '\n'.join([
                                        f"[{datetime.fromisoformat(c.get('timestamp', '').replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M')}] {c.get('content', '')}"
                                        for c in comments
                                    ])
                                    row_data.append(comment_text)
                                else:
                                    row_data.append('--')
                            else:
                                # sheet模式：显示超链接文本（超链接会在后面添加）
                                if comments:
                                    row_data.append(f"查看评论({len(comments)})")
                                else:
                                    row_data.append('--')
                        else:
                            row_data.append('--')
                    
                    # 收集评论数据（方式二）
                    if comment_mode == 'sheet':
                        for comment in comments:
                            comment_sheet_rows.append({
                                'project_name': task.get('project_name', '--'),
                                'task_summary': task.get('summary', '--'),
                                'timestamp': comment.get('timestamp', ''),
                                'content': comment.get('content', ''),
                                'task_id': task.get('id', ''),
                                'row_num': comment_row_counter,
                            })
                            comment_row_counter += 1
                    
                    ws.append(row_data)
                    
                    # 设置行样式（交替行颜色和边框）
                    for col_idx in range(1, len(headers) + 1):
                        cell = ws.cell(row=row_num, column=col_idx)
                        cell.fill = row_fill
                        cell.border = cell_border
                        
                        # 设置对齐方式
                        col_key = header_col_map.get(col_idx - 1)
                        if col_key == 'description' or col_key == 'summary':
                            # 任务描述和任务简述：左对齐，上下居中，自动换行
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        elif col_key == 'comment' or col_key == 'change_count':
                            # 评论列：左对齐，上下居中
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        else:
                            # 其他列：居中，上下居中
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 设置评论列自动换行（方式一）
                    if comment_mode == 'inline' and comments:
                        # 找到评论列的索引
                        comment_col_idx = None
                        for header_idx, col_key in header_col_map.items():
                            if col_key == 'change_count' or col_key == 'comment':
                                comment_col_idx = header_idx + 1
                                break
                        if not comment_col_idx and 'change_count' not in column_order:
                            comment_col_idx = len(headers)  # 评论列在最后
                        if comment_col_idx:
                            comment_cell = ws.cell(row=row_num, column=comment_col_idx)
                            # 评论列：左对齐，上下居中，自动换行
                            comment_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    
                    # 添加超链接（方式二）
                    if comment_mode == 'sheet' and comments:
                        # 找到评论列的索引（可能是change_count列或comment列）
                        comment_col_idx = None
                        for header_idx, col_key in header_col_map.items():
                            if col_key == 'change_count' or col_key == 'comment':
                                comment_col_idx = header_idx + 1
                                break
                        if comment_col_idx:
                            comment_cell = ws.cell(row=row_num, column=comment_col_idx)
                            # 计算评论Sheet中第一条评论的行号（表头占第1行，所以从第2行开始）
                            first_comment_row = comment_row_counter - len(comments)
                            if first_comment_row < 2:
                                first_comment_row = 2
                            comment_cell.hyperlink = f"#评论!A{first_comment_row}"
                            comment_cell.font = Font(color='0000FF', underline='single')
                            # 如果单元格已经有值（从row_data添加的），就不需要再设置value
                            if not comment_cell.value:
                                comment_cell.value = f"查看评论({len(comments)})"
                    
                    row_num += 1
                
                # 处理最后一个项目名称的合并
                if merge_project_name and project_name_col_idx and current_project_name is not None and project_name_start_row is not None:
                    if row_num - 1 > project_name_start_row:
                        # 有多个连续行，需要合并
                        project_name_merge_ranges.append((project_name_start_row, row_num - 1, project_name_col_idx))
                
                # 执行项目名称合并
                if merge_project_name and project_name_merge_ranges:
                    for start_row, end_row, col_idx in project_name_merge_ranges:
                        if end_row > start_row:
                            # 合并单元格
                            col_letter = get_column_letter(col_idx)
                            ws.merge_cells(f'{col_letter}{start_row}:{col_letter}{end_row}')
                            # 设置合并后的单元格对齐方式为居中
                            merged_cell = ws[f'{col_letter}{start_row}']
                            merged_cell.alignment = Alignment(horizontal='center', vertical='center')

                # 根据内容自动调整列宽
                def calculate_column_width(ws, col_idx, min_width=10, max_width=80):
                    """根据列内容计算合适的列宽"""
                    col_letter = get_column_letter(col_idx)
                    max_length = 0
                    
                    # 检查表头
                    header_cell = ws.cell(row=1, column=col_idx)
                    if header_cell.value:
                        max_length = max(max_length, len(str(header_cell.value)))
                    
                    # 检查所有数据行
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            if cell.value:
                                # 对于可能换行的列，计算单行最大长度
                                cell_value = str(cell.value)
                                # 如果包含换行符，取最长的一行
                                if '\n' in cell_value:
                                    lines = cell_value.split('\n')
                                    max_line_length = max(len(line) for line in lines)
                                    max_length = max(max_length, max_line_length)
                                else:
                                    max_length = max(max_length, len(cell_value))
                    
                    # 设置列宽，添加一些边距
                    width = min(max(max_length + 2, min_width), max_width)
                    return width
                
                # 调整列宽
                for col_idx, header in enumerate(headers, 1):
                    col_letter = get_column_letter(col_idx)
                    col_key = header_col_map.get(col_idx - 1)
                    
                    # 对于需要换行的列（任务简述、详细描述、评论），设置固定宽度
                    if col_key == 'summary':
                        ws.column_dimensions[col_letter].width = 30  # 任务简述
                    elif col_key == 'description':
                        ws.column_dimensions[col_letter].width = 50  # 详细描述
                    elif col_key == 'comment' or (col_key == 'change_count' and comment_mode == 'inline'):
                        ws.column_dimensions[col_letter].width = 50  # 评论
                    else:
                        # 其他列根据内容自动调整
                        width = calculate_column_width(ws, col_idx, min_width=10, max_width=30)
                        ws.column_dimensions[col_letter].width = width

                # 创建评论Sheet（方式二）
                if comment_mode == 'sheet' and comment_sheet_rows:
                    comment_ws = wb.create_sheet('评论', 1)
                    comment_ws.append(['项目名称', '任务简述', '时间', '评论内容'])
                    
                    # 设置评论表头样式
                    for cell in comment_ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border
                    
                    # 添加评论数据行并设置对齐方式
                    for row_data in comment_sheet_rows:
                        timestamp = row_data['timestamp']
                        if timestamp:
                            try:
                                dt = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
                                timestamp = dt.strftime('%Y-%m-%d %H:%M')
                            except:
                                pass
                        row_num = comment_ws.max_row + 1
                        comment_ws.append([
                            row_data['project_name'],
                            row_data['task_summary'],
                            timestamp,
                            row_data['content'],
                        ])
                        
                        # 设置对齐方式：项目名称和时间居中，任务简述和评论内容左对齐
                        comment_ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center', vertical='center')  # 项目名称
                        comment_ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)  # 任务简述
                        comment_ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center', vertical='center')  # 时间
                        comment_ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)  # 评论内容
                    
                    # 调整评论Sheet列宽
                    comment_ws.column_dimensions['A'].width = 20
                    comment_ws.column_dimensions['B'].width = 30
                    comment_ws.column_dimensions['C'].width = 18
                    comment_ws.column_dimensions['D'].width = 50

            else:
                # 方式二：不同项目不同Sheet
                projects_dict = {}
                for task in all_tasks:
                    project_id = task['project_id']
                    if project_id not in projects_dict:
                        projects_dict[project_id] = {
                            'name': task['project_name'],
                            'tasks': []
                        }
                    projects_dict[project_id]['tasks'].append(task)

                comment_sheet_rows = []
                comment_row_counter = 2  # 评论Sheet的行号计数器
                global_task_counter = 1  # 全局任务计数器（用于总序号）
                
                for project_id, project_data in projects_dict.items():
                    project_name = project_data['name']
                    # Excel sheet名称不能超过31个字符，且不能包含某些特殊字符
                    sheet_name = project_name[:31].replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_')
                    
                    ws = wb.create_sheet(sheet_name, len(wb.sheetnames))
                    
                    # 设置表头
                    headers = []
                    header_col_map = {}  # 记录表头列索引到column_order的映射
                    col_idx = 0
                    
                    # 如果启用总序号，在第一列添加总序号
                    if add_global_index:
                        headers.append('总序号')
                        header_col_map[col_idx] = 'global_index'
                        col_idx += 1
                    
                    for col in column_order:
                        if col in column_definitions:
                            # 如果是评论数且使用内联模式，表头显示"评论"而不是"评论数"
                            if col == 'comment_count' and comment_mode == 'inline':
                                headers.append('评论')
                            else:
                                headers.append(column_definitions[col])
                            header_col_map[col_idx] = col
                            col_idx += 1
                    
                    # 如果使用sheet模式且没有change_count列，添加评论列用于超链接
                    if comment_mode == 'sheet' and 'change_count' not in column_order:
                        headers.append('评论')
                        header_col_map[col_idx] = 'comment'
                        col_idx += 1
                    elif comment_mode == 'inline' and 'change_count' not in column_order:
                        headers.append('评论')
                        header_col_map[col_idx] = 'comment'
                        col_idx += 1
                    
                    ws.append(headers)
                    
                    # 设置表头样式
                    header_fill = PatternFill(start_color='D6F2EA', end_color='C8ECE0', fill_type='solid')
                    header_font = Font(bold=True, color='1E5F4F', size=11)
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border

                    # 添加数据行
                    row_num = 2
                    even_row_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                    odd_row_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                    cell_border = Border(
                        left=Side(style='thin', color='E0E0E0'),
                        right=Side(style='thin', color='E0E0E0'),
                        top=Side(style='thin', color='E0E0E0'),
                        bottom=Side(style='thin', color='E0E0E0')
                    )
                    
                    for idx, task in enumerate(project_data['tasks'], 1):
                        row_data = []
                        comments = task.get('comments', [])
                        is_even = (row_num % 2 == 0)
                        row_fill = even_row_fill if is_even else odd_row_fill
                        
                        # 按照表头顺序构建数据行
                        for header_idx in range(len(headers)):
                            col_key = header_col_map.get(header_idx)
                            
                            if col_key == 'global_index':
                                row_data.append(global_task_counter)
                                global_task_counter += 1
                            elif col_key == 'index':
                                row_data.append(idx)
                            elif col_key == 'summary':
                                row_data.append(task.get('summary', '--'))
                            elif col_key == 'description':
                                row_data.append(task.get('description', '--'))
                            elif col_key == 'priority':
                                row_data.append(task.get('priority', 3))
                            elif col_key == 'progress':
                                row_data.append(f"{task.get('progress', 0)}%")
                            elif col_key == 'due_date':
                                due_date = task.get('due_date')
                                row_data.append(due_date if due_date else '--')
                            elif col_key == 'created_at':
                                created = task.get('created_at', '')
                                if created:
                                    try:
                                        dt = datetime.fromisoformat(created.replace('Z', '+00:00'))
                                        row_data.append(dt.strftime('%Y-%m-%d %H:%M'))
                                    except:
                                        row_data.append(created)
                                else:
                                    row_data.append('--')
                            elif col_key == 'updated_at':
                                updated = task.get('updated_at', '')
                                if updated:
                                    try:
                                        dt = datetime.fromisoformat(updated.replace('Z', '+00:00'))
                                        row_data.append(dt.strftime('%Y-%m-%d %H:%M'))
                                    except:
                                        row_data.append(updated)
                                else:
                                    row_data.append('--')
                            elif col_key == 'change_count':
                                # 改动数量：显示修改次数和评论数
                                update_history = task.get('update_history', [])
                                update_count = len([r for r in update_history if r.get('field') != 'comment'])
                                comment_count = len(comments)
                                row_data.append(f"{update_count}次修改 {comment_count}条评论")
                            elif col_key == 'comment':
                                # 额外的评论列（当comment_count不在column_order中时）
                                if comments:
                                    comment_text = '\n'.join([
                                        f"[{datetime.fromisoformat(c.get('timestamp', '').replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M')}] {c.get('content', '')}"
                                        for c in comments
                                    ])
                                    row_data.append(comment_text)
                                else:
                                    row_data.append('--')
                            else:
                                row_data.append('--')
                        
                        # 收集评论数据（方式二）
                        if comment_mode == 'sheet':
                            for comment in comments:
                                comment_sheet_rows.append({
                                    'project_name': project_name,
                                    'task_summary': task.get('summary', '--'),
                                    'timestamp': comment.get('timestamp', ''),
                                    'content': comment.get('content', ''),
                                    'task_id': task.get('id', ''),
                                    'row_num': comment_row_counter,
                                })
                                comment_row_counter += 1
                        
                        ws.append(row_data)
                        
                        # 设置行样式（交替行颜色和边框）
                        for col_idx in range(1, len(headers) + 1):
                            cell = ws.cell(row=row_num, column=col_idx)
                            cell.fill = row_fill
                            cell.border = cell_border
                            
                            # 设置对齐方式
                            col_key = header_col_map.get(col_idx - 1)
                            if col_key == 'description' or col_key == 'summary':
                                # 任务描述和任务简述：左对齐，上下居中，自动换行
                                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                            elif col_key == 'comment' or col_key == 'change_count':
                                # 评论列：左对齐，上下居中
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                            else:
                                # 其他列：居中，上下居中
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # 设置评论列自动换行（方式一）
                        if comment_mode == 'inline' and comments:
                            # 找到评论列的索引
                            comment_col_idx = None
                            for header_idx, col_key in header_col_map.items():
                                if col_key == 'change_count' or col_key == 'comment':
                                    comment_col_idx = header_idx + 1
                                    break
                            if comment_col_idx:
                                comment_cell = ws.cell(row=row_num, column=comment_col_idx)
                                # 评论列：左对齐，上下居中，自动换行
                                comment_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        
                        # 添加超链接（方式二）
                        if comment_mode == 'sheet' and comments:
                            # 找到评论列的索引（可能是change_count列或comment列）
                            comment_col_idx = None
                            for header_idx, col_key in header_col_map.items():
                                if col_key == 'change_count' or col_key == 'comment':
                                    comment_col_idx = header_idx + 1
                                    break
                            if comment_col_idx:
                                comment_cell = ws.cell(row=row_num, column=comment_col_idx)
                                # 计算评论Sheet中第一条评论的行号（表头占第1行，所以从第2行开始）
                                first_comment_row = comment_row_counter - len(comments)
                                if first_comment_row < 2:
                                    first_comment_row = 2
                                comment_cell.hyperlink = f"#评论!A{first_comment_row}"
                                comment_cell.font = Font(color='0000FF', underline='single')
                                # 如果单元格已经有值（从row_data添加的），就不需要再设置value
                                if not comment_cell.value:
                                    comment_cell.value = f"查看评论({len(comments)})"
                        
                        row_num += 1
                    
                    # 方式二中，如果启用合并项目名称且Sheet中有多个任务，合并项目名称列
                    if merge_project_name and 'project_name' in column_order and row_num > 3:
                        # 找到项目名称列的索引
                        project_name_col_idx = None
                        for header_idx, col_key in header_col_map.items():
                            if col_key == 'project_name':
                                project_name_col_idx = header_idx + 1
                                break
                        if project_name_col_idx:
                            # 合并从第2行（数据开始）到最后一行
                            col_letter = get_column_letter(project_name_col_idx)
                            ws.merge_cells(f'{col_letter}2:{col_letter}{row_num - 1}')
                            # 设置合并后的单元格对齐方式为居中
                            merged_cell = ws[f'{col_letter}2']
                            merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 根据内容自动调整列宽（复用方式一的函数）
                    def calculate_column_width(ws, col_idx, min_width=10, max_width=80):
                        """根据列内容计算合适的列宽"""
                        col_letter = get_column_letter(col_idx)
                        max_length = 0
                        
                        # 检查表头
                        header_cell = ws.cell(row=1, column=col_idx)
                        if header_cell.value:
                            max_length = max(max_length, len(str(header_cell.value)))
                        
                        # 检查所有数据行
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                            for cell in row:
                                if cell.value:
                                    # 对于可能换行的列，计算单行最大长度
                                    cell_value = str(cell.value)
                                    # 如果包含换行符，取最长的一行
                                    if '\n' in cell_value:
                                        lines = cell_value.split('\n')
                                        max_line_length = max(len(line) for line in lines)
                                        max_length = max(max_length, max_line_length)
                                    else:
                                        max_length = max(max_length, len(cell_value))
                        
                        # 设置列宽，添加一些边距
                        width = min(max(max_length + 2, min_width), max_width)
                        return width
                    
                    # 调整列宽
                    for col_idx, header in enumerate(headers, 1):
                        col_letter = get_column_letter(col_idx)
                        col_key = header_col_map.get(col_idx - 1)
                        
                        # 对于需要换行的列（任务简述、详细描述、评论），设置固定宽度
                        if col_key == 'summary':
                            ws.column_dimensions[col_letter].width = 30  # 任务简述
                        elif col_key == 'description':
                            ws.column_dimensions[col_letter].width = 50  # 详细描述
                        elif col_key == 'comment' or (col_key == 'change_count' and comment_mode == 'inline'):
                            ws.column_dimensions[col_letter].width = 50  # 评论
                        else:
                            # 其他列根据内容自动调整
                            width = calculate_column_width(ws, col_idx, min_width=10, max_width=30)
                            ws.column_dimensions[col_letter].width = width

                # 创建评论Sheet（方式二）
                if comment_mode == 'sheet' and comment_sheet_rows:
                    comment_ws = wb.create_sheet('评论', len(wb.sheetnames))
                    comment_ws.append(['项目名称', '任务简述', '时间', '评论内容'])
                    
                    # 设置评论表头样式
                    header_fill = PatternFill(start_color='D6F2EA', end_color='C8ECE0', fill_type='solid')
                    header_font = Font(bold=True, color='1E5F4F', size=11)
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    for cell in comment_ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border
                    
                    # 添加评论数据行并设置对齐方式
                    for row_data in comment_sheet_rows:
                        timestamp = row_data['timestamp']
                        if timestamp:
                            try:
                                dt = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
                                timestamp = dt.strftime('%Y-%m-%d %H:%M')
                            except:
                                pass
                        row_num = comment_ws.max_row + 1
                        comment_ws.append([
                            row_data['project_name'],
                            row_data['task_summary'],
                            timestamp,
                            row_data['content'],
                        ])
                        
                        # 设置对齐方式：项目名称和时间居中，任务简述和评论内容左对齐
                        comment_ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center', vertical='center')  # 项目名称
                        comment_ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)  # 任务简述
                        comment_ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center', vertical='center')  # 时间
                        comment_ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)  # 评论内容
                    
                    # 调整评论Sheet列宽
                    comment_ws.column_dimensions['A'].width = 20
                    comment_ws.column_dimensions['B'].width = 30
                    comment_ws.column_dimensions['C'].width = 18
                    comment_ws.column_dimensions['D'].width = 50

            for worksheet in wb.worksheets:
                _excel_apply_auto_dimensions(worksheet, min_width=8, max_width=60, header_row=1)

            # 保存到内存
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # 生成文件名
            if custom_file_name:
                filename = custom_file_name
                if not filename.endswith('.xlsx'):
                    filename += '.xlsx'
            else:
                filename = f"todo_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )

        except ImportError:
            return jsonify({'success': False, 'error': 'openpyxl库未安装，请运行: pip install openpyxl'}), 500
        except Exception as exc:
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': str(exc)}), 500

    # ===== 产品对比路由 =====
    
    @app.route('/product_compare')
    def product_compare_page():
        """产品对比主页面"""
        if not is_logged_in():
            return redirect(url_for('login'))
        return render_template('product_compare.html')

    @app.route('/api/product_compare/files', methods=['GET'])
    def product_compare_files():
        """获取所有产品对比文件列表"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未授权'}), 401
        
        manager = current_app.config['PRODUCT_COMPARE_MANAGER']
        try:
            files = manager.list_files()
            return jsonify({'success': True, 'files': files})
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 500

    @app.route('/api/product_compare/files', methods=['POST'])
    def product_compare_create_file():
        """创建新的产品对比文件"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未授权'}), 401
        
        manager = current_app.config['PRODUCT_COMPARE_MANAGER']
        payload = request.json or {}
        name = payload.get('name', '未命名对比')
        
        try:
            file_data = manager.create_file(name)
            _product_compare_kb_rebuild_all_async()
            return jsonify({'success': True, 'file': file_data})
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

    @app.route('/api/product_compare/files/<file_id>', methods=['GET', 'PUT', 'DELETE'])
    def product_compare_file(file_id):
        """获取、更新或删除产品对比文件"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未授权'}), 401
        
        manager = current_app.config['PRODUCT_COMPARE_MANAGER']
        
        if request.method == 'GET':
            try:
                file_data = manager.get_file(file_id)
                return jsonify({'success': True, 'file': file_data})
            except FileNotFoundError:
                return jsonify({'success': False, 'error': '文件不存在'}), 404
            except Exception as exc:
                return jsonify({'success': False, 'error': str(exc)}), 500
        
        if request.method == 'PUT':
            payload = request.json or {}
            try:
                file_data = manager.update_file(file_id, name=payload.get('name'))
                _product_compare_kb_rebuild_all_async()
                return jsonify({'success': True, 'file': file_data})
            except FileNotFoundError:
                return jsonify({'success': False, 'error': '文件不存在'}), 404
            except Exception as exc:
                return jsonify({'success': False, 'error': str(exc)}), 400
        
        # DELETE
        try:
            manager.delete_file(file_id)
            _product_compare_kb_rebuild_all_async()
            return jsonify({'success': True, 'file_id': file_id})
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 500

    @app.route('/api/product_compare/files/<file_id>/attributes', methods=['POST'])
    def product_compare_add_attribute(file_id):
        """添加属性"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未授权'}), 401
        
        manager = current_app.config['PRODUCT_COMPARE_MANAGER']
        payload = request.json or {}
        
        try:
            attr = manager.add_attribute(
                file_id,
                name=payload.get('name'),
                is_common=payload.get('is_common', False),
                attr_type=payload.get('type', 'text')
            )
            # 如果添加成功，更新单位和方向
            if attr and payload.get('unit') is not None or payload.get('direction') is not None:
                attr = manager.update_attribute(
                    file_id,
                    attr['id'],
                    unit=payload.get('unit', ''),
                    direction=payload.get('direction', 'higher')
                )
            file_data = manager.get_file(file_id)
            _product_compare_kb_rebuild_all_async()
            return jsonify({'success': True, 'attribute': attr, 'file': file_data})
        except FileNotFoundError:
            return jsonify({'success': False, 'error': '文件不存在'}), 404
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

    @app.route('/api/product_compare/files/<file_id>/attributes/<attr_id>', methods=['PUT', 'DELETE'])
    def product_compare_attribute(file_id, attr_id):
        """更新或删除属性"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未授权'}), 401
        
        manager = current_app.config['PRODUCT_COMPARE_MANAGER']
        
        if request.method == 'PUT':
            payload = request.json or {}
            try:
                attr = manager.update_attribute(
                    file_id,
                    attr_id,
                    name=payload.get('name'),
                    is_common=payload.get('is_common'),
                    attr_type=payload.get('type'),
                    order=payload.get('order'),
                    unit=payload.get('unit'),
                    direction=payload.get('direction')
                )
                file_data = manager.get_file(file_id)
                _product_compare_kb_rebuild_all_async()
                return jsonify({'success': True, 'attribute': attr, 'file': file_data})
            except FileNotFoundError:
                return jsonify({'success': False, 'error': '文件不存在'}), 404
            except Exception as exc:
                return jsonify({'success': False, 'error': str(exc)}), 400
        
        # DELETE
        try:
            manager.delete_attribute(file_id, attr_id)
            file_data = manager.get_file(file_id)
            _product_compare_kb_rebuild_all_async()
            return jsonify({'success': True, 'attribute_id': attr_id, 'file': file_data})
        except FileNotFoundError:
            return jsonify({'success': False, 'error': '文件不存在'}), 404
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

    @app.route('/api/product_compare/files/<file_id>/attributes/reorder', methods=['POST'])
    def product_compare_reorder_attributes(file_id):
        """重新排序属性"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未授权'}), 401
        
        manager = current_app.config['PRODUCT_COMPARE_MANAGER']
        payload = request.json or {}
        attr_orders = payload.get('orders', [])
        
        try:
            manager.reorder_attributes(file_id, attr_orders)
            file_data = manager.get_file(file_id)
            _product_compare_kb_rebuild_all_async()
            return jsonify({'success': True, 'file': file_data})
        except FileNotFoundError:
            return jsonify({'success': False, 'error': '文件不存在'}), 404
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

    @app.route('/api/product_compare/files/<file_id>/products', methods=['POST'])
    def product_compare_add_product(file_id):
        """添加产品"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未授权'}), 401
        
        manager = current_app.config['PRODUCT_COMPARE_MANAGER']
        payload = request.json or {}
        
        try:
            product = manager.add_product(
                file_id,
                name=payload.get('name'),
                belonging=payload.get('belonging'),
                attributes=payload.get('attributes', {}),
                link=payload.get('link')
            )
            file_data = manager.get_file(file_id)
            _product_compare_kb_rebuild_all_async()
            return jsonify({'success': True, 'product': product, 'file': file_data})
        except FileNotFoundError:
            return jsonify({'success': False, 'error': '文件不存在'}), 404
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

    @app.route('/api/product_compare/files/<file_id>/products/<product_id>', methods=['PUT', 'DELETE'])
    def product_compare_product(file_id, product_id):
        """更新或删除产品"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未授权'}), 401
        
        manager = current_app.config['PRODUCT_COMPARE_MANAGER']
        
        if request.method == 'PUT':
            payload = request.json or {}
            try:
                product = manager.update_product(
                    file_id,
                    product_id,
                    name=payload.get('name'),
                    belonging=payload.get('belonging'),
                    attributes=payload.get('attributes'),
                    link=payload.get('link')
                )
                file_data = manager.get_file(file_id)
                _product_compare_kb_rebuild_all_async()
                return jsonify({'success': True, 'product': product, 'file': file_data})
            except FileNotFoundError:
                return jsonify({'success': False, 'error': '文件不存在'}), 404
            except Exception as exc:
                return jsonify({'success': False, 'error': str(exc)}), 400
        
        # DELETE
        try:
            manager.delete_product(file_id, product_id)
            file_data = manager.get_file(file_id)
            _product_compare_kb_rebuild_all_async()
            return jsonify({'success': True, 'product_id': product_id, 'file': file_data})
        except FileNotFoundError:
            return jsonify({'success': False, 'error': '文件不存在'}), 404
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

    @app.route('/api/product_compare/files/<file_id>/products/reorder', methods=['POST'])
    def product_compare_reorder_products(file_id):
        """重新排序产品"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未授权'}), 401
        
        manager = current_app.config['PRODUCT_COMPARE_MANAGER']
        payload = request.json or {}
        product_orders = payload.get('orders', [])
        
        try:
            manager.reorder_products(file_id, product_orders)
            file_data = manager.get_file(file_id)
            _product_compare_kb_rebuild_all_async()
            return jsonify({'success': True, 'file': file_data})
        except FileNotFoundError:
            return jsonify({'success': False, 'error': '文件不存在'}), 404
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

    @app.route('/api/product_compare/files/<file_id>/belongings', methods=['GET'])
    def product_compare_belongings(file_id):
        """获取文件中所有已使用的归属列表（包含颜色）"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未授权'}), 401
        
        manager = current_app.config['PRODUCT_COMPARE_MANAGER']
        try:
            belongings = manager.get_belongings(file_id)
            return jsonify({'success': True, 'belongings': belongings})
        except FileNotFoundError:
            return jsonify({'success': False, 'error': '文件不存在'}), 404
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 500
    
    @app.route('/api/product_compare/files/<file_id>/belongings/<belonging>/color', methods=['PUT'])
    def product_compare_set_belonging_color(file_id, belonging):
        """设置归属的颜色"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未授权'}), 401
        
        manager = current_app.config['PRODUCT_COMPARE_MANAGER']
        payload = request.json or {}
        color = payload.get('color', '').strip()
        
        if not color or not color.startswith('#'):
            return jsonify({'success': False, 'error': '颜色格式无效'}), 400
        
        try:
            manager._set_belonging_color(file_id, belonging, color)
            # 更新所有使用该归属的产品的颜色
            data = manager._load_file(file_id)
            for product in data.get('products', []):
                if product.get('belonging', '').strip() == belonging:
                    product['color'] = color
            manager._save_file(file_id, data)
            _product_compare_kb_rebuild_all_async()
            return jsonify({'success': True, 'file': data})
        except FileNotFoundError:
            return jsonify({'success': False, 'error': '文件不存在'}), 404
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 500

    @app.route('/api/favorites', methods=['GET'])
    def get_favorites():
        """获取收藏列表"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        try:
            data = load_favorites()
            return jsonify({'success': True, 'data': data})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    @app.route('/api/favorites/groups/<group_name>', methods=['GET'])
    def get_favorites_by_group(group_name):
        """获取指定分组的收藏文件"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        try:
            data = load_favorites()
            items = [item for item in data.get('items', []) if item.get('group') == group_name]
            return jsonify({'success': True, 'items': items, 'group': group_name})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500

    @app.route('/file_browser')
    def file_browser():
        """文件浏览器页面，显示文件列表"""
        # 添加登录验证
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        
        root_dir = current_app.config.get('ROOT_DIR')
        if not root_dir or not os.path.exists(root_dir) or not os.path.isdir(root_dir):
            # 如果没有设置根目录或根目录无效，重定向到设置页面
            return redirect(url_for('set_root'))
        
        # 获取请求路径
        path = request.args.get('path', '')
        if path:
            current_path = os.path.join(root_dir, path)
        else:
            current_path = root_dir
        
        # 安全检查：防止路径遍历
        try:
            current_path = os.path.normpath(current_path)
            if not current_path.startswith(os.path.normpath(root_dir)):
                abort(404)
        except Exception:
            abort(404)
        
        # 检查路径是否存在且是目录
        if not os.path.exists(current_path) or not os.path.isdir(current_path):
            abort(404)
        
        # 获取目录内容
        try:
            items = os.listdir(current_path)
        except Exception:
            items = []
        
        # 分类文件和目录
        directories = []
        files = []
        
        for item in items:
            item_path = os.path.join(current_path, item)
            item_rel_path = posixpath.join(path, item)
            
            if os.path.isdir(item_path):
                directories.append({
                    'name': item,
                    'path': item_rel_path,
                    'is_dir': True
                })
            else:
                _, ext = os.path.splitext(item.lower())
                files.append({
                    'name': item,
                    'path': item_rel_path,
                    'is_dir': False,
                    'is_markdown': ext in MARKDOWN_EXTENSIONS,
                    'is_image': ext in IMAGE_EXTENSIONS,
                    'is_pdf': ext in PDF_EXTENSIONS,
                    'is_office': ext in OFFICE_EXTENSIONS,
                    'is_video': ext in VIDEO_EXTENSIONS
                })
        
        # 排序：目录在前，按名称排序
        directories.sort(key=lambda x: x['name'].lower())
        files.sort(key=lambda x: x['name'].lower())
        
        # 构建面包屑导航
        path_parts = []
        current_dir = ''
        for part in path.split(os.sep):
            if part:
                current_dir = posixpath.join(current_dir, part)
                path_parts.append({
                    'name': part,
                    'path': current_dir
                })
        
        # 生成目录路径的上一级目录
        parent_dir = os.path.dirname(current_path)
        if parent_dir and parent_dir != root_dir:
            parent_rel_path = posixpath.dirname(path)
        else:
            parent_rel_path = ''
        
        # 加载收藏数据
        favorites_data = load_favorites()
        
        return render_template('index.html', 
                              directories=directories, 
                              files=files, 
                              current_path=path, 
                              path_parts=path_parts,
                              parent_rel_path=parent_rel_path,
                              favorites_groups=favorites_data.get('groups', []))
    
    @app.route('/view/<path:filepath>')
    def view_file(filepath):
        """旧的文件预览路由 (完整页面) - 保留以防万一或直接访问"""
        # 添加登录验证
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        root_dir = current_app.config.get('ROOT_DIR')
        if not root_dir:
            abort(404)

        full_path = os.path.join(root_dir, filepath)

        try:
            full_path = os.path.normpath(full_path)
            if not full_path.startswith(os.path.normpath(root_dir)):
                abort(404)
        except Exception:
            abort(404)

        if not os.path.exists(full_path) or os.path.isdir(full_path):
            abort(404)

        filename = os.path.basename(full_path)
        _, ext = os.path.splitext(filename.lower())
        file_type = 'unknown'
        content = ''

        if ext in OFFICE_EXTENSIONS:
            file_type = 'office'
        elif ext in IMAGE_EXTENSIONS:
            file_type = 'image'
        elif ext in MARKDOWN_EXTENSIONS:
            file_type = 'markdown'
            try:
                with open(full_path, 'r', encoding='utf-8') as f:
                    file_content = f.read()
                
                # 使用新的markdown-it-py渲染
                content = render_markdown_content(file_content, filepath)
            except Exception as e:
                content = f"<p>读取文件时出错: {e}</p>"
        elif ext in PDF_EXTENSIONS:
            file_type = 'pdf'
        elif ext in VIDEO_EXTENSIONS:
            file_type = 'video'

        parent_dir = posixpath.dirname(filepath)
        if not parent_dir:
            back_url = url_for('file_browser')
        else:
            back_url = url_for('file_browser', path=parent_dir)

        return render_template('view_file.html',
                               filename=filename,
                               filepath=filepath,
                               file_type=file_type,
                               content=content,
                               back_url=back_url)
    
    @app.route('/view_file')
    def view_file_compat():
        """兼容的文件预览路由，接受path参数"""
        filepath = request.args.get('path', '')
        root = request.args.get('root', '')  # 支持root参数，用于Git仓库文件预览
        
        if filepath:
            # 如果有root参数，使用root作为基础路径（用于Git仓库文件）
            if root:
                try:
                    if os.path.isabs(root):
                        base_path = os.path.normpath(root)
                    else:
                        # 如果是相对路径，尝试使用Git工作目录
                        git_workdir = current_app.config.get('GIT_WORKDIR', '')
                        if git_workdir and os.path.exists(git_workdir):
                            base_path = os.path.normpath(os.path.join(git_workdir, root))
                        else:
                            root_dir = current_app.config.get('ROOT_DIR')
                            base_path = os.path.normpath(os.path.join(root_dir or os.getcwd(), root))
                    
                    full_path = os.path.normpath(os.path.join(base_path, filepath))
                    # 验证路径安全性
                    if not full_path.startswith(base_path):
                        return render_template('view_file.html', filename=filepath, filepath=filepath, file_type='text', content="<p>无效的文件路径</p>", back_url=url_for('index'))
                except Exception:
                    return render_template('view_file.html', filename=filepath, filepath=filepath, file_type='text', content="<p>文件路径无效</p>", back_url=url_for('index'))
            else:
                # 不再重定向，直接处理文件预览逻辑
                root_dir = current_app.config.get('ROOT_DIR')
                if not root_dir:
                    # 如果没有设置ROOT_DIR，尝试在当前目录查找
                    full_path = os.path.join(os.getcwd(), filepath)
                else:
                    full_path = os.path.join(root_dir, filepath)
                
            try:
                full_path = os.path.normpath(full_path)
            except Exception:
                return render_template('view_file.html', filename=filepath, filepath=filepath, file_type='text', content="<p>文件路径无效</p>", back_url=url_for('index'))

            if not os.path.exists(full_path) or os.path.isdir(full_path):
                # 尝试在当前目录直接查找
                alt_path = os.path.join(os.getcwd(), filepath)
                if os.path.exists(alt_path) and not os.path.isdir(alt_path):
                    full_path = alt_path
                else:
                    return render_template('view_file.html', filename=filepath, filepath=filepath, file_type='text', content="<p>文件不存在</p>", back_url=url_for('index'))

            filename = os.path.basename(full_path)
            _, ext = os.path.splitext(filename.lower())
            file_type = 'unknown'
            content = ''

            if ext in MARKDOWN_EXTENSIONS:
                file_type = 'markdown'
                try:
                    with open(full_path, 'r', encoding='utf-8') as f:
                        file_content = f.read()
                    
                    # 使用新的markdown-it-py渲染
                    content = render_markdown_content(file_content, filepath)
                except Exception as e:
                    content = f"<p>读取文件时出错: {e}</p>"
            else:
                file_type = 'text'
                try:
                    with open(full_path, 'r', encoding='utf-8') as f:
                        content = f.read().replace('\n', '<br>')
                except Exception as e:
                    content = f"<p>读取文件时出错: {e}</p>"

            return render_template('view_file.html',
                                   filename=filename,
                                   filepath=filepath,
                                   file_type=file_type,
                                   content=content,
                                   back_url=url_for('file_browser'))
        return render_template('view_file.html', filename="", filepath="", file_type='text', content="<p>未指定文件</p>", back_url=url_for('file_browser'))
    
    @app.route('/download/<path:filepath>')
    def download_file(filepath):
        """下载文件路由"""
        root_dir = current_app.config.get('ROOT_DIR')
        if not root_dir:
            # 如果没有设置ROOT_DIR，尝试在当前目录查找
            root_dir = os.getcwd()
        
        # 安全检查：防止路径遍历
        try:
            full_path = os.path.join(root_dir, filepath)
            full_path = os.path.normpath(full_path)
            if not full_path.startswith(os.path.normpath(root_dir)):
                abort(404)
        except Exception:
            abort(404)
        
        if not os.path.exists(full_path) or os.path.isdir(full_path):
            # 尝试直接在当前目录查找
            alt_path = os.path.join(os.getcwd(), filepath)
            if os.path.exists(alt_path) and not os.path.isdir(alt_path):
                full_path = alt_path
            else:
                abort(404)
        
        # 获取文件所在目录和文件名
        directory = os.path.dirname(full_path)
        filename = os.path.basename(full_path)
        
        # 使用send_from_directory发送文件
        return send_from_directory(directory, filename, as_attachment=True)
    
    @app.route('/preview/<path:filepath>')
    def preview_file(filepath):
        """预览文件路由（非下载）"""
        root_dir = current_app.config.get('ROOT_DIR')
        if not root_dir:
            # 如果没有设置ROOT_DIR，尝试在当前目录查找
            root_dir = os.getcwd()
        
        # 安全检查：防止路径遍历
        try:
            full_path = os.path.join(root_dir, filepath)
            full_path = os.path.normpath(full_path)
            if not full_path.startswith(os.path.normpath(root_dir)):
                abort(404)
        except Exception:
            abort(404)
        
        if not os.path.exists(full_path) or os.path.isdir(full_path):
            # 尝试直接在当前目录查找
            alt_path = os.path.join(os.getcwd(), filepath)
            if os.path.exists(alt_path) and not os.path.isdir(alt_path):
                full_path = alt_path
            else:
                abort(404)
        
        # 获取文件所在目录和文件名
        directory = os.path.dirname(full_path)
        filename = os.path.basename(full_path)
        
        # 根据文件扩展名设置正确的MIME类型
        _, ext = os.path.splitext(filename.lower())
        mimetype = None
        
        if ext in IMAGE_EXTENSIONS:
            if ext in ['.jpg', '.jpeg']:
                mimetype = 'image/jpeg'
            elif ext == '.png':
                mimetype = 'image/png'
            elif ext == '.gif':
                mimetype = 'image/gif'
            elif ext == '.svg':
                mimetype = 'image/svg+xml'
            elif ext == '.webp':
                mimetype = 'image/webp'
            elif ext == '.avif':
                mimetype = 'image/avif'
            elif ext == '.ico':
                mimetype = 'image/x-icon'
            else:
                mimetype = 'image/jpeg'
        elif ext == '.pdf':
            mimetype = 'application/pdf'
        elif ext in VIDEO_EXTENSIONS:
            if ext == '.mp4':
                mimetype = 'video/mp4'
            elif ext == '.avi':
                mimetype = 'video/x-msvideo'
            elif ext == '.mov':
                mimetype = 'video/quicktime'
            elif ext == '.wmv':
                mimetype = 'video/x-ms-wmv'
            elif ext == '.webm':
                mimetype = 'video/webm'
            elif ext == '.m4v':
                mimetype = 'video/x-m4v'
            else:
                mimetype = 'video/mp4'
        elif ext in AUDIO_EXTENSIONS:
            if ext == '.mp3':
                mimetype = 'audio/mpeg'
            elif ext == '.wav':
                mimetype = 'audio/wav'
            elif ext == '.ogg':
                mimetype = 'audio/ogg'
            elif ext == '.flac':
                mimetype = 'audio/flac'
            elif ext == '.wma':
                mimetype = 'audio/x-ms-wma'
            elif ext in ['.m4a', '.aac']:
                mimetype = 'audio/mp4'
            elif ext in ['.opus', '.weba']:
                mimetype = 'audio/ogg'
            elif ext in ['.aiff', '.aif']:
                mimetype = 'audio/aiff'
            elif ext == '.amr':
                mimetype = 'audio/amr'
            else:
                mimetype = 'audio/mpeg'
        
        # 不设置as_attachment，这样浏览器会尝试预览而不是下载
        return send_from_directory(directory, filename, as_attachment=False, mimetype=mimetype)
    
    @app.route('/set_root', methods=['GET', 'POST'])
    def set_root():
        """设置根目录（需要管理员密码）"""
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        
        current_root = current_app.config.get('ROOT_DIR')
        
        if request.method == 'POST':
            new_root = request.form.get('root_path')
            admin_password = request.form.get('admin_password')  # 获取管理员密码
            
            # 验证管理员密码
            correct_admin_password = current_app.config.get('ADMIN_PASSWORD', 'admin123')
            if admin_password != correct_admin_password:
                return render_template('set_root.html', 
                                     error="管理员密码错误！无法修改目录。", 
                                     current_root=current_root)
            
            if new_root and os.path.exists(new_root) and os.path.isdir(new_root):
                current_app.config['ROOT_DIR'] = new_root
                # 保存到配置文件，保留所有配置
                config = configparser.ConfigParser()
                current_password = current_app.config.get('PASSWORD', 'ats123')
                config['settings'] = {
                    'root_dir': new_root,
                    'password': current_password,
                    'admin_password': correct_admin_password
                }
                # 获取配置文件路径
                config_file = current_app.config.get('CONFIG_FILE', 'config.ini')
                with open(config_file, 'w', encoding='utf-8') as f:
                    config.write(f)
                return redirect(url_for('file_browser'))
            else:
                return render_template('set_root.html', error="无效的目录路径", current_root=current_root)
        else:
            # GET请求，显示设置页面
            return render_template('set_root.html', current_root=current_root)
    
    @app.route('/set_current_as_root', methods=['POST'])
    def set_current_as_root():
        """将当前浏览的文件夹设置为共享根目录（需要管理员密码）"""
        if 'logged_in' not in session:
            return jsonify({'error': '请先登录'}), 401
        
        data = request.get_json()
        current_path = data.get('path', '')
        admin_password = data.get('admin_password', '')  # 获取管理员密码
        
        # 验证管理员密码
        correct_admin_password = current_app.config.get('ADMIN_PASSWORD', 'admin123')
        if admin_password != correct_admin_password:
            return jsonify({'error': '管理员密码错误！无法修改目录。', 'require_password': True}), 403
        
        # 获取当前根目录
        root_dir = current_app.config.get('ROOT_DIR')
        if not root_dir:
            return jsonify({'error': '根目录未设置'}), 400
        
        # 计算新的根目录路径
        if current_path:
            new_root = os.path.join(root_dir, current_path)
        else:
            new_root = root_dir
        
        # 验证新路径
        try:
            new_root = os.path.normpath(new_root)
            if not os.path.exists(new_root) or not os.path.isdir(new_root):
                return jsonify({'error': '目标路径不存在或不是目录'}), 400
        except Exception as e:
            return jsonify({'error': f'路径验证失败: {e}'}), 400
        
        # 更新配置
        current_app.config['ROOT_DIR'] = new_root
        
        # 保存到配置文件
        try:
            config = configparser.ConfigParser()
            current_password = current_app.config.get('PASSWORD', 'ats123')
            
            config['settings'] = {
                'root_dir': new_root,
                'password': current_password,
                'admin_password': correct_admin_password
            }
            
            config_file = current_app.config.get('CONFIG_FILE', 'config.ini')
            with open(config_file, 'w', encoding='utf-8') as f:
                config.write(f)
            
            return jsonify({
                'success': True,
                'new_root': new_root,
                'message': '共享文件夹已更新'
            })
        except Exception as e:
            return jsonify({'error': f'保存配置失败: {e}'}), 500
    
    @app.route('/view_model/<path:filepath>')
    def view_model(filepath):
        """3D模型查看器页面"""
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        
        root_dir = current_app.config.get('ROOT_DIR')
        if not root_dir:
            abort(404)
        
        full_path = os.path.join(root_dir, filepath)
        
        # 安全检查
        try:
            full_path = os.path.normpath(full_path)
            if not full_path.startswith(os.path.normpath(root_dir)):
                abort(404)
        except Exception:
            abort(404)
        
        if not os.path.exists(full_path) or os.path.isdir(full_path):
            abort(404)
        
        filename = os.path.basename(full_path)
        _, ext = os.path.splitext(filename.lower())
        
        # 检查是否是3D模型文件
        if ext not in MODEL_3D_EXTENSIONS:
            abort(400)
        
        # 生成模型URL
        model_url = url_for('preview_file', filepath=filepath)
        
        # 返回按钮URL
        parent_dir = posixpath.dirname(filepath)
        if not parent_dir:
            back_url = url_for('file_browser')
        else:
            back_url = url_for('file_browser', path=parent_dir)
        
        return render_template('model_viewer.html',
                             filename=filename,
                             filepath=filepath,
                             file_ext=ext,
                             model_url=model_url,
                             back_url=back_url)
    
    @app.route('/get_preview_content', methods=['POST'])
    def get_preview_content():
        """获取文件预览内容"""
        # 添加登录验证
        if 'logged_in' not in session:
            return jsonify({'error': '请先登录'}), 401
        
        data = request.get_json()
        filepath = data.get('filepath')
        git_root = data.get('git_root')  # Git仓库根路径（可选）
        
        if not filepath:
            return jsonify({'error': '文件路径不能为空'}), 400
        
        # 如果提供了Git仓库根路径，使用Git工作目录
        if git_root:
            git_workdir = current_app.config.get('GIT_WORKDIR', '')
            if git_workdir:
                # 构建完整路径
                if os.path.isabs(git_root):
                    repo_root = os.path.normpath(git_root)
                else:
                    repo_root = os.path.normpath(os.path.join(git_workdir, git_root))
                
                # 验证路径在Git工作目录内
                git_workdir_norm = os.path.normpath(git_workdir)
                if not repo_root.startswith(git_workdir_norm):
                    return jsonify({'error': 'Git仓库路径不在工作目录内'}), 403
                
                # 构建文件完整路径
                full_path = os.path.normpath(os.path.join(repo_root, filepath))
                
                # 再次验证路径安全性（防止路径遍历）
                if not full_path.startswith(repo_root):
                    return jsonify({'error': '访问被拒绝：路径遍历检测'}), 403
            else:
                return jsonify({'error': 'Git工作目录未设置'}), 400
        else:
            # 使用默认的ROOT_DIR
            root_dir = current_app.config.get('ROOT_DIR')
            if not root_dir:
                # 如果没有设置ROOT_DIR，尝试在当前目录查找
                root_dir = os.getcwd()
            
            # 安全检查：防止路径遍历
            try:
                full_path = os.path.join(root_dir, filepath)
                full_path = os.path.normpath(full_path)
                if not full_path.startswith(os.path.normpath(root_dir)):
                    return jsonify({'error': '访问被拒绝'}), 403
            except Exception:
                return jsonify({'error': '路径解析错误'}), 400
        
        if not os.path.exists(full_path) or os.path.isdir(full_path):
            return jsonify({'error': '文件不存在'}), 404
        
        filename = os.path.basename(full_path)
        _, ext = os.path.splitext(filename.lower())
        file_type = 'unknown'
        content_html = ''
        download_url = url_for('download_file', filepath=filepath)
        preview_url = url_for('preview_file', filepath=filepath)
        
        # 代码文件扩展名列表
        CODE_EXTENSIONS = ['.py', '.js', '.html', '.css', '.scss', '.php', '.java', '.c', '.cpp',
                          '.cs', '.go', '.rb', '.sh', '.bat', '.sql', '.ts', '.tsx', '.jsx',
                          '.json', '.xml', '.yaml', '.yml', '.md', '.markdown', '.txt', '.csv', '.log',
                          '.ini', '.toml', '.env', '.conf', '.cfg', '.properties',
                          '.vue', '.svelte', '.kt', '.rs', '.swift', '.lua', '.ps1']
        
        if ext in MARKDOWN_EXTENSIONS:
            file_type = 'markdown'
            try:
                with open(full_path, 'r', encoding='utf-8') as f:
                    markdown_content = f.read()
                
                # 使用新的markdown-it-py渲染
                content_html = render_markdown_content(markdown_content, filepath)
                
                # 将下载链接替换为预览链接
                content_html = content_html.replace('/download/', '/preview/')
            except Exception as e:
                content_html = f'<p>读取文件时出错: {e}</p>'
            content_html = f'<article class="markdown-body">{content_html}</article>'
        elif ext in PDF_EXTENSIONS:
            file_type = 'pdf'
            content_html = f'<div class="pdf-container"><embed src="{preview_url}" type="application/pdf"></div>'
        elif ext in IMAGE_EXTENSIONS:
            file_type = 'image'
            content_html = f'<div class="image-container"><img src="{preview_url}" alt="{filename}"></div>'
        elif ext in VIDEO_EXTENSIONS:
            file_type = 'video'
            content_html = f'<div class="video-container"><video controls src="{preview_url}"></video></div>'
        elif ext in AUDIO_EXTENSIONS:
            file_type = 'audio'
            content_html = f'''
            <div class="audio-player-container">
                <div class="audio-info">
                    <h3><i class="fas fa-music"></i> {filename}</h3>
                    <p>音频文件</p>
                </div>
                <audio controls src="{preview_url}" style="width: 100%; max-width: 600px;">
                    您的浏览器不支持音频播放
                </audio>
            </div>
            '''
        elif ext in DRAWIO_EXTENSIONS:
            file_type = 'drawio'
            # 提供draw.io文件的预览功能，编辑按钮在header中
            content_html = f'''
            <div class="drawio-container">
                <iframe src="/drawio_embed?filepath={quote(filepath)}" class="drawio-preview" width="100%" height="600px" style="border: none;"></iframe>
            </div>
            '''
        elif ext in MODEL_3D_EXTENSIONS:
            file_type = '3d_model'
            # 3D模型预览 - iframe嵌入方式
            view_model_url = url_for('view_model', filepath=filepath)
            content_html = f'''
            <div class="model-3d-container" style="width: 100%; height: 700px; position: relative; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 12px rgba(0,0,0,0.1);">
                <iframe src="{view_model_url}" 
                        style="width: 100%; height: 100%; border: none; display: block;"
                        allowfullscreen>
                </iframe>
                <div style="position: absolute; top: 10px; left: 10px; background: rgba(0,0,0,0.7); color: white; padding: 8px 12px; border-radius: 6px; font-size: 12px; z-index: 100;">
                    <i class="fas fa-cube"></i> {filename} ({ext.upper()})
                </div>
            </div>
            <div style="margin-top: 15px; padding: 15px; background: #f8f9fa; border-radius: 8px; border-left: 4px solid #409EFF;">
                <p style="margin: 5px 0; color: #666;"><strong>💡 操作提示:</strong></p>
                <p style="margin: 5px 0; color: #666;">🖱️ <strong>鼠标左键</strong> - 旋转模型 | <strong>右键</strong> - 平移视角 | <strong>滚轮</strong> - 缩放</p>
                <p style="margin: 5px 0; color: #666;">🛠️ 使用右上角工具栏可以: 重置视角、切换线框模式、显示网格、改变背景色等</p>
            </div>
            '''
        elif ext in CODE_EXTENSIONS:
            file_type = 'code'
            try:
                with open(full_path, 'r', encoding='utf-8') as f:
                    code_content = f.read()
                    # 获取语言名称（从扩展名映射）
                    language_map = {
                        '.py': 'python',
                        '.js': 'javascript',
                        '.html': 'html',
                        '.css': 'css',
                        '.scss': 'scss',
                        '.php': 'php',
                        '.java': 'java',
                        '.c': 'c',
                        '.cpp': 'cpp',
                        '.cs': 'csharp',
                        '.go': 'go',
                        '.rb': 'ruby',
                        '.sh': 'bash',
                        '.bat': 'batch',
                        '.sql': 'sql',
                        '.ts': 'typescript',
                        '.tsx': 'typescript',
                        '.jsx': 'javascript',
                        '.json': 'json',
                        '.xml': 'xml',
                        '.yaml': 'yaml',
                        '.yml': 'yaml',
                        '.ini': 'ini',
                        '.toml': 'toml',
                        '.env': 'shell',
                        '.conf': 'ini',
                        '.cfg': 'ini',
                        '.properties': 'ini',
                        '.vue': 'html',
                        '.svelte': 'html',
                        '.kt': 'kotlin',
                        '.rs': 'rust',
                        '.swift': 'swift',
                        '.lua': 'lua',
                        '.ps1': 'powershell',
                        '.md': 'markdown',
                        '.markdown': 'markdown',
                        '.txt': 'text',
                        '.csv': 'csv',
                        '.log': 'text'
                    }
                    language = language_map.get(ext, 'text')
                    # 使用HTML转义并添加语法高亮标记
                    escaped_content = code_content.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                    content_html = f'<pre class="code-preview language-{language}"><code>{escaped_content}</code></pre>'
            except Exception as e:
                content_html = f'<p>无法预览此文件: {e}</p>'
        else:
            file_type = 'text'
            try:
                with open(full_path, 'r', encoding='utf-8') as f:
                    text_content = f.read()
                    content_html = f'<pre>{text_content}</pre>'
            except Exception as e:
                content_html = f'<p>无法预览此文件: {e}</p>'
        
        return jsonify({
            'filename': filename,
            'file_type': file_type,
            'content_html': content_html,
            'download_url': download_url,
            'preview_url': preview_url
        })
    
    # 用户认证相关路由
    @app.route('/login', methods=['GET', 'POST'])
    def login():
        """登录页面"""
        if request.method == 'POST':
            password = request.form.get('password')
            
            # 使用配置文件中的密码进行验证
            configured_password = current_app.config.get('PASSWORD')  # 修正：使用大写的键名

            _routes_logger.debug(
                "登录尝试（脱敏）: supplied_len=%s configured_len=%s config_file=%s match=%s",
                len(password or ''),
                len(configured_password or ''),
                current_app.config.get('CONFIG_FILE'),
                password == configured_password,
            )

            if not configured_password:
                # 如果没有配置密码，使用默认密码并记录警告
                configured_password = 'ats123'
                _routes_logger.warning("未找到配置密码，已使用内置默认值（请在设置中修改）")

            if password == configured_password:
                session['logged_in'] = True
                _routes_logger.info("用户登录成功")
                # 登录后重定向到选择页面
                return redirect(url_for('index'))
            else:
                _routes_logger.warning("登录失败: 密码不匹配")
                return render_template('login.html', error="密码错误")
        return render_template('login.html')
    

    
    @app.route('/logout')
    def logout():
        """登出功能"""
        session.pop('logged_in', None)
        return redirect(url_for('login'))
    
    @app.route('/help')
    def help_page():
        """帮助页面"""
        # 帮助页面不需要登录也可以查看
        return render_template('help.html')
    
    @app.route('/theme_preview')
    def theme_preview():
        """主题预览页面，展示新配色与控件示例"""
        return render_template('theme_preview.html')
    
    # 串口调试助手
    @app.route('/serial_tool')
    def serial_tool():
        """串口调试助手页面"""
        if not is_logged_in():
            return redirect(url_for('login'))
        serial_server_config = get_serial_server_config('/serial_tool')
        if (
            serial_server_config['remote_serial_enabled']
            and serial_server_config['compat_mode']
            and serial_server_config['serial_https_active']
            and not request.is_secure
        ):
            return redirect(serial_server_config['serial_https_url'])
        return render_template(
            'serial_tool.html',
            serial_server_config=serial_server_config
        )

    def _normalize_parity(value: str) -> str:
        mapping = {
            'none': 'N',
            'n': 'N',
            'even': 'E',
            'e': 'E',
            'odd': 'O',
            'o': 'O',
            'mark': 'M',
            'm': 'M',
            'space': 'S',
            's': 'S',
        }
        if not value:
            return 'N'
        value = value.strip()
        return mapping.get(value.lower(), 'N')

    def _is_safe_port_id(value: str) -> bool:
        return bool(value) and re.fullmatch(r'[A-Za-z0-9_]+', value) is not None

    @app.route('/api/serial/ports', methods=['GET'])
    def api_serial_ports():
        """列出服务器可用串口"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        try:
            ports = serial_manager.list_available_ports()
            for port in ports:
                device = port.get('device')
                if device:
                    port['recommended_id'] = serial_manager.generate_port_id(device, 9600)
            return jsonify({'success': True, 'ports': ports})
        except Exception as exc:  # pylint: disable=broad-except
            return jsonify({'success': False, 'error': str(exc)}), 500
    @app.route('/api/serial/channels', methods=['GET'])
    def api_serial_channels():
        if not is_logged_in():
            return jsonify({'success': False, 'error': 'not_logged_in'}), 401
        try:
            return jsonify({'success': True, 'channels': shared_serial_hub.list_channels()})
        except Exception as exc:  # pylint: disable=broad-except
            return jsonify({'success': False, 'error': str(exc)}), 500
            return jsonify({'success': False, 'error': str(exc)}), 500

    @app.route('/api/serial/capture/start', methods=['POST'])
    def api_serial_capture_start():
        """开启指定串口的持续日志"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401

        payload = request.json or {}
        device = (payload.get('device') or '').strip()
        if not device:
            return jsonify({'success': False, 'error': '缺少 device 参数'}), 400

        try:
            baudrate = int(payload.get('baudrate', 9600))
            bytesize = int(payload.get('bytesize', 8))
            stopbits = int(payload.get('stopbits', 1))
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': '串口参数格式错误'}), 400

        parity = _normalize_parity(payload.get('parity', 'N'))
        port_id = payload.get('port_id')
        try:
            session = serial_manager.start_logging(
                device=device,
                baudrate=baudrate,
                bytesize=bytesize,
                parity=parity,
                stopbits=stopbits,
                port_id=port_id,
            )
            port_info = serial_manager.get_port_info(session['port_id'])
            session['port_info'] = port_info
            session['log_files'] = serial_manager.list_log_files(session['port_id'])
            return jsonify({'success': True, 'session': session})
        except ValueError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400
        except Exception as exc:  # pylint: disable=broad-except
            return jsonify({'success': False, 'error': str(exc)}), 500

    @app.route('/api/serial/capture/stop', methods=['POST'])
    def api_serial_capture_stop():
        """停止串口持续日志"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401

        payload = request.json or {}
        port_id = payload.get('port_id')
        if not _is_safe_port_id(port_id or ''):
            return jsonify({'success': False, 'error': '端口标识不合法'}), 400
        try:
            success = serial_manager.stop_logging(port_id)
            if not success:
                return jsonify({'success': False, 'error': '未找到对应日志会话'}), 404
            return jsonify({'success': True})
        except Exception as exc:  # pylint: disable=broad-except
            return jsonify({'success': False, 'error': str(exc)}), 500

    @app.route('/api/serial/capture/status', methods=['GET'])
    def api_serial_capture_status():
        """查询当前持续日志会话"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        try:
            sessions = serial_manager.get_logging_status()
            for session in sessions:
                port_id = session.get('port_id')
                if not port_id:
                    continue
                session['log_files'] = serial_manager.list_log_files(port_id)
                session['port_info'] = serial_manager.get_port_info(port_id)
            return jsonify({'success': True, 'sessions': sessions})
        except Exception as exc:  # pylint: disable=broad-except
            return jsonify({'success': False, 'error': str(exc)}), 500

    @app.route('/api/serial/capture/log/<port_id>', methods=['GET'])
    def api_serial_capture_log(port_id):
        """读取指定串口的日志条目"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        if not _is_safe_port_id(port_id):
            return jsonify({'success': False, 'error': '端口标识不合法'}), 400

        try:
            limit = int(request.args.get('limit', 500))
        except ValueError:
            return jsonify({'success': False, 'error': 'limit 参数错误'}), 400

        since = request.args.get('since')
        try:
            entries = serial_manager.read_log_entries(
                port_id=port_id,
                limit=max(0, limit),
                since=since,
            )
            return jsonify({'success': True, 'entries': entries})
        except Exception as exc:  # pylint: disable=broad-except
            return jsonify({'success': False, 'error': str(exc)}), 500
    
    # 保存串口日志
    def _serial_action_runner():
        socketio = getattr(current_app, 'socketio', None)
        return getattr(socketio, 'serial_action_runner', None) if socketio else None

    def _serial_emit_refresh(channel_id=None):
        socketio = getattr(current_app, 'socketio', None)
        if not socketio:
            return
        emit_channels = getattr(socketio, 'serial_emit_shared_channels', None)
        emit_channel_state = getattr(socketio, 'serial_emit_channel_state', None)
        if emit_channels:
            emit_channels()
        if channel_id and emit_channel_state:
            emit_channel_state(channel_id)

    def api_serial_channels_v2():
        if not is_logged_in():
            return jsonify({'success': False, 'error': 'not_logged_in'}), 401
        try:
            return jsonify({'success': True, 'channels': shared_serial_hub.list_channels()})
        except Exception as exc:  # pylint: disable=broad-except
            return jsonify({'success': False, 'error': str(exc)}), 500

    def api_serial_capture_start_v2():
        if not is_logged_in():
            return jsonify({'success': False, 'error': 'not_logged_in'}), 401

        payload = request.json or {}
        channel_id = (payload.get('channel_id') or payload.get('port_id') or '').strip()
        try:
            if not channel_id:
                device = (payload.get('device') or '').strip()
                if not device:
                    return jsonify({'success': False, 'error': 'missing_device'}), 400
                channel_info, _created = shared_serial_hub.ensure_server_channel(
                    device=device,
                    config={
                        'baudrate': int(payload.get('baudrate', 9600)),
                        'bytesize': int(payload.get('bytesize', 8)),
                        'parity': _normalize_parity(payload.get('parity', 'N')),
                        'stopbits': int(payload.get('stopbits', 1)),
                    },
                    display_name=payload.get('display_name') or device,
                )
                channel_id = channel_info['channel_id']

            session_info, actions, error = shared_serial_hub.start_capture(channel_id, started_by='http')
            if error or not session_info:
                return jsonify({'success': False, 'error': error or 'capture_start_failed'}), 400

            runner = _serial_action_runner()
            if runner and actions:
                runner(actions)

            channel = shared_serial_hub.get_channel(channel_id)
            if (
                channel
                and channel.get('source_type') == 'server_pyserial'
                and channel.get('capture_active')
                and channel.get('state') != 'active'
                and channel.get('last_error')
            ):
                shared_serial_hub.stop_capture(channel_id, reason='activation_failed')
                _serial_emit_refresh(channel_id)
                return jsonify({'success': False, 'error': channel.get('last_error') or 'server_open_failed'}), 400

            _serial_emit_refresh(channel_id)
            session_info['channel'] = channel
            session_info['port_info'] = channel
            return jsonify({'success': True, 'session': session_info})
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': 'invalid_serial_config'}), 400
        except Exception as exc:  # pylint: disable=broad-except
            return jsonify({'success': False, 'error': str(exc)}), 500

    def api_serial_capture_stop_v2():
        if not is_logged_in():
            return jsonify({'success': False, 'error': 'not_logged_in'}), 401

        payload = request.json or {}
        channel_id = (payload.get('channel_id') or payload.get('port_id') or '').strip()
        if not _is_safe_port_id(channel_id):
            return jsonify({'success': False, 'error': 'invalid_channel_id'}), 400
        try:
            _session_info, actions, error = shared_serial_hub.stop_capture(channel_id, reason='manual')
            if error:
                if error == 'capture_not_found':
                    return jsonify({'success': False, 'error': 'capture_not_found'}), 404
                return jsonify({'success': False, 'error': error}), 400
            runner = _serial_action_runner()
            if runner and actions:
                runner(actions)
            _serial_emit_refresh(channel_id)
            return jsonify({'success': True})
        except Exception as exc:  # pylint: disable=broad-except
            return jsonify({'success': False, 'error': str(exc)}), 500

    def api_serial_capture_status_v2():
        if not is_logged_in():
            return jsonify({'success': False, 'error': 'not_logged_in'}), 401
        try:
            sessions = shared_serial_hub.get_capture_sessions()
            for session_info in sessions:
                channel_id = session_info.get('channel_id')
                session_info['port_info'] = shared_serial_hub.get_channel(channel_id) if channel_id else None
            return jsonify({'success': True, 'sessions': sessions})
        except Exception as exc:  # pylint: disable=broad-except
            return jsonify({'success': False, 'error': str(exc)}), 500

    def api_serial_capture_log_v2(port_id):
        if not is_logged_in():
            return jsonify({'success': False, 'error': 'not_logged_in'}), 401
        if not _is_safe_port_id(port_id):
            return jsonify({'success': False, 'error': 'invalid_channel_id'}), 400

        try:
            limit = int(request.args.get('limit', 200))
        except ValueError:
            return jsonify({'success': False, 'error': 'invalid_limit'}), 400

        before_seq_raw = (request.args.get('before_seq') or '').strip()
        before_seq = None
        if before_seq_raw:
            try:
                before_seq = int(before_seq_raw)
            except ValueError:
                return jsonify({'success': False, 'error': 'invalid_before_seq'}), 400

        try:
            entries, next_cursor, has_more, error = shared_serial_hub.get_history(
                channel_id=port_id,
                before_seq=before_seq,
                limit=max(1, min(limit, 1000)),
            )
            if error:
                return jsonify({'success': False, 'error': error}), 404
            return jsonify(
                {
                    'success': True,
                    'entries': entries,
                    'next_cursor': next_cursor,
                    'has_more': has_more,
                }
            )
        except Exception as exc:  # pylint: disable=broad-except
            return jsonify({'success': False, 'error': str(exc)}), 500

    app.view_functions['api_serial_channels'] = api_serial_channels_v2
    app.view_functions['api_serial_capture_start'] = api_serial_capture_start_v2
    app.view_functions['api_serial_capture_stop'] = api_serial_capture_stop_v2
    app.view_functions['api_serial_capture_status'] = api_serial_capture_status_v2
    app.view_functions['api_serial_capture_log'] = api_serial_capture_log_v2

    @app.route('/save_serial_log', methods=['POST'])
    def save_serial_log():
        """保存串口日志到服务器"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        try:
            data = request.json
            log_content = data.get('log_content', '')
            port_name = data.get('port_name', 'unknown')
            
            # 创建 serial_logs 目录
            import os
            from datetime import datetime
            
            if getattr(sys, 'frozen', False):
                base_dir = project_base_dir()
            else:
                base_dir = project_base_dir()
            
            log_dir = os.path.join(base_dir, 'logs', 'serial_logs')
            os.makedirs(log_dir, exist_ok=True)
            
            # 生成日志文件名（精确到毫秒）
            now = datetime.now()
            milliseconds = now.microsecond // 1000
            timestamp = now.strftime(f'%Y-%m-%d_%H-%M-%S.{milliseconds:03d}')
            filename = f"{port_name}_{timestamp}.log"
            filepath = os.path.join(log_dir, filename)
            
            # 保存日志
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(log_content)
            
            return jsonify({
                'success': True,
                'filepath': filepath,
                'filename': filename
            })
        
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    # 获取串口历史记录列表
    @app.route('/get_serial_logs', methods=['GET'])
    def get_serial_logs():
        """获取串口历史记录列表"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        try:
            import os
            
            if getattr(sys, 'frozen', False):
                base_dir = project_base_dir()
            else:
                base_dir = project_base_dir()
            
            log_dir = os.path.join(base_dir, 'logs', 'serial_logs')
            
            if not os.path.exists(log_dir):
                return jsonify({'success': True, 'logs': []})
            
            # 获取所有日志文件
            logs = []
            for filename in os.listdir(log_dir):
                if filename.endswith('.log'):
                    filepath = os.path.join(log_dir, filename)
                    stat = os.stat(filepath)
                    logs.append({
                        'filename': filename,
                        'size': stat.st_size,
                        'modified': stat.st_mtime
                    })
            
            # 按修改时间倒序排序
            logs.sort(key=lambda x: x['modified'], reverse=True)
            
            return jsonify({'success': True, 'logs': logs})
        
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    # 读取串口历史记录
    @app.route('/read_serial_log/<filename>')
    def read_serial_log(filename):
        """读取串口历史记录"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        try:
            import os
            
            if getattr(sys, 'frozen', False):
                base_dir = project_base_dir()
            else:
                base_dir = project_base_dir()
            
            log_dir = os.path.join(base_dir, 'logs', 'serial_logs')
            filepath = os.path.join(log_dir, filename)
            
            # 安全检查
            if not os.path.abspath(filepath).startswith(os.path.abspath(log_dir)):
                return jsonify({'success': False, 'error': '非法路径'}), 403
            
            if not os.path.exists(filepath):
                return jsonify({'success': False, 'error': '文件不存在'}), 404
            
            # 读取日志
            with open(filepath, 'r', encoding='utf-8') as f:
                content = f.read()
            
            return jsonify({
                'success': True,
                'filename': filename,
                'content': content
            })
        
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    # 保存快捷指令
    @app.route('/save_serial_commands', methods=['POST'])
    def save_serial_commands():
        """保存快捷指令到服务器"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        try:
            import json
            import os
            
            data = request.json
            commands = data.get('commands', [])
            
            if getattr(sys, 'frozen', False):
                base_dir = project_base_dir()
            else:
                base_dir = project_base_dir()
            
            config_dir = os.path.join(base_dir, 'logs')
            os.makedirs(config_dir, exist_ok=True)
            
            filepath = os.path.join(config_dir, 'serial_commands.json')
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(commands, f, ensure_ascii=False, indent=2)
            
            return jsonify({'success': True})
        
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    # 加载快捷指令
    @app.route('/load_serial_commands', methods=['GET'])
    def load_serial_commands():
        """从服务器加载快捷指令"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        try:
            import json
            import os
            
            if getattr(sys, 'frozen', False):
                base_dir = project_base_dir()
            else:
                base_dir = project_base_dir()
            
            filepath = os.path.join(base_dir, 'logs', 'serial_commands.json')
            
            if not os.path.exists(filepath):
                return jsonify({'success': True, 'commands': []})
            
            with open(filepath, 'r', encoding='utf-8') as f:
                commands = json.load(f)
            
            return jsonify({'success': True, 'commands': commands})
        
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    # 串口诊断工具
    @app.route('/serial_diagnostic')
    def serial_diagnostic():
        """串口诊断工具页面"""
        if not is_logged_in():
            return redirect(url_for('login'))
        serial_server_config = get_serial_server_config('/serial_diagnostic')
        if (
            serial_server_config['remote_serial_enabled']
            and serial_server_config['compat_mode']
            and serial_server_config['serial_https_active']
            and not request.is_secure
        ):
            return redirect(serial_server_config['serial_https_url'])
        return render_template(
            'serial_diagnostic.html',
            serial_server_config=serial_server_config
        )
    
    # Draw.io主编辑器页面（带保存功能）
    @app.route('/drawio_main')
    def drawio_main():
        """Draw.io主编辑器页面，支持保存和上传"""
        if not is_logged_in():
            return redirect(url_for('login'))
        
        # 获取文件路径参数（如果从文件浏览器打开）
        filepath = request.args.get('filepath', '')
        diagram_content = ''
        
        _routes_logger.debug("drawio_main filepath=%r", filepath)
        
        if filepath:
            # 从服务器加载文件
            root_dir = current_app.config.get('ROOT_DIR')
            _routes_logger.debug("drawio_main ROOT_DIR=%r", root_dir)
            
            # 处理路径
            full_path = os.path.join(root_dir, filepath.lstrip('/'))
            _routes_logger.debug("drawio_main full_path=%r", full_path)
            
            try:
                full_path = os.path.normpath(full_path)
                _routes_logger.debug("drawio_main normpath=%r exists=%s", full_path, os.path.exists(full_path))
                
                if full_path.startswith(os.path.normpath(root_dir)) and os.path.exists(full_path):
                    with open(full_path, 'r', encoding='utf-8') as f:
                        diagram_content = f.read()
                    _routes_logger.debug("drawio_main read ok len=%s preview100=%r has_mxfile=%s", len(diagram_content), (diagram_content[:100] if diagram_content else ''), '<mxfile' in diagram_content)
                else:
                    _routes_logger.debug("drawio_main path invalid or missing")
            except Exception as e:
                _routes_logger.error("drawio_main load failed: %s", e, exc_info=True)
        
        return render_template('drawio_main.html', filepath=filepath, diagram_content=diagram_content)
    
    # 主draw.io路由 - 与原始实现兼容
    @app.route('/drawio')
    def drawio():
        """Draw.io编辑器主页面，与原始实现兼容"""
        if not is_logged_in():
            return redirect(url_for('login'))
        
        # 检查draw.io文件是否存在
        import os
        # 使用正确的路径获取方式，支持打包环境
        if getattr(sys, 'frozen', False):
            # 打包环境：static在exe所在目录
            base_dir = project_base_dir()
            drawio_dir = os.path.join(base_dir, 'static', 'drawio')
        else:
            # 开发环境
            drawio_dir = project_path('static', 'drawio')
        
        drawio_index = os.path.join(drawio_dir, 'index.html')
        
        # 构建离线模式URL参数
        # offline=1: 离线模式
        # stealth=1: 隐身模式，禁用追踪和外部服务
        # local=1: 只允许本地存储
        # sync=none: 禁用同步功能
        # lang=zh: 中文界面
        offline_params = '?offline=1&stealth=1&local=1&sync=none&mode=device&lang=zh'
        
        if not os.path.exists(drawio_index):
            # 如果没有index.html，提供一个简单的HTML页面来加载编辑器
            html_content = f'''
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>draw.io 编辑器</title>
    <style>
        body, html {{ margin: 0; padding: 0; height: 100%; overflow: hidden; }}
        iframe {{ width: 100%; height: 100%; border: none; }}
    </style>
</head>
<body>
    <iframe src="/drawio/{offline_params}" allowfullscreen></iframe>
</body>
</html>
            '''
            response = make_response(html_content)
            response.headers['Content-Type'] = 'text/html; charset=utf-8'
            return response
        
        # 读取并返回index.html，并注入离线模式配置
        try:
            with open(drawio_index, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # 在index.html中注入离线模式配置
            # 在head标签中添加配置脚本
            config_script = '''
    <script>
        // 配置drawio为离线模式
        window.DRAWIO_BASE_URL = window.location.origin;
        window.DRAWIO_CONFIG = {
            defaultLibraries: 'general',
            enableCustomLibraries: false,
            enabledLibraries: ['general', 'uml', 'entity', 'mockup', 'flowchart'],
            plugins: [],
            // 禁用云存储选项
            mode: 'device',
            offline: '1',
            stealth: '1',
            local: '1'
        };
    </script>
'''
            # 在</head>之前插入配置
            if '</head>' in content:
                content = content.replace('</head>', config_script + '</head>')
            
            response = make_response(content)
            response.headers['Content-Type'] = 'text/html; charset=utf-8'
            return response
        except Exception as e:
            _routes_logger.error("Error loading Draw.io index: %s", e)
            return "Error loading Draw.io editor", 500
    
    # 重新添加编辑功能路由，但使用正确的iframe src
    @app.route('/drawio_edit')
    def drawio_edit():
        """编辑Draw.io图表"""
        if not is_logged_in():
            return redirect(url_for('login'))
        
        filepath = request.args.get('filepath')
        if not filepath:
            # 不再使用不存在的error.html模板
            return make_response("文件路径不能为空", 400)
        
        # 安全检查：验证文件路径
        root_dir = current_app.config.get('ROOT_DIR')
        full_path = os.path.join(root_dir, filepath.lstrip('/'))
        
        # 确保路径在根目录内
        try:
            full_path = os.path.normpath(full_path)
            if not full_path.startswith(os.path.normpath(root_dir)):
                return make_response("无效的文件路径", 400)
        except Exception:
            return make_response("无效的文件路径", 400)
        
        # 读取现有图表内容（如果存在）
        diagram_content = ''
        if os.path.exists(full_path):
            try:
                with open(full_path, 'r', encoding='utf-8') as f:
                    diagram_content = f.read()
            except Exception:
                diagram_content = ''
        
        return render_template('drawio_edit.html', filepath=filepath, diagram_content=diagram_content)
    
    @app.route('/drawio_save', methods=['POST'])
    def drawio_save():
        """保存draw.io文件"""
        if not is_logged_in():
            return jsonify({'error': '请先登录'}), 401
        
        data = request.get_json()
        filepath = data.get('filepath')
        content = data.get('content')
        
        if not filepath or content is None:
            return jsonify({'error': '文件路径或内容不能为空'}), 400
        
        root_dir = current_app.config.get('ROOT_DIR')
        if not root_dir:
            root_dir = os.getcwd()
        
        # 安全检查
        try:
            full_path = os.path.join(root_dir, filepath)
            full_path = os.path.normpath(full_path)
            if not full_path.startswith(os.path.normpath(root_dir)):
                return jsonify({'error': '访问被拒绝'}), 403
        except Exception:
            return jsonify({'error': '路径解析错误'}), 400
        
        # 确保目录存在，如果不存在则创建
        directory = os.path.dirname(full_path)
        if not os.path.exists(directory):
            try:
                os.makedirs(directory, exist_ok=True)
                _routes_logger.info("创建目录: %s", directory)
            except Exception as e:
                _routes_logger.error("创建目录失败: %s", e)
                return jsonify({'error': f'创建目录失败: {e}'}), 400
        
        # 保存文件
        try:
            with open(full_path, 'w', encoding='utf-8') as f:
                f.write(content)
            return jsonify({'success': True, 'message': '文件保存成功'})
        except Exception as e:
            return jsonify({'error': f'保存文件失败: {e}'}), 500
    
    
    @app.route('/drawio_embed')
    def drawio_embed():
        """嵌入draw.io编辑器用于预览"""
        # 注意：此路由用于分享链接预览，不需要登录检查
        # 登录检查在分享链接路由中已完成
        
        filepath = request.args.get('filepath')
        readonly = request.args.get('readonly', '0')  # 是否只读模式
        
        if not filepath:
            return make_response("文件路径不能为空", 400)
        
        root_dir = current_app.config.get('ROOT_DIR')
        if not root_dir:
            root_dir = os.getcwd()
        
        # 安全检查
        try:
            full_path = os.path.join(root_dir, filepath)
            full_path = os.path.normpath(full_path)
            if not full_path.startswith(os.path.normpath(root_dir)):
                return make_response("访问被拒绝", 403)
        except Exception:
            return make_response("路径解析错误", 400)
        
        # 读取draw.io文件内容
        if os.path.exists(full_path) and not os.path.isdir(full_path):
            try:
                with open(full_path, 'r', encoding='utf-8') as f:
                    diagram_content = f.read()
            except Exception as e:
                return make_response(f'读取文件失败: {e}', 500)
        else:
            diagram_content = ''
        
        return render_template('drawio_embed.html', 
                              filepath=filepath, 
                              diagram_content=diagram_content,
                              readonly=(readonly == '1'))







    # Draw.io相关静态文件路由
    @app.route('/drawio/<path:filename>')
    def drawio_static(filename):
        """提供Draw.io静态文件，与原始实现一致"""
        # 处理空filename的情况（访问 /drawio/ 时）
        if not filename or filename == '':
            filename = 'index.html'
        
        # 安全检查：防止目录遍历
        if '..' in filename or '//' in filename or '\\' in filename:
            return make_response("访问被拒绝", 403)
        
        import os
        # 使用get_resource_path确保在打包环境下正确找到文件
        if getattr(sys, 'frozen', False):
            # 打包环境：static在exe所在目录
            base_dir = project_base_dir()
            drawio_dir = os.path.join(base_dir, 'static', 'drawio')
        else:
            # 开发环境
            drawio_dir = project_path('static', 'drawio')
        
        try:
            return send_from_directory(drawio_dir, filename)
        except FileNotFoundError:
            return make_response(f"文件未找到: {filename}", 404)
    
    # Draw.io根路径资源处理
    @app.route('/styles/<path:filename>')
    @app.route('/js/<path:filename>')
    @app.route('/images/<path:filename>')
    @app.route('/libs/<path:filename>')
    @app.route('/resources/<path:filename>')
    @app.route('/mxgraph/<path:filename>')
    @app.route('/math/<path:filename>')
    @app.route('/plugins/<path:filename>')
    @app.route('/shapes/<path:filename>')
    @app.route('/stencils/<path:filename>')
    @app.route('/templates/<path:filename>')
    @app.route('/connect/<path:filename>')
    @app.route('/newDiagramCats/<path:filename>')
    def drawio_root_resources(filename):
        """处理Draw.io根路径资源请求，与原始实现一致"""
        # 获取请求的路径
        path = request.path.lstrip('/')
        
        # 安全检查：防止目录遍历
        if '..' in path or '//' in path or '\\' in path:
            return make_response("访问被拒绝", 403)
        
        import os
        import sys
        # 使用正确的路径获取方式，支持打包环境
        if getattr(sys, 'frozen', False):
            # 打包环境：static在exe所在目录
            base_dir = project_base_dir()
            drawio_dir = os.path.join(base_dir, 'static', 'drawio')
        else:
            # 开发环境
            drawio_dir = project_path('static', 'drawio')
        
        try:
            return send_from_directory(drawio_dir, path)
        except FileNotFoundError:
            return make_response(f"资源未找到: {path}", 404)
    
    # Service Worker脚本路由
    @app.route('/service-worker.js')
    def service_worker():
        """提供Service Worker脚本，与原始实现一致"""
        try:
            import os
            import sys
            # 使用正确的路径获取方式，支持打包环境
            if getattr(sys, 'frozen', False):
                # 打包环境：static在exe所在目录
                base_dir = project_base_dir()
                drawio_dir = os.path.join(base_dir, 'static', 'drawio')
            else:
                # 开发环境
                drawio_dir = project_path('static', 'drawio')
            
            response = send_from_directory(drawio_dir, 'service-worker.js')
            response.headers['Content-Type'] = 'application/javascript'
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            return response
        except FileNotFoundError:
            return '', 404
    
    # 处理Draw.io的代理请求（禁用）
    @app.route('/proxy')
    @app.route('/proxt')
    def drawio_proxy():
        """禁用Draw.io的代理功能，返回404"""
        return '', 404

    # =============================
    # 分享链接功能路由
    # =============================
    
    @app.route('/create_share', methods=['POST'])
    def create_share():
        """创建分享链接"""
        if not is_logged_in():
            return jsonify({'error': '请先登录'}), 401
        
        data = request.get_json()
        file_path = data.get('file_path')
        is_directory = data.get('is_directory', False)
        password = data.get('password', '')
        expire_hours = data.get('expire_hours')
        max_visits = data.get('max_visits', -1)
        description = data.get('description', '')
        
        if not file_path:
            return jsonify({'error': '文件路径不能为空'}), 400
        
        # 验证文件是否存在
        root_dir = current_app.config.get('ROOT_DIR')
        full_path = os.path.join(root_dir, file_path)
        
        try:
            full_path = os.path.normpath(full_path)
            if not full_path.startswith(os.path.normpath(root_dir)):
                return jsonify({'error': '访问被拒绝'}), 403
            if not os.path.exists(full_path):
                return jsonify({'error': '文件不存在'}), 404
        except Exception as e:
            return jsonify({'error': f'路径验证失败: {e}'}), 400
        
        # 创建分享链接
        try:
            share_manager = current_app.config['SHARE_MANAGER']
            share_info = share_manager.create_share_link(
                file_path=file_path,
                is_directory=is_directory,
                password=password if password else None,
                expire_hours=int(expire_hours) if expire_hours else None,
                max_visits=int(max_visits) if max_visits else -1,
                created_by=session.get('username', 'anonymous'),
                description=description
            )
            
            # 生成完整的分享链接URL
            share_url = url_for('share_view', share_code=share_info['share_code'], _external=True)
            share_info['share_url'] = share_url
            
            return jsonify({
                'success': True,
                'message': '分享链接创建成功',
                'data': share_info
            })
        except Exception as e:
            return jsonify({'error': f'创建分享链接失败: {e}'}), 500
    
    @app.route('/share/<share_code>')
    def share_view(share_code):
        """访问分享链接"""
        share_manager = current_app.config['SHARE_MANAGER']
        
        # 检查是否需要密码
        link = share_manager.get_share_link(share_code)
        if not link:
            return render_template('share_error.html', 
                                 error='分享链接不存在或已被删除'), 404
        
        # 如果需要密码且未验证，显示密码输入页面
        if link['password'] and not session.get(f'share_verified_{share_code}'):
            return render_template('share_password.html', share_code=share_code)
        
        # 验证分享链接
        result = share_manager.verify_share_link(share_code)
        if not result['valid']:
            return render_template('share_error.html', error=result['reason']), 403
        
        # 记录访问
        client_ip = request.environ.get('REMOTE_ADDR', 'unknown')
        user_agent = request.headers.get('User-Agent', '')
        share_manager.record_visit(share_code, client_ip, user_agent)
        
        # 获取文件信息
        link_data = result['data']
        root_dir = current_app.config.get('ROOT_DIR')
        full_path = os.path.join(root_dir, link_data['file_path'])
        
        # 安全检查
        try:
            full_path = os.path.normpath(full_path)
            if not full_path.startswith(os.path.normpath(root_dir)):
                return render_template('share_error.html', error='访问被拒绝'), 403
        except Exception:
            return render_template('share_error.html', error='路径解析错误'), 400
        
        if not os.path.exists(full_path):
            return render_template('share_error.html', error='文件不存在'), 404
        
        filename = os.path.basename(full_path)
        _, ext = os.path.splitext(filename.lower())
        preview_kind = 'unsupported'

        if not link_data['is_directory']:
            if ext in DRAWIO_EXTENSIONS:
                preview_kind = 'drawio'
            elif ext in MARKDOWN_EXTENSIONS:
                preview_kind = 'markdown'
            elif ext in IMAGE_EXTENSIONS:
                preview_kind = 'image'
            elif ext == '.pdf':
                preview_kind = 'pdf'
            elif ext in VIDEO_EXTENSIONS:
                preview_kind = 'video'
            elif ext in AUDIO_EXTENSIONS:
                preview_kind = 'audio'

        preview_supported = preview_kind != 'unsupported'

        # 渲染分享页面
        return render_template('share_view.html',
                             share_code=share_code,
                             link_data=link_data,
                             file_path=link_data['file_path'],
                             is_directory=link_data['is_directory'],
                             filename=filename,
                             download_available=not link_data['is_directory'],
                             preview_supported=preview_supported,
                             preview_kind=preview_kind)
    
    @app.route('/share/<share_code>/verify', methods=['POST'])
    def share_verify_password(share_code):
        """验证分享链接密码"""
        data = request.get_json()
        password = data.get('password', '')
        
        share_manager = current_app.config['SHARE_MANAGER']
        result = share_manager.verify_share_link(share_code, password)
        
        if result['valid']:
            # 标记为已验证
            session[f'share_verified_{share_code}'] = True
            return jsonify({'success': True, 'message': '验证成功'})
        else:
            return jsonify({'success': False, 'message': result['reason']}), 403
    
    @app.route('/share/<share_code>/download')
    def share_download(share_code):
        """下载分享的文件"""
        share_manager = current_app.config['SHARE_MANAGER']
        
        # 检查是否已验证密码
        link = share_manager.get_share_link(share_code)
        if not link:
            abort(404)
        
        if link['password'] and not session.get(f'share_verified_{share_code}'):
            abort(403)
        
        # 验证分享链接
        result = share_manager.verify_share_link(share_code)
        if not result['valid']:
            abort(403)
        
        # 获取文件
        link_data = result['data']
        root_dir = current_app.config.get('ROOT_DIR')
        full_path = os.path.join(root_dir, link_data['file_path'])
        
        try:
            full_path = os.path.normpath(full_path)
            if not full_path.startswith(os.path.normpath(root_dir)):
                abort(403)
        except Exception:
            abort(400)
        
        if not os.path.exists(full_path) or os.path.isdir(full_path):
            abort(404)
        
        directory = os.path.dirname(full_path)
        filename = os.path.basename(full_path)
        
        return send_from_directory(directory, filename, as_attachment=True)
    
    @app.route('/share/<share_code>/preview')
    def share_preview(share_code):
        """预览分享的文件"""
        share_manager = current_app.config['SHARE_MANAGER']
        
        # 检查是否已验证密码
        link = share_manager.get_share_link(share_code)
        if not link:
            abort(404)
        
        if link['password'] and not session.get(f'share_verified_{share_code}'):
            abort(403)
        
        # 验证分享链接
        result = share_manager.verify_share_link(share_code)
        if not result['valid']:
            abort(403)
        
        # 获取文件
        link_data = result['data']
        root_dir = current_app.config.get('ROOT_DIR')
        full_path = os.path.join(root_dir, link_data['file_path'])
        
        try:
            full_path = os.path.normpath(full_path)
            if not full_path.startswith(os.path.normpath(root_dir)):
                abort(403)
        except Exception:
            abort(400)
        
        if not os.path.exists(full_path) or os.path.isdir(full_path):
            abort(404)
        
        directory = os.path.dirname(full_path)
        filename = os.path.basename(full_path)
        
        # 根据文件扩展名设置MIME类型或渲染预览页面
        _, ext = os.path.splitext(filename.lower())
        
        # 对于drawio文件，返回预览页面而不是文件本身
        if ext in DRAWIO_EXTENSIONS:
            # 读取drawio文件内容
            try:
                with open(full_path, 'r', encoding='utf-8') as f:
                    diagram_content = f.read()
            except Exception as e:
                diagram_content = ''
            
            # 返回drawio预览页面
            return render_template('share_drawio_preview.html',
                                 share_code=share_code,
                                 filename=filename,
                                 diagram_content=diagram_content,
                                 file_path=link_data['file_path'])
        
        # 对于Markdown文件，返回渲染后的HTML
        elif ext in MARKDOWN_EXTENSIONS:
            try:
                with open(full_path, 'r', encoding='utf-8') as f:
                    markdown_content = f.read()
                html_content = render_markdown_content(markdown_content, link_data['file_path'])
            except Exception as e:
                html_content = f'<p>读取文件时出错: {e}</p>'
            
            return render_template('share_markdown_preview.html',
                                 share_code=share_code,
                                 filename=filename,
                                 content=html_content)
        
        # 对于其他文件类型，设置正确的MIME类型并发送
        mimetype = None
        
        if ext in IMAGE_EXTENSIONS:
            if ext in ['.jpg', '.jpeg']:
                mimetype = 'image/jpeg'
            elif ext == '.png':
                mimetype = 'image/png'
            elif ext == '.gif':
                mimetype = 'image/gif'
            elif ext == '.svg':
                mimetype = 'image/svg+xml'
            elif ext == '.webp':
                mimetype = 'image/webp'
            elif ext == '.avif':
                mimetype = 'image/avif'
            elif ext == '.ico':
                mimetype = 'image/x-icon'
        elif ext == '.pdf':
            mimetype = 'application/pdf'
        elif ext in VIDEO_EXTENSIONS:
            if ext == '.mp4':
                mimetype = 'video/mp4'
            elif ext == '.avi':
                mimetype = 'video/x-msvideo'
            elif ext == '.mov':
                mimetype = 'video/quicktime'
            elif ext == '.wmv':
                mimetype = 'video/x-ms-wmv'
            elif ext == '.webm':
                mimetype = 'video/webm'
            elif ext == '.m4v':
                mimetype = 'video/x-m4v'
        elif ext in AUDIO_EXTENSIONS:
            if ext == '.mp3':
                mimetype = 'audio/mpeg'
            elif ext == '.wav':
                mimetype = 'audio/wav'
            elif ext == '.ogg':
                mimetype = 'audio/ogg'
            elif ext == '.flac':
                mimetype = 'audio/flac'
            elif ext == '.wma':
                mimetype = 'audio/x-ms-wma'
            elif ext in ['.m4a', '.aac']:
                mimetype = 'audio/mp4'
            elif ext in ['.opus', '.weba']:
                mimetype = 'audio/ogg'
            elif ext in ['.aiff', '.aif']:
                mimetype = 'audio/aiff'
            elif ext == '.amr':
                mimetype = 'audio/amr'
        
        return send_from_directory(directory, filename, as_attachment=False, mimetype=mimetype)
    
    @app.route('/manage_shares')
    def manage_shares():
        """管理分享链接页面"""
        if not is_logged_in():
            return redirect(url_for('login'))
        
        share_manager = current_app.config['SHARE_MANAGER']
        
        # 清理过期链接
        share_manager.cleanup_expired_links()
        
        # 获取所有分享链接
        shares = share_manager.get_all_share_links()
        
        # 为每个分享链接生成完整URL
        for share in shares:
            share['share_url'] = url_for('share_view', 
                                        share_code=share['share_code'], 
                                        _external=True)
        
        from datetime import datetime
        return render_template('manage_shares.html', shares=shares, now=datetime.now().isoformat())
    
    @app.route('/delete_share/<share_code>', methods=['POST'])
    def delete_share(share_code):
        """删除分享链接"""
        if not is_logged_in():
            return jsonify({'error': '请先登录'}), 401
        
        share_manager = current_app.config['SHARE_MANAGER']
        
        if share_manager.delete_share_link(share_code):
            return jsonify({'success': True, 'message': '分享链接已删除'})
        else:
            return jsonify({'success': False, 'message': '删除失败'}), 500
    
    @app.route('/share_stats/<share_code>')
    def share_stats(share_code):
        """查看分享链接统计"""
        if not is_logged_in():
            return redirect(url_for('login'))
        
        share_manager = current_app.config['SHARE_MANAGER']
        stats = share_manager.get_visit_stats(share_code)
        
        if not stats['link']:
            abort(404)
        
        return render_template('share_stats.html', 
                             share_code=share_code,
                             stats=stats)
    
    @app.route('/get_file_content/<path:filepath>')
    def get_file_content(filepath):
        """获取文件内容用于编辑"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        root_path = current_app.config['ROOT_DIR']
        full_path = os.path.join(root_path, filepath)
        
        # 安全检查
        if not os.path.abspath(full_path).startswith(os.path.abspath(root_path)):
            abort(403)
        
        if not os.path.exists(full_path) or not os.path.isfile(full_path):
            abort(404)
        
        try:
            # 尝试以UTF-8读取
            with open(full_path, 'r', encoding='utf-8') as f:
                content = f.read()
            return content, 200, {'Content-Type': 'text/plain; charset=utf-8'}
        except UnicodeDecodeError:
            # 尝试GBK
            try:
                with open(full_path, 'r', encoding='gbk') as f:
                    content = f.read()
                return content, 200, {'Content-Type': 'text/plain; charset=utf-8'}
            except:
                return '无法读取文件：编码不支持', 400
        except Exception as e:
            return f'读取文件失败: {str(e)}', 500
    
    @app.route('/get_git_file_content', methods=['POST'])
    def get_git_file_content():
        """获取Git仓库中的文件内容用于编辑"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        data = request.get_json()
        filepath = data.get('filepath', '').strip()
        git_root = data.get('git_root', '').strip()
        
        log_message(f"获取Git文件内容: filepath={filepath}, git_root={git_root}")
        
        if not filepath:
            return jsonify({'success': False, 'error': '文件路径不能为空'}), 400
        
        if not git_root:
            return jsonify({'success': False, 'error': 'Git仓库根路径不能为空'}), 400
        
        # 处理Git仓库路径
        git_workdir = current_app.config.get('GIT_WORKDIR', '')
        if not git_workdir:
            log_message("错误: Git工作目录未设置", 'ERROR')
            return jsonify({'success': False, 'error': 'Git工作目录未设置'}), 400
        
        # 标准化路径（统一使用正斜杠，避免\t等转义字符问题）
        # 先替换反斜杠为正斜杠，避免\t被解释为制表符
        git_root_normalized = git_root.replace('\\', '/')
        filepath_normalized = filepath.replace('\\', '/')
        
        log_message(f"标准化后的路径: git_root={git_root_normalized}, filepath={filepath_normalized}")
        
        # 处理git_root路径
        if git_root_normalized.startswith('/') or (len(git_root_normalized) > 1 and git_root_normalized[1] == ':'):
            # 绝对路径（Unix风格或Windows风格）
            # Windows路径如 F:/Work/GitDict/test
            if git_root_normalized[1] == ':':
                # Windows绝对路径，转换为系统路径格式
                repo_root = os.path.normpath(git_root_normalized.replace('/', '\\'))
            else:
                # Unix绝对路径
                repo_root = os.path.normpath(git_root_normalized)
        else:
            # 相对路径，相对于Git工作目录
            repo_root = os.path.normpath(os.path.join(git_workdir, git_root_normalized.replace('/', os.sep)))
        
        git_workdir_norm = os.path.normpath(git_workdir)
        if not repo_root.startswith(git_workdir_norm):
            log_message(f"错误: Git仓库路径不在工作目录内: {repo_root} (工作目录: {git_workdir_norm})", 'ERROR')
            return jsonify({'success': False, 'error': f'Git仓库路径不在工作目录内: {repo_root}'}), 403
        
        # 构建完整文件路径
        # filepath可能是相对路径（相对于仓库根）
        if filepath_normalized.startswith('/') or (len(filepath_normalized) > 1 and filepath_normalized[1] == ':'):
            # 绝对路径
            if filepath_normalized[1] == ':':
                # Windows绝对路径
                filepath_norm = os.path.normpath(filepath_normalized.replace('/', '\\'))
            else:
                # Unix绝对路径
                filepath_norm = os.path.normpath(filepath_normalized)
            
            if not filepath_norm.startswith(repo_root):
                log_message(f"错误: 文件路径不在仓库内: {filepath_norm} (仓库: {repo_root})", 'ERROR')
                return jsonify({'success': False, 'error': '访问被拒绝：文件路径不在仓库内'}), 403
            full_path = filepath_norm
        else:
            # 相对路径，相对于仓库根
            full_path = os.path.normpath(os.path.join(repo_root, filepath_normalized.replace('/', os.sep)))
        
        # 再次验证路径在仓库内（防止路径遍历）
        if not full_path.startswith(repo_root):
            log_message(f"错误: 路径遍历检测失败: {full_path} (仓库根: {repo_root})", 'ERROR')
            return jsonify({'success': False, 'error': '访问被拒绝：路径遍历检测'}), 403
        
        log_message(f"解析后的文件路径: {full_path}")
        
        if not os.path.exists(full_path):
            log_message(f"错误: 文件不存在: {full_path}", 'ERROR')
            return jsonify({'success': False, 'error': f'文件不存在: {full_path}'}), 404
        
        if not os.path.isfile(full_path):
            log_message(f"错误: 路径不是文件: {full_path}", 'ERROR')
            return jsonify({'success': False, 'error': f'路径不是文件: {full_path}'}), 400
        
        try:
            # 尝试以UTF-8读取
            with open(full_path, 'r', encoding='utf-8') as f:
                content = f.read()
            # 返回纯文本内容，不是JSON
            return Response(content, mimetype='text/plain; charset=utf-8')
        except UnicodeDecodeError:
            # 尝试GBK
            try:
                with open(full_path, 'r', encoding='gbk') as f:
                    content = f.read()
                return Response(content, mimetype='text/plain; charset=utf-8')
            except Exception as e:
                return jsonify({'success': False, 'error': f'无法读取文件：编码不支持 ({str(e)})'}), 400
        except Exception as e:
            import traceback
            log_message(f"读取Git文件失败: {str(e)}\n{traceback.format_exc()}", 'ERROR')
            return jsonify({'success': False, 'error': f'读取文件失败: {str(e)}'}), 500
    
    @app.route('/save_git_file', methods=['POST'])
    def save_git_file():
        """保存Git仓库中的文件内容"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        data = request.get_json()
        filepath = data.get('filepath', '').strip()
        git_root = data.get('git_root', '').strip()
        content = data.get('content')
        
        log_message(f"保存Git文件: filepath={filepath}, git_root={git_root}, content_length={len(content) if content else 0}")
        
        if not filepath:
            return jsonify({'success': False, 'error': '文件路径不能为空'}), 400
        
        if not git_root:
            return jsonify({'success': False, 'error': 'Git仓库根路径不能为空'}), 400
        
        if content is None:
            return jsonify({'success': False, 'error': '文件内容不能为空'}), 400
        
        # 处理Git仓库路径
        git_workdir = current_app.config.get('GIT_WORKDIR', '')
        if not git_workdir:
            return jsonify({'success': False, 'error': 'Git工作目录未设置'}), 400
        
        # 标准化路径（统一使用正斜杠，避免\t等转义字符问题）
        git_root_normalized = git_root.replace('\\', '/')
        filepath_normalized = filepath.replace('\\', '/')
        
        log_message(f"标准化后的路径: git_root={git_root_normalized}, filepath={filepath_normalized}")
        
        # 处理git_root路径
        if git_root_normalized.startswith('/') or (len(git_root_normalized) > 1 and git_root_normalized[1] == ':'):
            # 绝对路径
            if git_root_normalized[1] == ':':
                # Windows绝对路径
                repo_root = os.path.normpath(git_root_normalized.replace('/', '\\'))
            else:
                # Unix绝对路径
                repo_root = os.path.normpath(git_root_normalized)
        else:
            # 相对路径
            repo_root = os.path.normpath(os.path.join(git_workdir, git_root_normalized.replace('/', os.sep)))
        
        git_workdir_norm = os.path.normpath(git_workdir)
        if not repo_root.startswith(git_workdir_norm):
            log_message(f"错误: Git仓库路径不在工作目录内: {repo_root} (工作目录: {git_workdir_norm})", 'ERROR')
            return jsonify({'success': False, 'error': f'Git仓库路径不在工作目录内: {repo_root}'}), 403
        
        # 构建完整文件路径
        if filepath_normalized.startswith('/') or (len(filepath_normalized) > 1 and filepath_normalized[1] == ':'):
            # 绝对路径
            if filepath_normalized[1] == ':':
                full_path = os.path.normpath(filepath_normalized.replace('/', '\\'))
            else:
                full_path = os.path.normpath(filepath_normalized)
            
            if not full_path.startswith(repo_root):
                log_message(f"错误: 文件路径不在仓库内: {full_path} (仓库: {repo_root})", 'ERROR')
                return jsonify({'success': False, 'error': '访问被拒绝：文件路径不在仓库内'}), 403
        else:
            # 相对路径
            full_path = os.path.normpath(os.path.join(repo_root, filepath_normalized.replace('/', os.sep)))
        
        # 再次验证路径在仓库内（防止路径遍历）
        if not full_path.startswith(repo_root):
            log_message(f"错误: 路径遍历检测失败: {full_path} (仓库根: {repo_root})", 'ERROR')
            return jsonify({'success': False, 'error': '访问被拒绝：路径遍历检测'}), 403
        
        log_message(f"解析后的文件路径: {full_path}")
        
        if not os.path.exists(full_path):
            log_message(f"错误: 文件不存在: {full_path}", 'ERROR')
            return jsonify({'success': False, 'error': f'文件不存在: {full_path}'}), 404
        
        if not os.path.isfile(full_path):
            log_message(f"错误: 路径不是文件: {full_path}", 'ERROR')
            return jsonify({'success': False, 'error': f'路径不是文件: {full_path}'}), 400
        
        try:
            # 保存文件（使用UTF-8编码）
            with open(full_path, 'w', encoding='utf-8') as f:
                f.write(content)
            return jsonify({'success': True})
        except Exception as e:
            return jsonify({'success': False, 'error': f'保存文件失败: {str(e)}'}), 500
    
    @app.route('/save_file', methods=['POST'])
    def save_file():
        """保存文件内容"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        try:
            data = request.json
            filepath = data.get('filepath')
            content = data.get('content')
            
            if not filepath:
                return jsonify({'success': False, 'error': '缺少文件路径'}), 400
            
            if content is None:
                return jsonify({'success': False, 'error': '缺少文件内容'}), 400
            
            root_path = current_app.config['ROOT_DIR']
            full_path = os.path.join(root_path, filepath)
            
            # 安全检查：确保路径在允许的目录内
            if not os.path.abspath(full_path).startswith(os.path.abspath(root_path)):
                return jsonify({'success': False, 'error': '无权访问该路径'}), 403
            
            # 确保目录存在
            os.makedirs(os.path.dirname(full_path), exist_ok=True)
            
            # 保存文件
            with open(full_path, 'w', encoding='utf-8') as f:
                f.write(content)
            
            return jsonify({'success': True, 'message': '保存成功'})
            
        except Exception as e:
            return jsonify({'success': False, 'error': f'保存失败: {str(e)}'}), 500
    
    @app.route('/create_file', methods=['POST'])
    def create_file():
        """创建新文件"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        try:
            data = request.json
            filename = data.get('filename')
            path = data.get('path', '')
            content = data.get('content', '')
            
            if not filename:
                return jsonify({'success': False, 'error': '文件名不能为空'}), 400
            
            root_path = current_app.config['ROOT_DIR']
            if path:
                full_path = os.path.join(root_path, path, filename)
            else:
                full_path = os.path.join(root_path, filename)
            
            # 安全检查
            if not os.path.abspath(full_path).startswith(os.path.abspath(root_path)):
                return jsonify({'success': False, 'error': '无权访问该路径'}), 403
            
            # 检查文件是否已存在
            if os.path.exists(full_path):
                return jsonify({'success': False, 'error': '文件已存在'}), 400
            
            # 确保目录存在
            os.makedirs(os.path.dirname(full_path), exist_ok=True)
            
            # 创建文件
            with open(full_path, 'w', encoding='utf-8') as f:
                f.write(content)
            
            return jsonify({'success': True, 'message': '文件创建成功'})
            
        except Exception as e:
            return jsonify({'success': False, 'error': f'创建失败: {str(e)}'}), 500
    
    @app.route('/create_folder', methods=['POST'])
    def create_folder():
        """创建新文件夹"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        try:
            data = request.json
            foldername = data.get('foldername')
            path = data.get('path', '')
            
            if not foldername:
                return jsonify({'success': False, 'error': '文件夹名称不能为空'}), 400
            
            root_path = current_app.config['ROOT_DIR']
            if path:
                full_path = os.path.join(root_path, path, foldername)
            else:
                full_path = os.path.join(root_path, foldername)
            
            # 安全检查
            if not os.path.abspath(full_path).startswith(os.path.abspath(root_path)):
                return jsonify({'success': False, 'error': '无权访问该路径'}), 403
            
            # 检查文件夹是否已存在
            if os.path.exists(full_path):
                return jsonify({'success': False, 'error': '文件夹已存在'}), 400
            
            # 创建文件夹
            os.makedirs(full_path)
            
            return jsonify({'success': True, 'message': '文件夹创建成功'})
            
        except Exception as e:
            return jsonify({'success': False, 'error': f'创建失败: {str(e)}'}), 500
    
    @app.route('/upload_files', methods=['POST'])
    def upload_files():
        """上传文件"""
        return handle_upload_files_request()




