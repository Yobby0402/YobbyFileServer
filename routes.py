# routes.py
import os
import sys
import re
import configparser
# 确保在文件顶部添加必要的导入
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_from_directory, send_file, make_response, abort, current_app
import markdown
from markdown_it import MarkdownIt
from mdit_py_plugins import tasklists, deflist, footnote
from urllib.parse import quote # 导入 quote 用于编码文件名
import posixpath # 用于处理 URL 路径
from share_links import ShareLinkManager  # 导入分享链接管理器
from todo_manager import todo_manager
from serial_manager import serial_manager
from product_compare_manager import product_compare_manager

# 检查用户是否已登录的函数
def is_logged_in():
    """检查用户是否已登录"""
    return 'logged_in' in session

# 创建markdown-it实例，支持多种扩展
def create_markdown_parser():
    """创建配置好的markdown-it解析器"""
    md = MarkdownIt("default", {"breaks": True, "html": True})
    
    # 启用内建规则以支持表格与删除线
    md.enable(["table", "strikethrough"]) 
    
    # 启用插件
    md.use(tasklists.tasklists_plugin)
    md.use(deflist.deflist_plugin)
    md.use(footnote.footnote_plugin)
    
    return md

# 全局markdown解析器实例
markdown_parser = create_markdown_parser()

# 处理图片路径的函数
def process_image_paths(content, current_file_path):
    """处理Markdown内容中的图片路径"""
    # 定义正则表达式匹配Markdown图片语法
    img_pattern = re.compile(r'!\[(.*?)\]\(([^\s\)]+)(?:\s+"([^"]*)")?\)')
    
    def replace_img_path(match):
        alt_text = match.group(1)
        img_path = match.group(2)
        title = match.group(3)
        
        # 跳过已经是绝对URL或以/download或/preview开头的路径
        if img_path.startswith(('http://', 'https://', '/download', '/preview')):
            return match.group(0)
        
        # 处理相对路径
        parent_dir = posixpath.dirname(current_file_path)
        if parent_dir:
            # 如果是相对于当前文件的路径
            new_path = posixpath.join(parent_dir, img_path)
        else:
            # 如果当前在根目录
            new_path = img_path
        
        # 生成预览链接
        preview_url = f'/preview/{new_path}'
        
        # 重新组合图片标记
        if title:
            return f'![{alt_text}]({preview_url} "{title}")'
        else:
            return f'![{alt_text}]({preview_url})'
    
    # 执行替换
    return img_pattern.sub(replace_img_path, content)

# 渲染Markdown内容
def render_markdown_content(content, filepath):
    """使用markdown-it-py渲染Markdown内容"""
    try:
        # 先处理图片路径
        processed_content = process_image_paths(content, filepath)
        
        # 使用markdown-it-py渲染
        html_content = markdown_parser.render(processed_content)
        
        return html_content
    except Exception as e:
        return f"<p>渲染Markdown时出错: {e}</p>"

# 文件类型常量定义
OFFICE_EXTENSIONS = ['.docx', '.xlsx', '.pptx']
IMAGE_EXTENSIONS = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.svg', '.webp']
MARKDOWN_EXTENSIONS = ['.md', '.markdown']
PDF_EXTENSIONS = ['.pdf']
VIDEO_EXTENSIONS = ['.mp4', '.avi', '.mov', '.wmv']
AUDIO_EXTENSIONS = ['.mp3', '.wav', '.flac', '.ogg', '.wma', '.m4a', '.aac']  # 音频文件格式
DRAWIO_EXTENSIONS = ['.drawio', '.diagram', '.dio', '.xml']  # 添加.xml作为draw.io格式
MODEL_3D_EXTENSIONS = ['.gltf', '.glb', '.obj', '.stl', '.fbx', '.step', '.stp']  # 3D模型文件格式

# 修复init_app函数内部的Draw.io路由

def init_app(app):
    """初始化路由"""
    global current_app
    current_app = app
    
    # 初始化分享链接管理器
    share_manager = ShareLinkManager()
    app.config['SHARE_MANAGER'] = share_manager

    # 初始化 ToDo 管理器
    app.config['TODO_MANAGER'] = todo_manager
    
    # 初始化产品对比管理器
    app.config['PRODUCT_COMPARE_MANAGER'] = product_compare_manager

    # 初始化Draw.io静态文件目录（静默检查，不影响运行）
    # Draw.io是可选功能，不存在也不影响文件浏览器功能
    
    def has_todo_access():
        return is_logged_in() or session.get('todo_direct_access')

    @app.route('/')
    def index():
        """首页，显示操作选择界面"""
        # 添加登录验证
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        
        return render_template('choice.html')

    @app.route('/todo')
    def todo_page():
        """ToDo 主界面（需登录）"""
        if not is_logged_in():
            return redirect(url_for('login'))

        session.pop('todo_direct_access', None)
        return render_template('todo.html', direct_access=False)

    @app.route('/todo/direct')
    def todo_direct():
        """ToDo 后门入口（免登录）"""
        session['todo_direct_access'] = True
        return render_template('todo.html', direct_access=True)

    @app.route('/todo/v2')
    def todo_v2_page():
        """新版ToDo界面（需登录）"""
        if not is_logged_in():
            return redirect(url_for('login'))
        session.pop('todo_direct_access', None)
        return render_template('todo_v2.html', direct_access=False)

    @app.route('/todo/v2/direct')
    def todo_v2_direct():
        """新版ToDo后门入口（免登录）"""
        session['todo_direct_access'] = True
        return render_template('todo_v2.html', direct_access=True)

    @app.route('/todo/v2/preview')
    def todo_v2_preview():
        """ToDo表格预览界面（需登录）"""
        if not is_logged_in():
            return redirect(url_for('login'))
        session.pop('todo_direct_access', None)
        return render_template('todo_v2_preview.html', direct_access=False)

    @app.route('/todo/v2/preview/direct')
    def todo_v2_preview_direct():
        """ToDo表格预览界面后门入口（免登录）"""
        session['todo_direct_access'] = True
        return render_template('todo_v2_preview.html', direct_access=True)

    @app.route('/api/todo/items', methods=['GET', 'POST'])
    def todo_collection():
        """ToDo 列表查询与创建"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']

        if request.method == 'GET':
            state = manager.list_todos()
            timeline = manager.get_timeline()
            return jsonify({
                'success': True,
                'todos': state['todos'],
                'projects': state.get('projects', {}),
                'timeline': timeline
            })

        payload = request.json or {}
        try:
            todo, event = manager.create_todo(payload)
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

        timeline = manager.get_timeline()
        return jsonify({
            'success': True,
            'todo': todo,
            'event': event,
            'timeline': timeline,
            'projects': manager.get_projects()
        })

    @app.route('/api/todo/items/<todo_id>', methods=['GET', 'PUT', 'DELETE'])
    def todo_item(todo_id):
        """ToDo 单项查询、更新与删除"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']

        if request.method == 'GET':
            try:
                todo = manager.get_todo(todo_id)
            except ValueError:
                return jsonify({'success': False, 'error': '未找到 ToDo'}), 404
            return jsonify({'success': True, 'todo': todo})

        if request.method == 'PUT':
            payload = request.json or {}
            try:
                todo, event = manager.update_todo(todo_id, payload)
            except ValueError as exc:
                return jsonify({'success': False, 'error': str(exc)}), 404
            except Exception as exc:
                return jsonify({'success': False, 'error': str(exc)}), 400

            timeline = manager.get_timeline()
            response = {
                'success': True,
                'todo': todo,
                'event': event,
                'timeline': timeline,
                'projects': manager.get_projects()
            }
            return jsonify(response)

        # DELETE
        try:
            manager.delete_todo(todo_id)
        except ValueError:
            return jsonify({'success': False, 'error': '未找到 ToDo'}), 404
        timeline = manager.get_timeline()
        return jsonify({
            'success': True,
            'todo_id': todo_id,
            'timeline': timeline,
            'projects': manager.get_projects()
        })

    @app.route('/api/todo/items/<todo_id>/comments', methods=['POST'])
    def todo_comment(todo_id):
        """为 ToDo 添加评论"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']
        data = request.json or {}
        comment_text = data.get('content')

        try:
            todo, event = manager.add_comment_legacy(todo_id, comment_text)
        except ValueError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 500

        timeline = manager.get_timeline()
        return jsonify({
            'success': True,
            'todo': todo,
            'event': event,
            'timeline': timeline,
            'projects': manager.get_projects()
        })
    
    @app.route('/api/todo/events/<event_id>', methods=['DELETE'])
    def todo_event(event_id):
        """删除单个时间轴事件"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']
        try:
            result = manager.delete_event(event_id)
        except ValueError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 404

        timeline = manager.get_timeline()
        return jsonify({
            'success': True,
            'event_id': event_id,
            'event_type': result.get('event_type'),
            'todo': result.get('todo'),
            'timeline': timeline,
            'projects': manager.get_projects()
        })

    # ===== 新版ToDo API（基于项目和任务的层级结构） =====

    @app.route('/api/todo/v2/data', methods=['GET'])
    def todo_v2_data():
        """获取所有项目和任务数据"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']
        try:
            data = manager.list_all()
            overview = manager.get_pending_overview()
            return jsonify({
                'success': True,
                'data': data,
                'overview': overview
            })
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 500

    @app.route('/api/todo/v2/projects', methods=['POST'])
    def todo_v2_create_project():
        """创建新项目"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']
        payload = request.json or {}
        try:
            project = manager.create_project(payload)
            return jsonify({'success': True, 'project': project})
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

    @app.route('/api/todo/v2/projects/<project_id>', methods=['GET', 'PUT', 'DELETE'])
    def todo_v2_project(project_id):
        """项目操作：查询、更新、删除"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']

        if request.method == 'GET':
            try:
                project = manager.get_project(project_id)
                return jsonify({'success': True, 'project': project})
            except ValueError:
                return jsonify({'success': False, 'error': '项目不存在'}), 404

        if request.method == 'PUT':
            payload = request.json or {}
            try:
                project = manager.update_project(project_id, payload)
                return jsonify({'success': True, 'project': project})
            except ValueError:
                return jsonify({'success': False, 'error': '项目不存在'}), 404
            except Exception as exc:
                return jsonify({'success': False, 'error': str(exc)}), 400

        # DELETE
        try:
            project = manager.delete_project(project_id)
            return jsonify({'success': True, 'project': project})
        except ValueError:
            return jsonify({'success': False, 'error': '项目不存在'}), 404

    @app.route('/api/todo/v2/projects/<project_id>/tasks', methods=['POST'])
    def todo_v2_create_task(project_id):
        """在项目中创建新任务"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']
        payload = request.json or {}
        try:
            task, update_record = manager.create_task(project_id, payload)
            return jsonify({'success': True, 'task': task, 'update_record': update_record})
        except ValueError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 404
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

    @app.route('/api/todo/v2/projects/<project_id>/tasks/<task_id>', methods=['GET', 'PUT', 'DELETE'])
    def todo_v2_task(project_id, task_id):
        """任务操作：查询、更新、删除"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']

        if request.method == 'GET':
            try:
                project = manager.get_project(project_id)
                task = next((t for t in project.get('tasks', []) if t.get('id') == task_id), None)
                if not task:
                    return jsonify({'success': False, 'error': '任务不存在'}), 404
                return jsonify({'success': True, 'task': task})
            except ValueError:
                return jsonify({'success': False, 'error': '项目不存在'}), 404

        if request.method == 'PUT':
            payload = request.json or {}
            try:
                task, update_records = manager.update_task(project_id, task_id, payload)
                return jsonify({'success': True, 'task': task, 'update_records': update_records})
            except ValueError as exc:
                return jsonify({'success': False, 'error': str(exc)}), 404
            except Exception as exc:
                return jsonify({'success': False, 'error': str(exc)}), 400

        # DELETE
        try:
            task = manager.delete_task(project_id, task_id)
            return jsonify({'success': True, 'task': task})
        except ValueError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 404

    @app.route('/api/todo/v2/projects/<project_id>/tasks/reorder', methods=['POST'])
    def todo_v2_reorder_tasks(project_id):
        """重新排序任务"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']
        payload = request.json or {}
        task_ids = payload.get('task_ids', [])

        if not isinstance(task_ids, list):
            return jsonify({'success': False, 'error': 'task_ids必须是数组'}), 400

        try:
            project = manager.reorder_tasks(project_id, task_ids)
            return jsonify({'success': True, 'project': project})
        except ValueError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 404

    @app.route('/api/todo/v2/projects/<project_id>/tasks/<task_id>/comments', methods=['POST'])
    def todo_v2_add_comment(project_id, task_id):
        """为任务添加评论"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']
        payload = request.json or {}
        comment_text = payload.get('content', '')

        try:
            task, update_record = manager.add_comment(project_id, task_id, comment_text)
            return jsonify({'success': True, 'task': task, 'update_record': update_record})
        except ValueError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 500

    @app.route('/api/todo/v2/projects/<project_id>/tasks/<task_id>/comments/<comment_id>', methods=['DELETE'])
    def todo_v2_delete_comment(project_id, task_id, comment_id):
        """删除评论"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']
        try:
            task = manager.delete_comment(project_id, task_id, comment_id)
            return jsonify({'success': True, 'task': task})
        except ValueError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 404

    @app.route('/api/todo/v2/overview', methods=['GET'])
    def todo_v2_overview():
        """获取待完成概览"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        manager = current_app.config['TODO_MANAGER']
        try:
            overview = manager.get_pending_overview()
            return jsonify({'success': True, 'overview': overview})
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 500

    @app.route('/api/todo/v2/export/excel', methods=['POST'])
    def todo_v2_export_excel():
        """导出TODO数据为Excel文件"""
        if not has_todo_access():
            return jsonify({'success': False, 'error': '未授权'}), 401

        try:
            from flask import send_file
            from io import BytesIO
            from datetime import datetime
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            config = request.json or {}
            manager = current_app.config['TODO_MANAGER']
            
            # 获取自定义文件名
            custom_file_name = config.get('fileName')
            
            # 检查是否是预览模式
            time_range = config.get('timeRange', 'all')
            if time_range == 'preview':
                # 预览模式：直接使用传入的预览数据
                preview_data = config.get('previewData', [])
                if not preview_data:
                    return jsonify({'success': False, 'error': '预览数据为空'}), 400
                all_tasks = preview_data
            else:
                # 正常模式：从数据库获取数据
                data = manager.list_all()
                
                # 获取所有任务并扁平化
                all_tasks = []
                for project in data.get('projects', []):
                    for task in project.get('tasks', []):
                        task_with_project = {
                            **task,
                            'project_id': project['id'],
                            'project_name': project['name'],
                            'project_color': project.get('color', '#4facfe'),
                        }
                        all_tasks.append(task_with_project)

                # 应用时间范围筛选
            if time_range == 'completed':
                all_tasks = [t for t in all_tasks if (t.get('progress') or 0) >= 100]
            elif time_range == 'pending':
                all_tasks = [t for t in all_tasks if (t.get('progress') or 0) < 100]
            elif time_range == 'custom':
                custom_range = config.get('customTimeRange')
                if custom_range:
                    start_date = datetime.strptime(custom_range['start'], '%Y-%m-%d')
                    end_date = datetime.strptime(custom_range['end'], '%Y-%m-%d')
                    field = custom_range.get('field', 'created')
                    
                    filtered_tasks = []
                    for task in all_tasks:
                        date_str = None
                        if field == 'created':
                            date_str = task.get('created_at')
                        elif field == 'updated':
                            date_str = task.get('updated_at')
                        elif field == 'due':
                            date_str = task.get('due_date')
                        
                        if date_str:
                            try:
                                task_date = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                                if start_date <= task_date <= end_date:
                                    filtered_tasks.append(task)
                            except:
                                pass
                    all_tasks = filtered_tasks

            # 获取列配置
            columns = config.get('columns', [])
            column_order = config.get('columnOrder', columns)
            project_mode = config.get('projectMode', 'single')
            comment_mode = config.get('commentMode', 'inline')
            merge_project_name = config.get('mergeProjectName', False)
            add_global_index = config.get('addGlobalIndex', False)

            # 列定义映射
            column_definitions = {
                'project_name': '项目名称',
                'index': '序号',
                'summary': '任务简述',
                'description': '详细描述',
                'priority': '优先级',
                'progress': '进度',
                'due_date': '预计完成时间',
                'created_at': '创建时间',
                'updated_at': '最后更新时间',
                'change_count': '改动数量',
            }

            # 创建Excel工作簿
            wb = openpyxl.Workbook()
            
            # 删除默认sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            if project_mode == 'single':
                # 方式一：同一Sheet
                ws = wb.create_sheet('任务列表', 0)
                
                # 设置表头
                headers = []
                header_col_map = {}  # 记录表头列索引到column_order的映射
                col_idx = 0
                
                # 如果启用总序号，在第一列添加总序号
                if add_global_index:
                    headers.append('总序号')
                    header_col_map[col_idx] = 'global_index'
                    col_idx += 1
                
                if 'project_name' in column_order:
                    headers.append('项目名称')
                    header_col_map[col_idx] = 'project_name'
                    col_idx += 1
                
                for col in column_order:
                    if col == 'project_name':
                        continue
                    if col in column_definitions:
                        # 如果是评论数且使用内联模式，表头显示"评论"而不是"评论数"
                        if col == 'change_count' and comment_mode == 'inline':
                            headers.append('评论')
                        else:
                            headers.append(column_definitions[col])
                        header_col_map[col_idx] = col
                        col_idx += 1
                
                # 如果使用sheet模式且没有change_count列，添加评论列用于超链接
                if comment_mode == 'sheet' and 'change_count' not in column_order:
                    headers.append('评论')
                    header_col_map[col_idx] = 'comment'
                    col_idx += 1
                elif comment_mode == 'inline' and 'change_count' not in column_order:
                    headers.append('评论')
                    header_col_map[col_idx] = 'comment'
                    col_idx += 1
                
                ws.append(headers)
                
                # 设置表头样式
                header_fill = PatternFill(start_color='D6F2EA', end_color='C8ECE0', fill_type='solid')
                header_font = Font(bold=True, color='1E5F4F', size=11)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border

                # 添加数据行
                row_num = 2
                comment_sheet_rows = []  # 用于方式二的评论数据
                comment_row_counter = 2  # 评论Sheet的行号计数器（从2开始，因为第1行是表头）
                
                # 定义交替行颜色
                even_row_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                odd_row_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                cell_border = Border(
                    left=Side(style='thin', color='E0E0E0'),
                    right=Side(style='thin', color='E0E0E0'),
                    top=Side(style='thin', color='E0E0E0'),
                    bottom=Side(style='thin', color='E0E0E0')
                )
                
                # 用于记录项目名称合并信息
                project_name_merge_ranges = []  # 存储需要合并的范围 [(start_row, end_row, col_idx)]
                current_project_name = None
                project_name_start_row = None
                project_name_col_idx = None
                
                # 找到项目名称列的索引
                if 'project_name' in column_order:
                    for header_idx, col_key in header_col_map.items():
                        if col_key == 'project_name':
                            project_name_col_idx = header_idx + 1  # Excel列索引从1开始
                            break
                
                # 计算每个任务在其项目中的序号
                project_task_counters = {}  # {project_id: counter}
                
                for global_idx, task in enumerate(all_tasks, 1):
                    row_data = []
                    comments = task.get('comments', [])
                    is_even = (row_num % 2 == 0)
                    row_fill = even_row_fill if is_even else odd_row_fill
                    
                    task_project_name = task.get('project_name', '--')
                    
                    # 处理项目名称合并逻辑
                    if merge_project_name and project_name_col_idx:
                        if task_project_name != current_project_name:
                            # 项目名称变化，如果之前有连续的项目名称，记录合并范围
                            if current_project_name is not None and project_name_start_row is not None:
                                if row_num - 1 > project_name_start_row:
                                    # 有多个连续行，需要合并
                                    project_name_merge_ranges.append((project_name_start_row, row_num - 1, project_name_col_idx))
                                # 重置
                            current_project_name = task_project_name
                            project_name_start_row = row_num
                    
                    # 按照表头顺序构建数据行
                    for header_idx in range(len(headers)):
                        col_key = header_col_map.get(header_idx)
                        
                        if col_key == 'global_index':
                            row_data.append(global_idx)
                        elif col_key == 'project_name':
                            row_data.append(task_project_name)
                        elif col_key == 'index':
                            # 使用任务在其项目中的序号
                            project_id = task.get('project_id', '')
                            if project_id not in project_task_counters:
                                project_task_counters[project_id] = 0
                            project_task_counters[project_id] += 1
                            row_data.append(project_task_counters[project_id])
                        elif col_key == 'summary':
                            row_data.append(task.get('summary', '--'))
                        elif col_key == 'description':
                            row_data.append(task.get('description', '--'))
                        elif col_key == 'priority':
                            row_data.append(task.get('priority', 3))
                        elif col_key == 'progress':
                            row_data.append(f"{task.get('progress', 0)}%")
                        elif col_key == 'due_date':
                            due_date = task.get('due_date')
                            row_data.append(due_date if due_date else '--')
                        elif col_key == 'created_at':
                            created = task.get('created_at', '')
                            if created:
                                try:
                                    dt = datetime.fromisoformat(created.replace('Z', '+00:00'))
                                    row_data.append(dt.strftime('%Y-%m-%d %H:%M'))
                                except:
                                    row_data.append(created)
                            else:
                                row_data.append('--')
                        elif col_key == 'updated_at':
                            updated = task.get('updated_at', '')
                            if updated:
                                try:
                                    dt = datetime.fromisoformat(updated.replace('Z', '+00:00'))
                                    row_data.append(dt.strftime('%Y-%m-%d %H:%M'))
                                except:
                                    row_data.append(updated)
                            else:
                                row_data.append('--')
                        elif col_key == 'change_count':
                            # 改动数量：显示修改次数和评论数
                            update_history = task.get('update_history', [])
                            update_count = len([r for r in update_history if r.get('field') != 'comment'])
                            comment_count = len(comments)
                            row_data.append(f"{update_count}次修改 {comment_count}条评论")
                        elif col_key == 'comment':
                            # 额外的评论列
                            if comment_mode == 'inline':
                                # 内联模式：显示评论内容
                                if comments:
                                    comment_text = '\n'.join([
                                        f"[{datetime.fromisoformat(c.get('timestamp', '').replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M')}] {c.get('content', '')}"
                                        for c in comments
                                    ])
                                    row_data.append(comment_text)
                                else:
                                    row_data.append('--')
                            else:
                                # sheet模式：显示超链接文本（超链接会在后面添加）
                                if comments:
                                    row_data.append(f"查看评论({len(comments)})")
                                else:
                                    row_data.append('--')
                        else:
                            row_data.append('--')
                    
                    # 收集评论数据（方式二）
                    if comment_mode == 'sheet':
                        for comment in comments:
                            comment_sheet_rows.append({
                                'project_name': task.get('project_name', '--'),
                                'task_summary': task.get('summary', '--'),
                                'timestamp': comment.get('timestamp', ''),
                                'content': comment.get('content', ''),
                                'task_id': task.get('id', ''),
                                'row_num': comment_row_counter,
                            })
                            comment_row_counter += 1
                    
                    ws.append(row_data)
                    
                    # 设置行样式（交替行颜色和边框）
                    for col_idx in range(1, len(headers) + 1):
                        cell = ws.cell(row=row_num, column=col_idx)
                        cell.fill = row_fill
                        cell.border = cell_border
                        
                        # 设置对齐方式
                        col_key = header_col_map.get(col_idx - 1)
                        if col_key == 'description' or col_key == 'summary':
                            # 任务描述和任务简述：左对齐，上下居中，自动换行
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        elif col_key == 'comment' or col_key == 'change_count':
                            # 评论列：左对齐，上下居中
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        else:
                            # 其他列：居中，上下居中
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 设置评论列自动换行（方式一）
                    if comment_mode == 'inline' and comments:
                        # 找到评论列的索引
                        comment_col_idx = None
                        for header_idx, col_key in header_col_map.items():
                            if col_key == 'change_count' or col_key == 'comment':
                                comment_col_idx = header_idx + 1
                                break
                        if not comment_col_idx and 'change_count' not in column_order:
                            comment_col_idx = len(headers)  # 评论列在最后
                        if comment_col_idx:
                            comment_cell = ws.cell(row=row_num, column=comment_col_idx)
                            # 评论列：左对齐，上下居中，自动换行
                            comment_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    
                    # 添加超链接（方式二）
                    if comment_mode == 'sheet' and comments:
                        # 找到评论列的索引（可能是change_count列或comment列）
                        comment_col_idx = None
                        for header_idx, col_key in header_col_map.items():
                            if col_key == 'change_count' or col_key == 'comment':
                                comment_col_idx = header_idx + 1
                                break
                        if comment_col_idx:
                            comment_cell = ws.cell(row=row_num, column=comment_col_idx)
                            # 计算评论Sheet中第一条评论的行号（表头占第1行，所以从第2行开始）
                            first_comment_row = comment_row_counter - len(comments)
                            if first_comment_row < 2:
                                first_comment_row = 2
                            comment_cell.hyperlink = f"#评论!A{first_comment_row}"
                            comment_cell.font = Font(color='0000FF', underline='single')
                            # 如果单元格已经有值（从row_data添加的），就不需要再设置value
                            if not comment_cell.value:
                                comment_cell.value = f"查看评论({len(comments)})"
                    
                    row_num += 1
                
                # 处理最后一个项目名称的合并
                if merge_project_name and project_name_col_idx and current_project_name is not None and project_name_start_row is not None:
                    if row_num - 1 > project_name_start_row:
                        # 有多个连续行，需要合并
                        project_name_merge_ranges.append((project_name_start_row, row_num - 1, project_name_col_idx))
                
                # 执行项目名称合并
                if merge_project_name and project_name_merge_ranges:
                    for start_row, end_row, col_idx in project_name_merge_ranges:
                        if end_row > start_row:
                            # 合并单元格
                            col_letter = get_column_letter(col_idx)
                            ws.merge_cells(f'{col_letter}{start_row}:{col_letter}{end_row}')
                            # 设置合并后的单元格对齐方式为居中
                            merged_cell = ws[f'{col_letter}{start_row}']
                            merged_cell.alignment = Alignment(horizontal='center', vertical='center')

                # 根据内容自动调整列宽
                def calculate_column_width(ws, col_idx, min_width=10, max_width=80):
                    """根据列内容计算合适的列宽"""
                    col_letter = get_column_letter(col_idx)
                    max_length = 0
                    
                    # 检查表头
                    header_cell = ws.cell(row=1, column=col_idx)
                    if header_cell.value:
                        max_length = max(max_length, len(str(header_cell.value)))
                    
                    # 检查所有数据行
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            if cell.value:
                                # 对于可能换行的列，计算单行最大长度
                                cell_value = str(cell.value)
                                # 如果包含换行符，取最长的一行
                                if '\n' in cell_value:
                                    lines = cell_value.split('\n')
                                    max_line_length = max(len(line) for line in lines)
                                    max_length = max(max_length, max_line_length)
                                else:
                                    max_length = max(max_length, len(cell_value))
                    
                    # 设置列宽，添加一些边距
                    width = min(max(max_length + 2, min_width), max_width)
                    return width
                
                # 调整列宽
                for col_idx, header in enumerate(headers, 1):
                    col_letter = get_column_letter(col_idx)
                    col_key = header_col_map.get(col_idx - 1)
                    
                    # 对于需要换行的列（任务简述、详细描述、评论），设置固定宽度
                    if col_key == 'summary':
                        ws.column_dimensions[col_letter].width = 30  # 任务简述
                    elif col_key == 'description':
                        ws.column_dimensions[col_letter].width = 50  # 详细描述
                    elif col_key == 'comment' or (col_key == 'change_count' and comment_mode == 'inline'):
                        ws.column_dimensions[col_letter].width = 50  # 评论
                    else:
                        # 其他列根据内容自动调整
                        width = calculate_column_width(ws, col_idx, min_width=10, max_width=30)
                        ws.column_dimensions[col_letter].width = width

                # 创建评论Sheet（方式二）
                if comment_mode == 'sheet' and comment_sheet_rows:
                    comment_ws = wb.create_sheet('评论', 1)
                    comment_ws.append(['项目名称', '任务简述', '时间', '评论内容'])
                    
                    # 设置评论表头样式
                    for cell in comment_ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border
                    
                    # 添加评论数据行并设置对齐方式
                    for row_data in comment_sheet_rows:
                        timestamp = row_data['timestamp']
                        if timestamp:
                            try:
                                dt = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
                                timestamp = dt.strftime('%Y-%m-%d %H:%M')
                            except:
                                pass
                        row_num = comment_ws.max_row + 1
                        comment_ws.append([
                            row_data['project_name'],
                            row_data['task_summary'],
                            timestamp,
                            row_data['content'],
                        ])
                        
                        # 设置对齐方式：项目名称和时间居中，任务简述和评论内容左对齐
                        comment_ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center', vertical='center')  # 项目名称
                        comment_ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)  # 任务简述
                        comment_ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center', vertical='center')  # 时间
                        comment_ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)  # 评论内容
                    
                    # 调整评论Sheet列宽
                    comment_ws.column_dimensions['A'].width = 20
                    comment_ws.column_dimensions['B'].width = 30
                    comment_ws.column_dimensions['C'].width = 18
                    comment_ws.column_dimensions['D'].width = 50

            else:
                # 方式二：不同项目不同Sheet
                projects_dict = {}
                for task in all_tasks:
                    project_id = task['project_id']
                    if project_id not in projects_dict:
                        projects_dict[project_id] = {
                            'name': task['project_name'],
                            'tasks': []
                        }
                    projects_dict[project_id]['tasks'].append(task)

                comment_sheet_rows = []
                comment_row_counter = 2  # 评论Sheet的行号计数器
                global_task_counter = 1  # 全局任务计数器（用于总序号）
                
                for project_id, project_data in projects_dict.items():
                    project_name = project_data['name']
                    # Excel sheet名称不能超过31个字符，且不能包含某些特殊字符
                    sheet_name = project_name[:31].replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_')
                    
                    ws = wb.create_sheet(sheet_name, len(wb.sheetnames))
                    
                    # 设置表头
                    headers = []
                    header_col_map = {}  # 记录表头列索引到column_order的映射
                    col_idx = 0
                    
                    # 如果启用总序号，在第一列添加总序号
                    if add_global_index:
                        headers.append('总序号')
                        header_col_map[col_idx] = 'global_index'
                        col_idx += 1
                    
                    for col in column_order:
                        if col in column_definitions:
                            # 如果是评论数且使用内联模式，表头显示"评论"而不是"评论数"
                            if col == 'comment_count' and comment_mode == 'inline':
                                headers.append('评论')
                            else:
                                headers.append(column_definitions[col])
                            header_col_map[col_idx] = col
                            col_idx += 1
                    
                    # 如果使用sheet模式且没有change_count列，添加评论列用于超链接
                    if comment_mode == 'sheet' and 'change_count' not in column_order:
                        headers.append('评论')
                        header_col_map[col_idx] = 'comment'
                        col_idx += 1
                    elif comment_mode == 'inline' and 'change_count' not in column_order:
                        headers.append('评论')
                        header_col_map[col_idx] = 'comment'
                        col_idx += 1
                    
                    ws.append(headers)
                    
                    # 设置表头样式
                    header_fill = PatternFill(start_color='D6F2EA', end_color='C8ECE0', fill_type='solid')
                    header_font = Font(bold=True, color='1E5F4F', size=11)
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border

                    # 添加数据行
                    row_num = 2
                    even_row_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                    odd_row_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                    cell_border = Border(
                        left=Side(style='thin', color='E0E0E0'),
                        right=Side(style='thin', color='E0E0E0'),
                        top=Side(style='thin', color='E0E0E0'),
                        bottom=Side(style='thin', color='E0E0E0')
                    )
                    
                    for idx, task in enumerate(project_data['tasks'], 1):
                        row_data = []
                        comments = task.get('comments', [])
                        is_even = (row_num % 2 == 0)
                        row_fill = even_row_fill if is_even else odd_row_fill
                        
                        # 按照表头顺序构建数据行
                        for header_idx in range(len(headers)):
                            col_key = header_col_map.get(header_idx)
                            
                            if col_key == 'global_index':
                                row_data.append(global_task_counter)
                                global_task_counter += 1
                            elif col_key == 'index':
                                row_data.append(idx)
                            elif col_key == 'summary':
                                row_data.append(task.get('summary', '--'))
                            elif col_key == 'description':
                                row_data.append(task.get('description', '--'))
                            elif col_key == 'priority':
                                row_data.append(task.get('priority', 3))
                            elif col_key == 'progress':
                                row_data.append(f"{task.get('progress', 0)}%")
                            elif col_key == 'due_date':
                                due_date = task.get('due_date')
                                row_data.append(due_date if due_date else '--')
                            elif col_key == 'created_at':
                                created = task.get('created_at', '')
                                if created:
                                    try:
                                        dt = datetime.fromisoformat(created.replace('Z', '+00:00'))
                                        row_data.append(dt.strftime('%Y-%m-%d %H:%M'))
                                    except:
                                        row_data.append(created)
                                else:
                                    row_data.append('--')
                            elif col_key == 'updated_at':
                                updated = task.get('updated_at', '')
                                if updated:
                                    try:
                                        dt = datetime.fromisoformat(updated.replace('Z', '+00:00'))
                                        row_data.append(dt.strftime('%Y-%m-%d %H:%M'))
                                    except:
                                        row_data.append(updated)
                                else:
                                    row_data.append('--')
                            elif col_key == 'change_count':
                                # 改动数量：显示修改次数和评论数
                                update_history = task.get('update_history', [])
                                update_count = len([r for r in update_history if r.get('field') != 'comment'])
                                comment_count = len(comments)
                                row_data.append(f"{update_count}次修改 {comment_count}条评论")
                            elif col_key == 'comment':
                                # 额外的评论列（当comment_count不在column_order中时）
                                if comments:
                                    comment_text = '\n'.join([
                                        f"[{datetime.fromisoformat(c.get('timestamp', '').replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M')}] {c.get('content', '')}"
                                        for c in comments
                                    ])
                                    row_data.append(comment_text)
                                else:
                                    row_data.append('--')
                            else:
                                row_data.append('--')
                        
                        # 收集评论数据（方式二）
                        if comment_mode == 'sheet':
                            for comment in comments:
                                comment_sheet_rows.append({
                                    'project_name': project_name,
                                    'task_summary': task.get('summary', '--'),
                                    'timestamp': comment.get('timestamp', ''),
                                    'content': comment.get('content', ''),
                                    'task_id': task.get('id', ''),
                                    'row_num': comment_row_counter,
                                })
                                comment_row_counter += 1
                        
                        ws.append(row_data)
                        
                        # 设置行样式（交替行颜色和边框）
                        for col_idx in range(1, len(headers) + 1):
                            cell = ws.cell(row=row_num, column=col_idx)
                            cell.fill = row_fill
                            cell.border = cell_border
                            
                            # 设置对齐方式
                            col_key = header_col_map.get(col_idx - 1)
                            if col_key == 'description' or col_key == 'summary':
                                # 任务描述和任务简述：左对齐，上下居中，自动换行
                                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                            elif col_key == 'comment' or col_key == 'change_count':
                                # 评论列：左对齐，上下居中
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                            else:
                                # 其他列：居中，上下居中
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # 设置评论列自动换行（方式一）
                        if comment_mode == 'inline' and comments:
                            # 找到评论列的索引
                            comment_col_idx = None
                            for header_idx, col_key in header_col_map.items():
                                if col_key == 'change_count' or col_key == 'comment':
                                    comment_col_idx = header_idx + 1
                                    break
                            if comment_col_idx:
                                comment_cell = ws.cell(row=row_num, column=comment_col_idx)
                                # 评论列：左对齐，上下居中，自动换行
                                comment_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        
                        # 添加超链接（方式二）
                        if comment_mode == 'sheet' and comments:
                            # 找到评论列的索引（可能是change_count列或comment列）
                            comment_col_idx = None
                            for header_idx, col_key in header_col_map.items():
                                if col_key == 'change_count' or col_key == 'comment':
                                    comment_col_idx = header_idx + 1
                                    break
                            if comment_col_idx:
                                comment_cell = ws.cell(row=row_num, column=comment_col_idx)
                                # 计算评论Sheet中第一条评论的行号（表头占第1行，所以从第2行开始）
                                first_comment_row = comment_row_counter - len(comments)
                                if first_comment_row < 2:
                                    first_comment_row = 2
                                comment_cell.hyperlink = f"#评论!A{first_comment_row}"
                                comment_cell.font = Font(color='0000FF', underline='single')
                                # 如果单元格已经有值（从row_data添加的），就不需要再设置value
                                if not comment_cell.value:
                                    comment_cell.value = f"查看评论({len(comments)})"
                        
                        row_num += 1
                    
                    # 方式二中，如果启用合并项目名称且Sheet中有多个任务，合并项目名称列
                    if merge_project_name and 'project_name' in column_order and row_num > 3:
                        # 找到项目名称列的索引
                        project_name_col_idx = None
                        for header_idx, col_key in header_col_map.items():
                            if col_key == 'project_name':
                                project_name_col_idx = header_idx + 1
                                break
                        if project_name_col_idx:
                            # 合并从第2行（数据开始）到最后一行
                            col_letter = get_column_letter(project_name_col_idx)
                            ws.merge_cells(f'{col_letter}2:{col_letter}{row_num - 1}')
                            # 设置合并后的单元格对齐方式为居中
                            merged_cell = ws[f'{col_letter}2']
                            merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 根据内容自动调整列宽（复用方式一的函数）
                    def calculate_column_width(ws, col_idx, min_width=10, max_width=80):
                        """根据列内容计算合适的列宽"""
                        col_letter = get_column_letter(col_idx)
                        max_length = 0
                        
                        # 检查表头
                        header_cell = ws.cell(row=1, column=col_idx)
                        if header_cell.value:
                            max_length = max(max_length, len(str(header_cell.value)))
                        
                        # 检查所有数据行
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                            for cell in row:
                                if cell.value:
                                    # 对于可能换行的列，计算单行最大长度
                                    cell_value = str(cell.value)
                                    # 如果包含换行符，取最长的一行
                                    if '\n' in cell_value:
                                        lines = cell_value.split('\n')
                                        max_line_length = max(len(line) for line in lines)
                                        max_length = max(max_length, max_line_length)
                                    else:
                                        max_length = max(max_length, len(cell_value))
                        
                        # 设置列宽，添加一些边距
                        width = min(max(max_length + 2, min_width), max_width)
                        return width
                    
                    # 调整列宽
                    for col_idx, header in enumerate(headers, 1):
                        col_letter = get_column_letter(col_idx)
                        col_key = header_col_map.get(col_idx - 1)
                        
                        # 对于需要换行的列（任务简述、详细描述、评论），设置固定宽度
                        if col_key == 'summary':
                            ws.column_dimensions[col_letter].width = 30  # 任务简述
                        elif col_key == 'description':
                            ws.column_dimensions[col_letter].width = 50  # 详细描述
                        elif col_key == 'comment' or (col_key == 'change_count' and comment_mode == 'inline'):
                            ws.column_dimensions[col_letter].width = 50  # 评论
                        else:
                            # 其他列根据内容自动调整
                            width = calculate_column_width(ws, col_idx, min_width=10, max_width=30)
                            ws.column_dimensions[col_letter].width = width

                # 创建评论Sheet（方式二）
                if comment_mode == 'sheet' and comment_sheet_rows:
                    comment_ws = wb.create_sheet('评论', len(wb.sheetnames))
                    comment_ws.append(['项目名称', '任务简述', '时间', '评论内容'])
                    
                    # 设置评论表头样式
                    header_fill = PatternFill(start_color='D6F2EA', end_color='C8ECE0', fill_type='solid')
                    header_font = Font(bold=True, color='1E5F4F', size=11)
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    for cell in comment_ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border
                    
                    # 添加评论数据行并设置对齐方式
                    for row_data in comment_sheet_rows:
                        timestamp = row_data['timestamp']
                        if timestamp:
                            try:
                                dt = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
                                timestamp = dt.strftime('%Y-%m-%d %H:%M')
                            except:
                                pass
                        row_num = comment_ws.max_row + 1
                        comment_ws.append([
                            row_data['project_name'],
                            row_data['task_summary'],
                            timestamp,
                            row_data['content'],
                        ])
                        
                        # 设置对齐方式：项目名称和时间居中，任务简述和评论内容左对齐
                        comment_ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center', vertical='center')  # 项目名称
                        comment_ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)  # 任务简述
                        comment_ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center', vertical='center')  # 时间
                        comment_ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)  # 评论内容
                    
                    # 调整评论Sheet列宽
                    comment_ws.column_dimensions['A'].width = 20
                    comment_ws.column_dimensions['B'].width = 30
                    comment_ws.column_dimensions['C'].width = 18
                    comment_ws.column_dimensions['D'].width = 50

            # 保存到内存
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # 生成文件名
            if custom_file_name:
                filename = custom_file_name
                if not filename.endswith('.xlsx'):
                    filename += '.xlsx'
            else:
                filename = f"todo_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )

        except ImportError:
            return jsonify({'success': False, 'error': 'openpyxl库未安装，请运行: pip install openpyxl'}), 500
        except Exception as exc:
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': str(exc)}), 500

    # ===== 产品对比路由 =====
    
    @app.route('/product_compare')
    def product_compare_page():
        """产品对比主页面"""
        if not is_logged_in():
            return redirect(url_for('login'))
        return render_template('product_compare.html')

    @app.route('/api/product_compare/files', methods=['GET'])
    def product_compare_files():
        """获取所有产品对比文件列表"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未授权'}), 401
        
        manager = current_app.config['PRODUCT_COMPARE_MANAGER']
        try:
            files = manager.list_files()
            return jsonify({'success': True, 'files': files})
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 500

    @app.route('/api/product_compare/files', methods=['POST'])
    def product_compare_create_file():
        """创建新的产品对比文件"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未授权'}), 401
        
        manager = current_app.config['PRODUCT_COMPARE_MANAGER']
        payload = request.json or {}
        name = payload.get('name', '未命名对比')
        
        try:
            file_data = manager.create_file(name)
            return jsonify({'success': True, 'file': file_data})
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

    @app.route('/api/product_compare/files/<file_id>', methods=['GET', 'PUT', 'DELETE'])
    def product_compare_file(file_id):
        """获取、更新或删除产品对比文件"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未授权'}), 401
        
        manager = current_app.config['PRODUCT_COMPARE_MANAGER']
        
        if request.method == 'GET':
            try:
                file_data = manager.get_file(file_id)
                return jsonify({'success': True, 'file': file_data})
            except FileNotFoundError:
                return jsonify({'success': False, 'error': '文件不存在'}), 404
            except Exception as exc:
                return jsonify({'success': False, 'error': str(exc)}), 500
        
        if request.method == 'PUT':
            payload = request.json or {}
            try:
                file_data = manager.update_file(file_id, name=payload.get('name'))
                return jsonify({'success': True, 'file': file_data})
            except FileNotFoundError:
                return jsonify({'success': False, 'error': '文件不存在'}), 404
            except Exception as exc:
                return jsonify({'success': False, 'error': str(exc)}), 400
        
        # DELETE
        try:
            manager.delete_file(file_id)
            return jsonify({'success': True, 'file_id': file_id})
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 500

    @app.route('/api/product_compare/files/<file_id>/attributes', methods=['POST'])
    def product_compare_add_attribute(file_id):
        """添加属性"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未授权'}), 401
        
        manager = current_app.config['PRODUCT_COMPARE_MANAGER']
        payload = request.json or {}
        
        try:
            attr = manager.add_attribute(
                file_id,
                name=payload.get('name'),
                is_common=payload.get('is_common', False),
                attr_type=payload.get('type', 'text')
            )
            # 如果添加成功，更新单位和方向
            if attr and payload.get('unit') is not None or payload.get('direction') is not None:
                attr = manager.update_attribute(
                    file_id,
                    attr['id'],
                    unit=payload.get('unit', ''),
                    direction=payload.get('direction', 'higher')
                )
            file_data = manager.get_file(file_id)
            return jsonify({'success': True, 'attribute': attr, 'file': file_data})
        except FileNotFoundError:
            return jsonify({'success': False, 'error': '文件不存在'}), 404
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

    @app.route('/api/product_compare/files/<file_id>/attributes/<attr_id>', methods=['PUT', 'DELETE'])
    def product_compare_attribute(file_id, attr_id):
        """更新或删除属性"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未授权'}), 401
        
        manager = current_app.config['PRODUCT_COMPARE_MANAGER']
        
        if request.method == 'PUT':
            payload = request.json or {}
            try:
                attr = manager.update_attribute(
                    file_id,
                    attr_id,
                    name=payload.get('name'),
                    is_common=payload.get('is_common'),
                    attr_type=payload.get('type'),
                    order=payload.get('order'),
                    unit=payload.get('unit'),
                    direction=payload.get('direction')
                )
                file_data = manager.get_file(file_id)
                return jsonify({'success': True, 'attribute': attr, 'file': file_data})
            except FileNotFoundError:
                return jsonify({'success': False, 'error': '文件不存在'}), 404
            except Exception as exc:
                return jsonify({'success': False, 'error': str(exc)}), 400
        
        # DELETE
        try:
            manager.delete_attribute(file_id, attr_id)
            file_data = manager.get_file(file_id)
            return jsonify({'success': True, 'attribute_id': attr_id, 'file': file_data})
        except FileNotFoundError:
            return jsonify({'success': False, 'error': '文件不存在'}), 404
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

    @app.route('/api/product_compare/files/<file_id>/attributes/reorder', methods=['POST'])
    def product_compare_reorder_attributes(file_id):
        """重新排序属性"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未授权'}), 401
        
        manager = current_app.config['PRODUCT_COMPARE_MANAGER']
        payload = request.json or {}
        attr_orders = payload.get('orders', [])
        
        try:
            manager.reorder_attributes(file_id, attr_orders)
            file_data = manager.get_file(file_id)
            return jsonify({'success': True, 'file': file_data})
        except FileNotFoundError:
            return jsonify({'success': False, 'error': '文件不存在'}), 404
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

    @app.route('/api/product_compare/files/<file_id>/products', methods=['POST'])
    def product_compare_add_product(file_id):
        """添加产品"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未授权'}), 401
        
        manager = current_app.config['PRODUCT_COMPARE_MANAGER']
        payload = request.json or {}
        
        try:
            product = manager.add_product(
                file_id,
                name=payload.get('name'),
                belonging=payload.get('belonging'),
                attributes=payload.get('attributes', {}),
                link=payload.get('link')
            )
            file_data = manager.get_file(file_id)
            return jsonify({'success': True, 'product': product, 'file': file_data})
        except FileNotFoundError:
            return jsonify({'success': False, 'error': '文件不存在'}), 404
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

    @app.route('/api/product_compare/files/<file_id>/products/<product_id>', methods=['PUT', 'DELETE'])
    def product_compare_product(file_id, product_id):
        """更新或删除产品"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未授权'}), 401
        
        manager = current_app.config['PRODUCT_COMPARE_MANAGER']
        
        if request.method == 'PUT':
            payload = request.json or {}
            try:
                product = manager.update_product(
                    file_id,
                    product_id,
                    name=payload.get('name'),
                    belonging=payload.get('belonging'),
                    attributes=payload.get('attributes'),
                    link=payload.get('link')
                )
                file_data = manager.get_file(file_id)
                return jsonify({'success': True, 'product': product, 'file': file_data})
            except FileNotFoundError:
                return jsonify({'success': False, 'error': '文件不存在'}), 404
            except Exception as exc:
                return jsonify({'success': False, 'error': str(exc)}), 400
        
        # DELETE
        try:
            manager.delete_product(file_id, product_id)
            file_data = manager.get_file(file_id)
            return jsonify({'success': True, 'product_id': product_id, 'file': file_data})
        except FileNotFoundError:
            return jsonify({'success': False, 'error': '文件不存在'}), 404
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

    @app.route('/api/product_compare/files/<file_id>/products/reorder', methods=['POST'])
    def product_compare_reorder_products(file_id):
        """重新排序产品"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未授权'}), 401
        
        manager = current_app.config['PRODUCT_COMPARE_MANAGER']
        payload = request.json or {}
        product_orders = payload.get('orders', [])
        
        try:
            manager.reorder_products(file_id, product_orders)
            file_data = manager.get_file(file_id)
            return jsonify({'success': True, 'file': file_data})
        except FileNotFoundError:
            return jsonify({'success': False, 'error': '文件不存在'}), 404
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

    @app.route('/api/product_compare/files/<file_id>/belongings', methods=['GET'])
    def product_compare_belongings(file_id):
        """获取文件中所有已使用的归属列表（包含颜色）"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未授权'}), 401
        
        manager = current_app.config['PRODUCT_COMPARE_MANAGER']
        try:
            belongings = manager.get_belongings(file_id)
            return jsonify({'success': True, 'belongings': belongings})
        except FileNotFoundError:
            return jsonify({'success': False, 'error': '文件不存在'}), 404
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 500
    
    @app.route('/api/product_compare/files/<file_id>/belongings/<belonging>/color', methods=['PUT'])
    def product_compare_set_belonging_color(file_id, belonging):
        """设置归属的颜色"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未授权'}), 401
        
        manager = current_app.config['PRODUCT_COMPARE_MANAGER']
        payload = request.json or {}
        color = payload.get('color', '').strip()
        
        if not color or not color.startswith('#'):
            return jsonify({'success': False, 'error': '颜色格式无效'}), 400
        
        try:
            manager._set_belonging_color(file_id, belonging, color)
            # 更新所有使用该归属的产品的颜色
            data = manager._load_file(file_id)
            for product in data.get('products', []):
                if product.get('belonging', '').strip() == belonging:
                    product['color'] = color
            manager._save_file(file_id, data)
            return jsonify({'success': True, 'file': data})
        except FileNotFoundError:
            return jsonify({'success': False, 'error': '文件不存在'}), 404
        except Exception as exc:
            return jsonify({'success': False, 'error': str(exc)}), 500

    @app.route('/file_browser')
    def file_browser():
        """文件浏览器页面，显示文件列表"""
        # 添加登录验证
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        
        root_dir = current_app.config.get('ROOT_DIR')
        if not root_dir or not os.path.exists(root_dir) or not os.path.isdir(root_dir):
            # 如果没有设置根目录或根目录无效，重定向到设置页面
            return redirect(url_for('set_root'))
        
        # 获取请求路径
        path = request.args.get('path', '')
        if path:
            current_path = os.path.join(root_dir, path)
        else:
            current_path = root_dir
        
        # 安全检查：防止路径遍历
        try:
            current_path = os.path.normpath(current_path)
            if not current_path.startswith(os.path.normpath(root_dir)):
                abort(404)
        except Exception:
            abort(404)
        
        # 检查路径是否存在且是目录
        if not os.path.exists(current_path) or not os.path.isdir(current_path):
            abort(404)
        
        # 获取目录内容
        try:
            items = os.listdir(current_path)
        except Exception:
            items = []
        
        # 分类文件和目录
        directories = []
        files = []
        
        for item in items:
            item_path = os.path.join(current_path, item)
            item_rel_path = posixpath.join(path, item)
            
            if os.path.isdir(item_path):
                directories.append({
                    'name': item,
                    'path': item_rel_path,
                    'is_dir': True
                })
            else:
                _, ext = os.path.splitext(item.lower())
                files.append({
                    'name': item,
                    'path': item_rel_path,
                    'is_dir': False,
                    'is_markdown': ext in MARKDOWN_EXTENSIONS,
                    'is_image': ext in IMAGE_EXTENSIONS,
                    'is_pdf': ext in PDF_EXTENSIONS,
                    'is_office': ext in OFFICE_EXTENSIONS,
                    'is_video': ext in VIDEO_EXTENSIONS
                })
        
        # 排序：目录在前，按名称排序
        directories.sort(key=lambda x: x['name'].lower())
        files.sort(key=lambda x: x['name'].lower())
        
        # 构建面包屑导航
        path_parts = []
        current_dir = ''
        for part in path.split(os.sep):
            if part:
                current_dir = posixpath.join(current_dir, part)
                path_parts.append({
                    'name': part,
                    'path': current_dir
                })
        
        # 生成目录路径的上一级目录
        parent_dir = os.path.dirname(current_path)
        if parent_dir and parent_dir != root_dir:
            parent_rel_path = posixpath.dirname(path)
        else:
            parent_rel_path = ''
        
        return render_template('index.html', 
                              directories=directories, 
                              files=files, 
                              current_path=path, 
                              path_parts=path_parts,
                              parent_rel_path=parent_rel_path)
    
    @app.route('/view/<path:filepath>')
    def view_file(filepath):
        """旧的文件预览路由 (完整页面) - 保留以防万一或直接访问"""
        # 添加登录验证
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        root_dir = current_app.config.get('ROOT_DIR')
        if not root_dir:
            abort(404)

        full_path = os.path.join(root_dir, filepath)

        try:
            full_path = os.path.normpath(full_path)
            if not full_path.startswith(os.path.normpath(root_dir)):
                abort(404)
        except Exception:
            abort(404)

        if not os.path.exists(full_path) or os.path.isdir(full_path):
            abort(404)

        filename = os.path.basename(full_path)
        _, ext = os.path.splitext(filename.lower())
        file_type = 'unknown'
        content = ''

        if ext in OFFICE_EXTENSIONS:
            file_type = 'office'
        elif ext in IMAGE_EXTENSIONS:
            file_type = 'image'
        elif ext in MARKDOWN_EXTENSIONS:
            file_type = 'markdown'
            try:
                with open(full_path, 'r', encoding='utf-8') as f:
                    file_content = f.read()
                
                # 使用新的markdown-it-py渲染
                content = render_markdown_content(file_content, filepath)
            except Exception as e:
                content = f"<p>读取文件时出错: {e}</p>"
        elif ext in PDF_EXTENSIONS:
            file_type = 'pdf'
        elif ext in VIDEO_EXTENSIONS:
            file_type = 'video'

        parent_dir = posixpath.dirname(filepath)
        if not parent_dir:
            back_url = url_for('file_browser')
        else:
            back_url = url_for('file_browser', path=parent_dir)

        return render_template('view_file.html',
                               filename=filename,
                               filepath=filepath,
                               file_type=file_type,
                               content=content,
                               back_url=back_url)
    
    @app.route('/view_file')
    def view_file_compat():
        """兼容的文件预览路由，接受path参数"""
        filepath = request.args.get('path', '')
        if filepath:
            # 不再重定向，直接处理文件预览逻辑
            root_dir = current_app.config.get('ROOT_DIR')
            if not root_dir:
                # 如果没有设置ROOT_DIR，尝试在当前目录查找
                full_path = os.path.join(os.getcwd(), filepath)
            else:
                full_path = os.path.join(root_dir, filepath)
                
            try:
                full_path = os.path.normpath(full_path)
            except Exception:
                return render_template('view_file.html', filename=filepath, filepath=filepath, file_type='text', content="<p>文件路径无效</p>", back_url=url_for('index'))

            if not os.path.exists(full_path) or os.path.isdir(full_path):
                # 尝试在当前目录直接查找
                alt_path = os.path.join(os.getcwd(), filepath)
                if os.path.exists(alt_path) and not os.path.isdir(alt_path):
                    full_path = alt_path
                else:
                    return render_template('view_file.html', filename=filepath, filepath=filepath, file_type='text', content="<p>文件不存在</p>", back_url=url_for('index'))

            filename = os.path.basename(full_path)
            _, ext = os.path.splitext(filename.lower())
            file_type = 'unknown'
            content = ''

            if ext in MARKDOWN_EXTENSIONS:
                file_type = 'markdown'
                try:
                    with open(full_path, 'r', encoding='utf-8') as f:
                        file_content = f.read()
                    
                    # 使用新的markdown-it-py渲染
                    content = render_markdown_content(file_content, filepath)
                except Exception as e:
                    content = f"<p>读取文件时出错: {e}</p>"
            else:
                file_type = 'text'
                try:
                    with open(full_path, 'r', encoding='utf-8') as f:
                        content = f.read().replace('\n', '<br>')
                except Exception as e:
                    content = f"<p>读取文件时出错: {e}</p>"

            return render_template('view_file.html',
                                   filename=filename,
                                   filepath=filepath,
                                   file_type=file_type,
                                   content=content,
                                   back_url=url_for('file_browser'))
        return render_template('view_file.html', filename="", filepath="", file_type='text', content="<p>未指定文件</p>", back_url=url_for('file_browser'))
    
    @app.route('/download/<path:filepath>')
    def download_file(filepath):
        """下载文件路由"""
        root_dir = current_app.config.get('ROOT_DIR')
        if not root_dir:
            # 如果没有设置ROOT_DIR，尝试在当前目录查找
            root_dir = os.getcwd()
        
        # 安全检查：防止路径遍历
        try:
            full_path = os.path.join(root_dir, filepath)
            full_path = os.path.normpath(full_path)
            if not full_path.startswith(os.path.normpath(root_dir)):
                abort(404)
        except Exception:
            abort(404)
        
        if not os.path.exists(full_path) or os.path.isdir(full_path):
            # 尝试直接在当前目录查找
            alt_path = os.path.join(os.getcwd(), filepath)
            if os.path.exists(alt_path) and not os.path.isdir(alt_path):
                full_path = alt_path
            else:
                abort(404)
        
        # 获取文件所在目录和文件名
        directory = os.path.dirname(full_path)
        filename = os.path.basename(full_path)
        
        # 使用send_from_directory发送文件
        return send_from_directory(directory, filename, as_attachment=True)
    
    @app.route('/preview/<path:filepath>')
    def preview_file(filepath):
        """预览文件路由（非下载）"""
        root_dir = current_app.config.get('ROOT_DIR')
        if not root_dir:
            # 如果没有设置ROOT_DIR，尝试在当前目录查找
            root_dir = os.getcwd()
        
        # 安全检查：防止路径遍历
        try:
            full_path = os.path.join(root_dir, filepath)
            full_path = os.path.normpath(full_path)
            if not full_path.startswith(os.path.normpath(root_dir)):
                abort(404)
        except Exception:
            abort(404)
        
        if not os.path.exists(full_path) or os.path.isdir(full_path):
            # 尝试直接在当前目录查找
            alt_path = os.path.join(os.getcwd(), filepath)
            if os.path.exists(alt_path) and not os.path.isdir(alt_path):
                full_path = alt_path
            else:
                abort(404)
        
        # 获取文件所在目录和文件名
        directory = os.path.dirname(full_path)
        filename = os.path.basename(full_path)
        
        # 根据文件扩展名设置正确的MIME类型
        _, ext = os.path.splitext(filename.lower())
        mimetype = None
        
        if ext in IMAGE_EXTENSIONS:
            if ext == '.jpg' or ext == '.jpeg':
                mimetype = 'image/jpeg'
            elif ext == '.png':
                mimetype = 'image/png'
            elif ext == '.gif':
                mimetype = 'image/gif'
            elif ext == '.svg':
                mimetype = 'image/svg+xml'
            elif ext == '.webp':
                mimetype = 'image/webp'
            else:
                mimetype = 'image/jpeg'
        elif ext == '.pdf':
            mimetype = 'application/pdf'
        elif ext in VIDEO_EXTENSIONS:
            if ext == '.mp4':
                mimetype = 'video/mp4'
            elif ext == '.avi':
                mimetype = 'video/x-msvideo'
            elif ext == '.mov':
                mimetype = 'video/quicktime'
            else:
                mimetype = 'video/mp4'
        elif ext in ['.mp3', '.wav', '.flac', '.ogg', '.wma', '.m4a']:
            if ext == '.mp3':
                mimetype = 'audio/mpeg'
            elif ext == '.wav':
                mimetype = 'audio/wav'
            elif ext == '.ogg':
                mimetype = 'audio/ogg'
            else:
                mimetype = 'audio/mpeg'
        
        # 不设置as_attachment，这样浏览器会尝试预览而不是下载
        return send_from_directory(directory, filename, as_attachment=False, mimetype=mimetype)
    
    @app.route('/set_root', methods=['GET', 'POST'])
    def set_root():
        """设置根目录（需要管理员密码）"""
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        
        current_root = current_app.config.get('ROOT_DIR')
        
        if request.method == 'POST':
            new_root = request.form.get('root_path')
            admin_password = request.form.get('admin_password')  # 获取管理员密码
            
            # 验证管理员密码
            correct_admin_password = current_app.config.get('ADMIN_PASSWORD', 'admin123')
            if admin_password != correct_admin_password:
                return render_template('set_root.html', 
                                     error="管理员密码错误！无法修改目录。", 
                                     current_root=current_root)
            
            if new_root and os.path.exists(new_root) and os.path.isdir(new_root):
                current_app.config['ROOT_DIR'] = new_root
                # 保存到配置文件，保留所有配置
                config = configparser.ConfigParser()
                current_password = current_app.config.get('PASSWORD', 'ats123')
                config['settings'] = {
                    'root_dir': new_root,
                    'password': current_password,
                    'admin_password': correct_admin_password
                }
                # 获取配置文件路径
                config_file = current_app.config.get('CONFIG_FILE', 'config.ini')
                with open(config_file, 'w', encoding='utf-8') as f:
                    config.write(f)
                return redirect(url_for('file_browser'))
            else:
                return render_template('set_root.html', error="无效的目录路径", current_root=current_root)
        else:
            # GET请求，显示设置页面
            return render_template('set_root.html', current_root=current_root)
    
    @app.route('/set_current_as_root', methods=['POST'])
    def set_current_as_root():
        """将当前浏览的文件夹设置为共享根目录（需要管理员密码）"""
        if 'logged_in' not in session:
            return jsonify({'error': '请先登录'}), 401
        
        data = request.get_json()
        current_path = data.get('path', '')
        admin_password = data.get('admin_password', '')  # 获取管理员密码
        
        # 验证管理员密码
        correct_admin_password = current_app.config.get('ADMIN_PASSWORD', 'admin123')
        if admin_password != correct_admin_password:
            return jsonify({'error': '管理员密码错误！无法修改目录。', 'require_password': True}), 403
        
        # 获取当前根目录
        root_dir = current_app.config.get('ROOT_DIR')
        if not root_dir:
            return jsonify({'error': '根目录未设置'}), 400
        
        # 计算新的根目录路径
        if current_path:
            new_root = os.path.join(root_dir, current_path)
        else:
            new_root = root_dir
        
        # 验证新路径
        try:
            new_root = os.path.normpath(new_root)
            if not os.path.exists(new_root) or not os.path.isdir(new_root):
                return jsonify({'error': '目标路径不存在或不是目录'}), 400
        except Exception as e:
            return jsonify({'error': f'路径验证失败: {e}'}), 400
        
        # 更新配置
        current_app.config['ROOT_DIR'] = new_root
        
        # 保存到配置文件
        try:
            config = configparser.ConfigParser()
            current_password = current_app.config.get('PASSWORD', 'ats123')
            
            config['settings'] = {
                'root_dir': new_root,
                'password': current_password,
                'admin_password': correct_admin_password
            }
            
            config_file = current_app.config.get('CONFIG_FILE', 'config.ini')
            with open(config_file, 'w', encoding='utf-8') as f:
                config.write(f)
            
            return jsonify({
                'success': True,
                'new_root': new_root,
                'message': '共享文件夹已更新'
            })
        except Exception as e:
            return jsonify({'error': f'保存配置失败: {e}'}), 500
    
    @app.route('/view_model/<path:filepath>')
    def view_model(filepath):
        """3D模型查看器页面"""
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        
        root_dir = current_app.config.get('ROOT_DIR')
        if not root_dir:
            abort(404)
        
        full_path = os.path.join(root_dir, filepath)
        
        # 安全检查
        try:
            full_path = os.path.normpath(full_path)
            if not full_path.startswith(os.path.normpath(root_dir)):
                abort(404)
        except Exception:
            abort(404)
        
        if not os.path.exists(full_path) or os.path.isdir(full_path):
            abort(404)
        
        filename = os.path.basename(full_path)
        _, ext = os.path.splitext(filename.lower())
        
        # 检查是否是3D模型文件
        if ext not in MODEL_3D_EXTENSIONS:
            abort(400)
        
        # 生成模型URL
        model_url = url_for('preview_file', filepath=filepath)
        
        # 返回按钮URL
        parent_dir = posixpath.dirname(filepath)
        if not parent_dir:
            back_url = url_for('file_browser')
        else:
            back_url = url_for('file_browser', path=parent_dir)
        
        return render_template('model_viewer.html',
                             filename=filename,
                             filepath=filepath,
                             file_ext=ext,
                             model_url=model_url,
                             back_url=back_url)
    
    @app.route('/get_preview_content', methods=['POST'])
    def get_preview_content():
        """获取文件预览内容"""
        # 添加登录验证
        if 'logged_in' not in session:
            return jsonify({'error': '请先登录'}), 401
        
        data = request.get_json()
        filepath = data.get('filepath')
        
        if not filepath:
            return jsonify({'error': '文件路径不能为空'}), 400
        
        root_dir = current_app.config.get('ROOT_DIR')
        if not root_dir:
            # 如果没有设置ROOT_DIR，尝试在当前目录查找
            root_dir = os.getcwd()
        
        # 安全检查：防止路径遍历
        try:
            full_path = os.path.join(root_dir, filepath)
            full_path = os.path.normpath(full_path)
            if not full_path.startswith(os.path.normpath(root_dir)):
                return jsonify({'error': '访问被拒绝'}), 403
        except Exception:
            return jsonify({'error': '路径解析错误'}), 400
        
        if not os.path.exists(full_path) or os.path.isdir(full_path):
            return jsonify({'error': '文件不存在'}), 404
        
        filename = os.path.basename(full_path)
        _, ext = os.path.splitext(filename.lower())
        file_type = 'unknown'
        content_html = ''
        download_url = url_for('download_file', filepath=filepath)
        preview_url = url_for('preview_file', filepath=filepath)
        
        # 代码文件扩展名列表
        CODE_EXTENSIONS = ['.py', '.js', '.html', '.css', '.scss', '.php', '.java', '.c', '.cpp', 
                          '.cs', '.go', '.rb', '.sh', '.bat', '.sql', '.ts', '.tsx', '.jsx', 
                          '.json', '.xml', '.yaml', '.yml', '.md', '.markdown', '.txt', '.csv', '.log']
        
        if ext in MARKDOWN_EXTENSIONS:
            file_type = 'markdown'
            try:
                with open(full_path, 'r', encoding='utf-8') as f:
                    markdown_content = f.read()
                
                # 使用新的markdown-it-py渲染
                content_html = render_markdown_content(markdown_content, filepath)
                
                # 将下载链接替换为预览链接
                content_html = content_html.replace('/download/', '/preview/')
            except Exception as e:
                content_html = f'<p>读取文件时出错: {e}</p>'
            content_html = f'<article class="markdown-body">{content_html}</article>'
        elif ext in PDF_EXTENSIONS:
            file_type = 'pdf'
            content_html = f'<div class="pdf-container"><embed src="{preview_url}" type="application/pdf"></div>'
        elif ext in IMAGE_EXTENSIONS:
            file_type = 'image'
            content_html = f'<div class="image-container"><img src="{preview_url}" alt="{filename}"></div>'
        elif ext in VIDEO_EXTENSIONS:
            file_type = 'video'
            content_html = f'<div class="video-container"><video controls src="{preview_url}"></video></div>'
        elif ext in AUDIO_EXTENSIONS:
            file_type = 'audio'
            content_html = f'''
            <div class="audio-player-container">
                <div class="audio-info">
                    <h3><i class="fas fa-music"></i> {filename}</h3>
                    <p>音频文件</p>
                </div>
                <audio controls src="{preview_url}" style="width: 100%; max-width: 600px;">
                    您的浏览器不支持音频播放
                </audio>
            </div>
            '''
        elif ext in DRAWIO_EXTENSIONS:
            file_type = 'drawio'
            # 提供draw.io文件的预览功能，编辑按钮在header中
            content_html = f'''
            <div class="drawio-container">
                <iframe src="/drawio_embed?filepath={quote(filepath)}" class="drawio-preview" width="100%" height="600px" style="border: none;"></iframe>
            </div>
            '''
        elif ext in MODEL_3D_EXTENSIONS:
            file_type = '3d_model'
            # 3D模型预览 - iframe嵌入方式
            view_model_url = url_for('view_model', filepath=filepath)
            content_html = f'''
            <div class="model-3d-container" style="width: 100%; height: 700px; position: relative; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 12px rgba(0,0,0,0.1);">
                <iframe src="{view_model_url}" 
                        style="width: 100%; height: 100%; border: none; display: block;"
                        allowfullscreen>
                </iframe>
                <div style="position: absolute; top: 10px; left: 10px; background: rgba(0,0,0,0.7); color: white; padding: 8px 12px; border-radius: 6px; font-size: 12px; z-index: 100;">
                    <i class="fas fa-cube"></i> {filename} ({ext.upper()})
                </div>
            </div>
            <div style="margin-top: 15px; padding: 15px; background: #f8f9fa; border-radius: 8px; border-left: 4px solid #409EFF;">
                <p style="margin: 5px 0; color: #666;"><strong>💡 操作提示:</strong></p>
                <p style="margin: 5px 0; color: #666;">🖱️ <strong>鼠标左键</strong> - 旋转模型 | <strong>右键</strong> - 平移视角 | <strong>滚轮</strong> - 缩放</p>
                <p style="margin: 5px 0; color: #666;">🛠️ 使用右上角工具栏可以: 重置视角、切换线框模式、显示网格、改变背景色等</p>
            </div>
            '''
        elif ext in CODE_EXTENSIONS:
            file_type = 'code'
            try:
                with open(full_path, 'r', encoding='utf-8') as f:
                    code_content = f.read()
                    # 获取语言名称（从扩展名映射）
                    language_map = {
                        '.py': 'python',
                        '.js': 'javascript',
                        '.html': 'html',
                        '.css': 'css',
                        '.scss': 'scss',
                        '.php': 'php',
                        '.java': 'java',
                        '.c': 'c',
                        '.cpp': 'cpp',
                        '.cs': 'csharp',
                        '.go': 'go',
                        '.rb': 'ruby',
                        '.sh': 'bash',
                        '.bat': 'batch',
                        '.sql': 'sql',
                        '.ts': 'typescript',
                        '.tsx': 'typescript',
                        '.jsx': 'javascript',
                        '.json': 'json',
                        '.xml': 'xml',
                        '.yaml': 'yaml',
                        '.yml': 'yaml',
                        '.md': 'markdown',
                        '.markdown': 'markdown',
                        '.txt': 'text',
                        '.csv': 'csv',
                        '.log': 'text'
                    }
                    language = language_map.get(ext, 'text')
                    # 使用HTML转义并添加语法高亮标记
                    escaped_content = code_content.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                    content_html = f'<pre class="code-preview language-{language}"><code>{escaped_content}</code></pre>'
            except Exception as e:
                content_html = f'<p>无法预览此文件: {e}</p>'
        else:
            file_type = 'text'
            try:
                with open(full_path, 'r', encoding='utf-8') as f:
                    text_content = f.read()
                    content_html = f'<pre>{text_content}</pre>'
            except Exception as e:
                content_html = f'<p>无法预览此文件: {e}</p>'
        
        return jsonify({
            'filename': filename,
            'file_type': file_type,
            'content_html': content_html,
            'download_url': download_url,
            'preview_url': preview_url
        })
    
    # 用户认证相关路由
    @app.route('/login', methods=['GET', 'POST'])
    def login():
        """登录页面"""
        if request.method == 'POST':
            password = request.form.get('password')
            
            # 使用配置文件中的密码进行验证
            configured_password = current_app.config.get('PASSWORD')  # 修正：使用大写的键名
            
            # === 调试输出 ===
            print("=" * 60)
            print("[登录调试信息]")
            print(f"用户输入的密码: '{password}'")
            print(f"用户输入密码长度: {len(password) if password else 0}")
            print(f"配置的正确密码: '{configured_password}'")
            print(f"配置密码长度: {len(configured_password) if configured_password else 0}")
            print(f"配置文件路径: {current_app.config.get('CONFIG_FILE')}")
            print(f"密码匹配: {password == configured_password}")
            print("=" * 60)
            # === 调试输出结束 ===
            
            if not configured_password:
                # 如果没有配置密码，使用默认密码并记录警告
                configured_password = 'ats123'
                print("[警告] 未找到配置密码，使用默认密码 'ats123'")
            
            if password == configured_password:
                session['logged_in'] = True
                print(f"[成功] 用户登录成功")
                # 登录后重定向到选择页面
                return redirect(url_for('index'))
            else:
                print(f"[失败] 密码不匹配 - 用户输入: '{password}', 正确密码: '{configured_password}'")
                return render_template('login.html', error="密码错误")
        return render_template('login.html')
    

    
    @app.route('/logout')
    def logout():
        """登出功能"""
        session.pop('logged_in', None)
        return redirect(url_for('login'))
    
    @app.route('/help')
    def help_page():
        """帮助页面"""
        # 帮助页面不需要登录也可以查看
        return render_template('help.html')
    
    @app.route('/theme_preview')
    def theme_preview():
        """主题预览页面，展示新配色与控件示例"""
        return render_template('theme_preview.html')
    
    # 串口调试助手
    @app.route('/serial_tool')
    def serial_tool():
        """串口调试助手页面"""
        if not is_logged_in():
            return redirect(url_for('login'))
        return render_template('serial_tool.html')

    def _normalize_parity(value: str) -> str:
        mapping = {
            'none': 'N',
            'n': 'N',
            'even': 'E',
            'e': 'E',
            'odd': 'O',
            'o': 'O',
            'mark': 'M',
            'm': 'M',
            'space': 'S',
            's': 'S',
        }
        if not value:
            return 'N'
        value = value.strip()
        return mapping.get(value.lower(), 'N')

    def _is_safe_port_id(value: str) -> bool:
        return bool(value) and re.fullmatch(r'[A-Za-z0-9_]+', value) is not None

    @app.route('/api/serial/ports', methods=['GET'])
    def api_serial_ports():
        """列出服务器可用串口"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        try:
            ports = serial_manager.list_available_ports()
            for port in ports:
                device = port.get('device')
                if device:
                    port['recommended_id'] = serial_manager.generate_port_id(device, 9600)
            return jsonify({'success': True, 'ports': ports})
        except Exception as exc:  # pylint: disable=broad-except
            return jsonify({'success': False, 'error': str(exc)}), 500

    @app.route('/api/serial/capture/start', methods=['POST'])
    def api_serial_capture_start():
        """开启指定串口的持续日志"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401

        payload = request.json or {}
        device = (payload.get('device') or '').strip()
        if not device:
            return jsonify({'success': False, 'error': '缺少 device 参数'}), 400

        try:
            baudrate = int(payload.get('baudrate', 9600))
            bytesize = int(payload.get('bytesize', 8))
            stopbits = int(payload.get('stopbits', 1))
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': '串口参数格式错误'}), 400

        parity = _normalize_parity(payload.get('parity', 'N'))
        port_id = payload.get('port_id')
        try:
            session = serial_manager.start_logging(
                device=device,
                baudrate=baudrate,
                bytesize=bytesize,
                parity=parity,
                stopbits=stopbits,
                port_id=port_id,
            )
            port_info = serial_manager.get_port_info(session['port_id'])
            session['port_info'] = port_info
            session['log_files'] = serial_manager.list_log_files(session['port_id'])
            return jsonify({'success': True, 'session': session})
        except ValueError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400
        except Exception as exc:  # pylint: disable=broad-except
            return jsonify({'success': False, 'error': str(exc)}), 500

    @app.route('/api/serial/capture/stop', methods=['POST'])
    def api_serial_capture_stop():
        """停止串口持续日志"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401

        payload = request.json or {}
        port_id = payload.get('port_id')
        if not _is_safe_port_id(port_id or ''):
            return jsonify({'success': False, 'error': '端口标识不合法'}), 400
        try:
            success = serial_manager.stop_logging(port_id)
            if not success:
                return jsonify({'success': False, 'error': '未找到对应日志会话'}), 404
            return jsonify({'success': True})
        except Exception as exc:  # pylint: disable=broad-except
            return jsonify({'success': False, 'error': str(exc)}), 500

    @app.route('/api/serial/capture/status', methods=['GET'])
    def api_serial_capture_status():
        """查询当前持续日志会话"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        try:
            sessions = serial_manager.get_logging_status()
            for session in sessions:
                port_id = session.get('port_id')
                if not port_id:
                    continue
                session['log_files'] = serial_manager.list_log_files(port_id)
                session['port_info'] = serial_manager.get_port_info(port_id)
            return jsonify({'success': True, 'sessions': sessions})
        except Exception as exc:  # pylint: disable=broad-except
            return jsonify({'success': False, 'error': str(exc)}), 500

    @app.route('/api/serial/capture/log/<port_id>', methods=['GET'])
    def api_serial_capture_log(port_id):
        """读取指定串口的日志条目"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        if not _is_safe_port_id(port_id):
            return jsonify({'success': False, 'error': '端口标识不合法'}), 400

        try:
            limit = int(request.args.get('limit', 500))
        except ValueError:
            return jsonify({'success': False, 'error': 'limit 参数错误'}), 400

        since = request.args.get('since')
        try:
            entries = serial_manager.read_log_entries(
                port_id=port_id,
                limit=max(0, limit),
                since=since,
            )
            return jsonify({'success': True, 'entries': entries})
        except Exception as exc:  # pylint: disable=broad-except
            return jsonify({'success': False, 'error': str(exc)}), 500
    
    # 保存串口日志
    @app.route('/save_serial_log', methods=['POST'])
    def save_serial_log():
        """保存串口日志到服务器"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        try:
            data = request.json
            log_content = data.get('log_content', '')
            port_name = data.get('port_name', 'unknown')
            
            # 创建 serial_logs 目录
            import os
            from datetime import datetime
            
            if getattr(sys, 'frozen', False):
                base_dir = os.path.dirname(sys.executable)
            else:
                base_dir = os.path.dirname(os.path.abspath(__file__))
            
            log_dir = os.path.join(base_dir, 'logs', 'serial_logs')
            os.makedirs(log_dir, exist_ok=True)
            
            # 生成日志文件名（精确到毫秒）
            now = datetime.now()
            milliseconds = now.microsecond // 1000
            timestamp = now.strftime(f'%Y-%m-%d_%H-%M-%S.{milliseconds:03d}')
            filename = f"{port_name}_{timestamp}.log"
            filepath = os.path.join(log_dir, filename)
            
            # 保存日志
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(log_content)
            
            return jsonify({
                'success': True,
                'filepath': filepath,
                'filename': filename
            })
        
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    # 获取串口历史记录列表
    @app.route('/get_serial_logs', methods=['GET'])
    def get_serial_logs():
        """获取串口历史记录列表"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        try:
            import os
            
            if getattr(sys, 'frozen', False):
                base_dir = os.path.dirname(sys.executable)
            else:
                base_dir = os.path.dirname(os.path.abspath(__file__))
            
            log_dir = os.path.join(base_dir, 'logs', 'serial_logs')
            
            if not os.path.exists(log_dir):
                return jsonify({'success': True, 'logs': []})
            
            # 获取所有日志文件
            logs = []
            for filename in os.listdir(log_dir):
                if filename.endswith('.log'):
                    filepath = os.path.join(log_dir, filename)
                    stat = os.stat(filepath)
                    logs.append({
                        'filename': filename,
                        'size': stat.st_size,
                        'modified': stat.st_mtime
                    })
            
            # 按修改时间倒序排序
            logs.sort(key=lambda x: x['modified'], reverse=True)
            
            return jsonify({'success': True, 'logs': logs})
        
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    # 读取串口历史记录
    @app.route('/read_serial_log/<filename>')
    def read_serial_log(filename):
        """读取串口历史记录"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        try:
            import os
            
            if getattr(sys, 'frozen', False):
                base_dir = os.path.dirname(sys.executable)
            else:
                base_dir = os.path.dirname(os.path.abspath(__file__))
            
            log_dir = os.path.join(base_dir, 'logs', 'serial_logs')
            filepath = os.path.join(log_dir, filename)
            
            # 安全检查
            if not os.path.abspath(filepath).startswith(os.path.abspath(log_dir)):
                return jsonify({'success': False, 'error': '非法路径'}), 403
            
            if not os.path.exists(filepath):
                return jsonify({'success': False, 'error': '文件不存在'}), 404
            
            # 读取日志
            with open(filepath, 'r', encoding='utf-8') as f:
                content = f.read()
            
            return jsonify({
                'success': True,
                'filename': filename,
                'content': content
            })
        
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    # 保存快捷指令
    @app.route('/save_serial_commands', methods=['POST'])
    def save_serial_commands():
        """保存快捷指令到服务器"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        try:
            import json
            import os
            
            data = request.json
            commands = data.get('commands', [])
            
            if getattr(sys, 'frozen', False):
                base_dir = os.path.dirname(sys.executable)
            else:
                base_dir = os.path.dirname(os.path.abspath(__file__))
            
            config_dir = os.path.join(base_dir, 'logs')
            os.makedirs(config_dir, exist_ok=True)
            
            filepath = os.path.join(config_dir, 'serial_commands.json')
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(commands, f, ensure_ascii=False, indent=2)
            
            return jsonify({'success': True})
        
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    # 加载快捷指令
    @app.route('/load_serial_commands', methods=['GET'])
    def load_serial_commands():
        """从服务器加载快捷指令"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        try:
            import json
            import os
            
            if getattr(sys, 'frozen', False):
                base_dir = os.path.dirname(sys.executable)
            else:
                base_dir = os.path.dirname(os.path.abspath(__file__))
            
            filepath = os.path.join(base_dir, 'logs', 'serial_commands.json')
            
            if not os.path.exists(filepath):
                return jsonify({'success': True, 'commands': []})
            
            with open(filepath, 'r', encoding='utf-8') as f:
                commands = json.load(f)
            
            return jsonify({'success': True, 'commands': commands})
        
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    # 串口诊断工具
    @app.route('/serial_diagnostic')
    def serial_diagnostic():
        """串口诊断工具页面"""
        if not is_logged_in():
            return redirect(url_for('login'))
        return render_template('serial_diagnostic.html')
    
    # Draw.io主编辑器页面（带保存功能）
    @app.route('/drawio_main')
    def drawio_main():
        """Draw.io主编辑器页面，支持保存和上传"""
        if not is_logged_in():
            return redirect(url_for('login'))
        
        # 获取文件路径参数（如果从文件浏览器打开）
        filepath = request.args.get('filepath', '')
        diagram_content = ''
        
        print(f"[DEBUG] drawio_main - filepath参数: '{filepath}'")
        
        if filepath:
            # 从服务器加载文件
            root_dir = current_app.config.get('ROOT_DIR')
            print(f"[DEBUG] ROOT_DIR: '{root_dir}'")
            
            # 处理路径
            full_path = os.path.join(root_dir, filepath.lstrip('/'))
            print(f"[DEBUG] 完整路径: '{full_path}'")
            
            try:
                full_path = os.path.normpath(full_path)
                print(f"[DEBUG] 标准化路径: '{full_path}'")
                print(f"[DEBUG] 文件存在: {os.path.exists(full_path)}")
                
                if full_path.startswith(os.path.normpath(root_dir)) and os.path.exists(full_path):
                    with open(full_path, 'r', encoding='utf-8') as f:
                        diagram_content = f.read()
                    print(f"[DEBUG] 成功读取文件，内容长度: {len(diagram_content)}")
                    print(f"[DEBUG] 文件前100字符: {diagram_content[:100] if diagram_content else 'EMPTY'}")
                    print(f"[DEBUG] 是否包含<mxfile: {'<mxfile' in diagram_content}")
                else:
                    print(f"[DEBUG] 路径验证失败或文件不存在")
            except Exception as e:
                print(f"[ERROR] 加载文件失败: {e}")
                import traceback
                traceback.print_exc()
        
        return render_template('drawio_main.html', filepath=filepath, diagram_content=diagram_content)
    
    # 主draw.io路由 - 与原始实现兼容
    @app.route('/drawio')
    def drawio():
        """Draw.io编辑器主页面，与原始实现兼容"""
        if not is_logged_in():
            return redirect(url_for('login'))
        
        # 检查draw.io文件是否存在
        import os
        # 使用正确的路径获取方式，支持打包环境
        if getattr(sys, 'frozen', False):
            # 打包环境：static在exe所在目录
            base_dir = os.path.dirname(sys.executable)
            drawio_dir = os.path.join(base_dir, 'static', 'drawio')
        else:
            # 开发环境
            drawio_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'drawio')
        
        drawio_index = os.path.join(drawio_dir, 'index.html')
        
        # 构建离线模式URL参数
        # offline=1: 离线模式
        # stealth=1: 隐身模式，禁用追踪和外部服务
        # local=1: 只允许本地存储
        # sync=none: 禁用同步功能
        # lang=zh: 中文界面
        offline_params = '?offline=1&stealth=1&local=1&sync=none&mode=device&lang=zh'
        
        if not os.path.exists(drawio_index):
            # 如果没有index.html，提供一个简单的HTML页面来加载编辑器
            html_content = f'''
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>draw.io 编辑器</title>
    <style>
        body, html {{ margin: 0; padding: 0; height: 100%; overflow: hidden; }}
        iframe {{ width: 100%; height: 100%; border: none; }}
    </style>
</head>
<body>
    <iframe src="/drawio/{offline_params}" allowfullscreen></iframe>
</body>
</html>
            '''
            response = make_response(html_content)
            response.headers['Content-Type'] = 'text/html; charset=utf-8'
            return response
        
        # 读取并返回index.html，并注入离线模式配置
        try:
            with open(drawio_index, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # 在index.html中注入离线模式配置
            # 在head标签中添加配置脚本
            config_script = '''
    <script>
        // 配置drawio为离线模式
        window.DRAWIO_BASE_URL = window.location.origin;
        window.DRAWIO_CONFIG = {
            defaultLibraries: 'general',
            enableCustomLibraries: false,
            enabledLibraries: ['general', 'uml', 'entity', 'mockup', 'flowchart'],
            plugins: [],
            // 禁用云存储选项
            mode: 'device',
            offline: '1',
            stealth: '1',
            local: '1'
        };
    </script>
'''
            # 在</head>之前插入配置
            if '</head>' in content:
                content = content.replace('</head>', config_script + '</head>')
            
            response = make_response(content)
            response.headers['Content-Type'] = 'text/html; charset=utf-8'
            return response
        except Exception as e:
            # 使用print替代logger
            print(f"Error loading Draw.io index: {str(e)}")
            return "Error loading Draw.io editor", 500
    
    # 重新添加编辑功能路由，但使用正确的iframe src
    @app.route('/drawio_edit')
    def drawio_edit():
        """编辑Draw.io图表"""
        if not is_logged_in():
            return redirect(url_for('login'))
        
        filepath = request.args.get('filepath')
        if not filepath:
            # 不再使用不存在的error.html模板
            return make_response("文件路径不能为空", 400)
        
        # 安全检查：验证文件路径
        root_dir = current_app.config.get('ROOT_DIR')
        full_path = os.path.join(root_dir, filepath.lstrip('/'))
        
        # 确保路径在根目录内
        try:
            full_path = os.path.normpath(full_path)
            if not full_path.startswith(os.path.normpath(root_dir)):
                return make_response("无效的文件路径", 400)
        except Exception:
            return make_response("无效的文件路径", 400)
        
        # 读取现有图表内容（如果存在）
        diagram_content = ''
        if os.path.exists(full_path):
            try:
                with open(full_path, 'r', encoding='utf-8') as f:
                    diagram_content = f.read()
            except Exception:
                diagram_content = ''
        
        return render_template('drawio_edit.html', filepath=filepath, diagram_content=diagram_content)
    
    @app.route('/drawio_save', methods=['POST'])
    def drawio_save():
        """保存draw.io文件"""
        if not is_logged_in():
            return jsonify({'error': '请先登录'}), 401
        
        data = request.get_json()
        filepath = data.get('filepath')
        content = data.get('content')
        
        if not filepath or content is None:
            return jsonify({'error': '文件路径或内容不能为空'}), 400
        
        root_dir = current_app.config.get('ROOT_DIR')
        if not root_dir:
            root_dir = os.getcwd()
        
        # 安全检查
        try:
            full_path = os.path.join(root_dir, filepath)
            full_path = os.path.normpath(full_path)
            if not full_path.startswith(os.path.normpath(root_dir)):
                return jsonify({'error': '访问被拒绝'}), 403
        except Exception:
            return jsonify({'error': '路径解析错误'}), 400
        
        # 确保目录存在，如果不存在则创建
        directory = os.path.dirname(full_path)
        if not os.path.exists(directory):
            try:
                os.makedirs(directory, exist_ok=True)
                print(f"[INFO] 创建目录: {directory}")
            except Exception as e:
                print(f"[ERROR] 创建目录失败: {e}")
                return jsonify({'error': f'创建目录失败: {e}'}), 400
        
        # 保存文件
        try:
            with open(full_path, 'w', encoding='utf-8') as f:
                f.write(content)
            return jsonify({'success': True, 'message': '文件保存成功'})
        except Exception as e:
            return jsonify({'error': f'保存文件失败: {e}'}), 500
    
    
    @app.route('/drawio_embed')
    def drawio_embed():
        """嵌入draw.io编辑器用于预览"""
        # 注意：此路由用于分享链接预览，不需要登录检查
        # 登录检查在分享链接路由中已完成
        
        filepath = request.args.get('filepath')
        readonly = request.args.get('readonly', '0')  # 是否只读模式
        
        if not filepath:
            return make_response("文件路径不能为空", 400)
        
        root_dir = current_app.config.get('ROOT_DIR')
        if not root_dir:
            root_dir = os.getcwd()
        
        # 安全检查
        try:
            full_path = os.path.join(root_dir, filepath)
            full_path = os.path.normpath(full_path)
            if not full_path.startswith(os.path.normpath(root_dir)):
                return make_response("访问被拒绝", 403)
        except Exception:
            return make_response("路径解析错误", 400)
        
        # 读取draw.io文件内容
        if os.path.exists(full_path) and not os.path.isdir(full_path):
            try:
                with open(full_path, 'r', encoding='utf-8') as f:
                    diagram_content = f.read()
            except Exception as e:
                return make_response(f'读取文件失败: {e}', 500)
        else:
            diagram_content = ''
        
        return render_template('drawio_embed.html', 
                              filepath=filepath, 
                              diagram_content=diagram_content,
                              readonly=(readonly == '1'))







    # Draw.io相关静态文件路由
    @app.route('/drawio/<path:filename>')
    def drawio_static(filename):
        """提供Draw.io静态文件，与原始实现一致"""
        # 处理空filename的情况（访问 /drawio/ 时）
        if not filename or filename == '':
            filename = 'index.html'
        
        # 安全检查：防止目录遍历
        if '..' in filename or '//' in filename or '\\' in filename:
            return make_response("访问被拒绝", 403)
        
        import os
        # 使用get_resource_path确保在打包环境下正确找到文件
        if getattr(sys, 'frozen', False):
            # 打包环境：static在exe所在目录
            base_dir = os.path.dirname(sys.executable)
            drawio_dir = os.path.join(base_dir, 'static', 'drawio')
        else:
            # 开发环境
            drawio_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'drawio')
        
        try:
            return send_from_directory(drawio_dir, filename)
        except FileNotFoundError:
            return make_response(f"文件未找到: {filename}", 404)
    
    # Draw.io根路径资源处理
    @app.route('/styles/<path:filename>')
    @app.route('/js/<path:filename>')
    @app.route('/images/<path:filename>')
    @app.route('/libs/<path:filename>')
    @app.route('/resources/<path:filename>')
    @app.route('/mxgraph/<path:filename>')
    @app.route('/math/<path:filename>')
    @app.route('/plugins/<path:filename>')
    @app.route('/shapes/<path:filename>')
    @app.route('/stencils/<path:filename>')
    @app.route('/templates/<path:filename>')
    @app.route('/connect/<path:filename>')
    @app.route('/newDiagramCats/<path:filename>')
    def drawio_root_resources(filename):
        """处理Draw.io根路径资源请求，与原始实现一致"""
        # 获取请求的路径
        path = request.path.lstrip('/')
        
        # 安全检查：防止目录遍历
        if '..' in path or '//' in path or '\\' in path:
            return make_response("访问被拒绝", 403)
        
        import os
        import sys
        # 使用正确的路径获取方式，支持打包环境
        if getattr(sys, 'frozen', False):
            # 打包环境：static在exe所在目录
            base_dir = os.path.dirname(sys.executable)
            drawio_dir = os.path.join(base_dir, 'static', 'drawio')
        else:
            # 开发环境
            drawio_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'drawio')
        
        try:
            return send_from_directory(drawio_dir, path)
        except FileNotFoundError:
            return make_response(f"资源未找到: {path}", 404)
    
    # Service Worker脚本路由
    @app.route('/service-worker.js')
    def service_worker():
        """提供Service Worker脚本，与原始实现一致"""
        try:
            import os
            import sys
            # 使用正确的路径获取方式，支持打包环境
            if getattr(sys, 'frozen', False):
                # 打包环境：static在exe所在目录
                base_dir = os.path.dirname(sys.executable)
                drawio_dir = os.path.join(base_dir, 'static', 'drawio')
            else:
                # 开发环境
                drawio_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'drawio')
            
            response = send_from_directory(drawio_dir, 'service-worker.js')
            response.headers['Content-Type'] = 'application/javascript'
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            return response
        except FileNotFoundError:
            return '', 404
    
    # 处理Draw.io的代理请求（禁用）
    @app.route('/proxy')
    @app.route('/proxt')
    def drawio_proxy():
        """禁用Draw.io的代理功能，返回404"""
        return '', 404

    # =============================
    # 分享链接功能路由
    # =============================
    
    @app.route('/create_share', methods=['POST'])
    def create_share():
        """创建分享链接"""
        if not is_logged_in():
            return jsonify({'error': '请先登录'}), 401
        
        data = request.get_json()
        file_path = data.get('file_path')
        is_directory = data.get('is_directory', False)
        password = data.get('password', '')
        expire_hours = data.get('expire_hours')
        max_visits = data.get('max_visits', -1)
        description = data.get('description', '')
        
        if not file_path:
            return jsonify({'error': '文件路径不能为空'}), 400
        
        # 验证文件是否存在
        root_dir = current_app.config.get('ROOT_DIR')
        full_path = os.path.join(root_dir, file_path)
        
        try:
            full_path = os.path.normpath(full_path)
            if not full_path.startswith(os.path.normpath(root_dir)):
                return jsonify({'error': '访问被拒绝'}), 403
            if not os.path.exists(full_path):
                return jsonify({'error': '文件不存在'}), 404
        except Exception as e:
            return jsonify({'error': f'路径验证失败: {e}'}), 400
        
        # 创建分享链接
        try:
            share_manager = current_app.config['SHARE_MANAGER']
            share_info = share_manager.create_share_link(
                file_path=file_path,
                is_directory=is_directory,
                password=password if password else None,
                expire_hours=int(expire_hours) if expire_hours else None,
                max_visits=int(max_visits) if max_visits else -1,
                created_by=session.get('username', 'anonymous'),
                description=description
            )
            
            # 生成完整的分享链接URL
            share_url = url_for('share_view', share_code=share_info['share_code'], _external=True)
            share_info['share_url'] = share_url
            
            return jsonify({
                'success': True,
                'message': '分享链接创建成功',
                'data': share_info
            })
        except Exception as e:
            return jsonify({'error': f'创建分享链接失败: {e}'}), 500
    
    @app.route('/share/<share_code>')
    def share_view(share_code):
        """访问分享链接"""
        share_manager = current_app.config['SHARE_MANAGER']
        
        # 检查是否需要密码
        link = share_manager.get_share_link(share_code)
        if not link:
            return render_template('share_error.html', 
                                 error='分享链接不存在或已被删除'), 404
        
        # 如果需要密码且未验证，显示密码输入页面
        if link['password'] and not session.get(f'share_verified_{share_code}'):
            return render_template('share_password.html', share_code=share_code)
        
        # 验证分享链接
        result = share_manager.verify_share_link(share_code)
        if not result['valid']:
            return render_template('share_error.html', error=result['reason']), 403
        
        # 记录访问
        client_ip = request.environ.get('REMOTE_ADDR', 'unknown')
        user_agent = request.headers.get('User-Agent', '')
        share_manager.record_visit(share_code, client_ip, user_agent)
        
        # 获取文件信息
        link_data = result['data']
        root_dir = current_app.config.get('ROOT_DIR')
        full_path = os.path.join(root_dir, link_data['file_path'])
        
        # 安全检查
        try:
            full_path = os.path.normpath(full_path)
            if not full_path.startswith(os.path.normpath(root_dir)):
                return render_template('share_error.html', error='访问被拒绝'), 403
        except Exception:
            return render_template('share_error.html', error='路径解析错误'), 400
        
        if not os.path.exists(full_path):
            return render_template('share_error.html', error='文件不存在'), 404
        
        # 渲染分享页面
        return render_template('share_view.html',
                             share_code=share_code,
                             link_data=link_data,
                             file_path=link_data['file_path'],
                             is_directory=link_data['is_directory'],
                             filename=os.path.basename(full_path))
    
    @app.route('/share/<share_code>/verify', methods=['POST'])
    def share_verify_password(share_code):
        """验证分享链接密码"""
        data = request.get_json()
        password = data.get('password', '')
        
        share_manager = current_app.config['SHARE_MANAGER']
        result = share_manager.verify_share_link(share_code, password)
        
        if result['valid']:
            # 标记为已验证
            session[f'share_verified_{share_code}'] = True
            return jsonify({'success': True, 'message': '验证成功'})
        else:
            return jsonify({'success': False, 'message': result['reason']}), 403
    
    @app.route('/share/<share_code>/download')
    def share_download(share_code):
        """下载分享的文件"""
        share_manager = current_app.config['SHARE_MANAGER']
        
        # 检查是否已验证密码
        link = share_manager.get_share_link(share_code)
        if not link:
            abort(404)
        
        if link['password'] and not session.get(f'share_verified_{share_code}'):
            abort(403)
        
        # 验证分享链接
        result = share_manager.verify_share_link(share_code)
        if not result['valid']:
            abort(403)
        
        # 获取文件
        link_data = result['data']
        root_dir = current_app.config.get('ROOT_DIR')
        full_path = os.path.join(root_dir, link_data['file_path'])
        
        try:
            full_path = os.path.normpath(full_path)
            if not full_path.startswith(os.path.normpath(root_dir)):
                abort(403)
        except Exception:
            abort(400)
        
        if not os.path.exists(full_path) or os.path.isdir(full_path):
            abort(404)
        
        directory = os.path.dirname(full_path)
        filename = os.path.basename(full_path)
        
        return send_from_directory(directory, filename, as_attachment=True)
    
    @app.route('/share/<share_code>/preview')
    def share_preview(share_code):
        """预览分享的文件"""
        share_manager = current_app.config['SHARE_MANAGER']
        
        # 检查是否已验证密码
        link = share_manager.get_share_link(share_code)
        if not link:
            abort(404)
        
        if link['password'] and not session.get(f'share_verified_{share_code}'):
            abort(403)
        
        # 验证分享链接
        result = share_manager.verify_share_link(share_code)
        if not result['valid']:
            abort(403)
        
        # 获取文件
        link_data = result['data']
        root_dir = current_app.config.get('ROOT_DIR')
        full_path = os.path.join(root_dir, link_data['file_path'])
        
        try:
            full_path = os.path.normpath(full_path)
            if not full_path.startswith(os.path.normpath(root_dir)):
                abort(403)
        except Exception:
            abort(400)
        
        if not os.path.exists(full_path) or os.path.isdir(full_path):
            abort(404)
        
        directory = os.path.dirname(full_path)
        filename = os.path.basename(full_path)
        
        # 根据文件扩展名设置MIME类型或渲染预览页面
        _, ext = os.path.splitext(filename.lower())
        
        # 对于drawio文件，返回预览页面而不是文件本身
        if ext in DRAWIO_EXTENSIONS:
            # 读取drawio文件内容
            try:
                with open(full_path, 'r', encoding='utf-8') as f:
                    diagram_content = f.read()
            except Exception as e:
                diagram_content = ''
            
            # 返回drawio预览页面
            return render_template('share_drawio_preview.html',
                                 share_code=share_code,
                                 filename=filename,
                                 diagram_content=diagram_content,
                                 file_path=link_data['file_path'])
        
        # 对于Markdown文件，返回渲染后的HTML
        elif ext in MARKDOWN_EXTENSIONS:
            try:
                with open(full_path, 'r', encoding='utf-8') as f:
                    markdown_content = f.read()
                html_content = render_markdown_content(markdown_content, link_data['file_path'])
            except Exception as e:
                html_content = f'<p>读取文件时出错: {e}</p>'
            
            return render_template('share_markdown_preview.html',
                                 share_code=share_code,
                                 filename=filename,
                                 content=html_content)
        
        # 对于其他文件类型，设置正确的MIME类型并发送
        mimetype = None
        
        if ext in IMAGE_EXTENSIONS:
            if ext in ['.jpg', '.jpeg']:
                mimetype = 'image/jpeg'
            elif ext == '.png':
                mimetype = 'image/png'
            elif ext == '.gif':
                mimetype = 'image/gif'
            elif ext == '.svg':
                mimetype = 'image/svg+xml'
            elif ext == '.webp':
                mimetype = 'image/webp'
        elif ext == '.pdf':
            mimetype = 'application/pdf'
        elif ext in VIDEO_EXTENSIONS:
            if ext == '.mp4':
                mimetype = 'video/mp4'
            elif ext == '.avi':
                mimetype = 'video/x-msvideo'
            elif ext == '.mov':
                mimetype = 'video/quicktime'
        elif ext in AUDIO_EXTENSIONS:
            if ext == '.mp3':
                mimetype = 'audio/mpeg'
            elif ext == '.wav':
                mimetype = 'audio/wav'
            elif ext == '.ogg':
                mimetype = 'audio/ogg'
            elif ext == '.m4a':
                mimetype = 'audio/mp4'
        
        return send_from_directory(directory, filename, as_attachment=False, mimetype=mimetype)
    
    @app.route('/manage_shares')
    def manage_shares():
        """管理分享链接页面"""
        if not is_logged_in():
            return redirect(url_for('login'))
        
        share_manager = current_app.config['SHARE_MANAGER']
        
        # 清理过期链接
        share_manager.cleanup_expired_links()
        
        # 获取所有分享链接
        shares = share_manager.get_all_share_links()
        
        # 为每个分享链接生成完整URL
        for share in shares:
            share['share_url'] = url_for('share_view', 
                                        share_code=share['share_code'], 
                                        _external=True)
        
        from datetime import datetime
        return render_template('manage_shares.html', shares=shares, now=datetime.now().isoformat())
    
    @app.route('/delete_share/<share_code>', methods=['POST'])
    def delete_share(share_code):
        """删除分享链接"""
        if not is_logged_in():
            return jsonify({'error': '请先登录'}), 401
        
        share_manager = current_app.config['SHARE_MANAGER']
        
        if share_manager.delete_share_link(share_code):
            return jsonify({'success': True, 'message': '分享链接已删除'})
        else:
            return jsonify({'success': False, 'message': '删除失败'}), 500
    
    @app.route('/share_stats/<share_code>')
    def share_stats(share_code):
        """查看分享链接统计"""
        if not is_logged_in():
            return redirect(url_for('login'))
        
        share_manager = current_app.config['SHARE_MANAGER']
        stats = share_manager.get_visit_stats(share_code)
        
        if not stats['link']:
            abort(404)
        
        return render_template('share_stats.html', 
                             share_code=share_code,
                             stats=stats)
    
    @app.route('/get_file_content/<path:filepath>')
    def get_file_content(filepath):
        """获取文件内容用于编辑"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        root_path = current_app.config['ROOT_DIR']
        full_path = os.path.join(root_path, filepath)
        
        # 安全检查
        if not os.path.abspath(full_path).startswith(os.path.abspath(root_path)):
            abort(403)
        
        if not os.path.exists(full_path) or not os.path.isfile(full_path):
            abort(404)
        
        try:
            # 尝试以UTF-8读取
            with open(full_path, 'r', encoding='utf-8') as f:
                content = f.read()
            return content, 200, {'Content-Type': 'text/plain; charset=utf-8'}
        except UnicodeDecodeError:
            # 尝试GBK
            try:
                with open(full_path, 'r', encoding='gbk') as f:
                    content = f.read()
                return content, 200, {'Content-Type': 'text/plain; charset=utf-8'}
            except:
                return '无法读取文件：编码不支持', 400
        except Exception as e:
            return f'读取文件失败: {str(e)}', 500
    
    @app.route('/save_file', methods=['POST'])
    def save_file():
        """保存文件内容"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        try:
            data = request.json
            filepath = data.get('filepath')
            content = data.get('content')
            
            if not filepath:
                return jsonify({'success': False, 'error': '缺少文件路径'}), 400
            
            if content is None:
                return jsonify({'success': False, 'error': '缺少文件内容'}), 400
            
            root_path = current_app.config['ROOT_DIR']
            full_path = os.path.join(root_path, filepath)
            
            # 安全检查：确保路径在允许的目录内
            if not os.path.abspath(full_path).startswith(os.path.abspath(root_path)):
                return jsonify({'success': False, 'error': '无权访问该路径'}), 403
            
            # 确保目录存在
            os.makedirs(os.path.dirname(full_path), exist_ok=True)
            
            # 保存文件
            with open(full_path, 'w', encoding='utf-8') as f:
                f.write(content)
            
            return jsonify({'success': True, 'message': '保存成功'})
            
        except Exception as e:
            return jsonify({'success': False, 'error': f'保存失败: {str(e)}'}), 500
    
    @app.route('/create_file', methods=['POST'])
    def create_file():
        """创建新文件"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        try:
            data = request.json
            filename = data.get('filename')
            path = data.get('path', '')
            content = data.get('content', '')
            
            if not filename:
                return jsonify({'success': False, 'error': '文件名不能为空'}), 400
            
            root_path = current_app.config['ROOT_DIR']
            if path:
                full_path = os.path.join(root_path, path, filename)
            else:
                full_path = os.path.join(root_path, filename)
            
            # 安全检查
            if not os.path.abspath(full_path).startswith(os.path.abspath(root_path)):
                return jsonify({'success': False, 'error': '无权访问该路径'}), 403
            
            # 检查文件是否已存在
            if os.path.exists(full_path):
                return jsonify({'success': False, 'error': '文件已存在'}), 400
            
            # 确保目录存在
            os.makedirs(os.path.dirname(full_path), exist_ok=True)
            
            # 创建文件
            with open(full_path, 'w', encoding='utf-8') as f:
                f.write(content)
            
            return jsonify({'success': True, 'message': '文件创建成功'})
            
        except Exception as e:
            return jsonify({'success': False, 'error': f'创建失败: {str(e)}'}), 500
    
    @app.route('/create_folder', methods=['POST'])
    def create_folder():
        """创建新文件夹"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        try:
            data = request.json
            foldername = data.get('foldername')
            path = data.get('path', '')
            
            if not foldername:
                return jsonify({'success': False, 'error': '文件夹名称不能为空'}), 400
            
            root_path = current_app.config['ROOT_DIR']
            if path:
                full_path = os.path.join(root_path, path, foldername)
            else:
                full_path = os.path.join(root_path, foldername)
            
            # 安全检查
            if not os.path.abspath(full_path).startswith(os.path.abspath(root_path)):
                return jsonify({'success': False, 'error': '无权访问该路径'}), 403
            
            # 检查文件夹是否已存在
            if os.path.exists(full_path):
                return jsonify({'success': False, 'error': '文件夹已存在'}), 400
            
            # 创建文件夹
            os.makedirs(full_path)
            
            return jsonify({'success': True, 'message': '文件夹创建成功'})
            
        except Exception as e:
            return jsonify({'success': False, 'error': f'创建失败: {str(e)}'}), 500
    
    @app.route('/upload_files', methods=['POST'])
    def upload_files():
        """上传文件"""
        if not is_logged_in():
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        try:
            if 'files' not in request.files:
                return jsonify({'success': False, 'error': '没有文件'}), 400
            
            files = request.files.getlist('files')
            path = request.form.get('path', '')
            
            root_path = current_app.config['ROOT_DIR']
            if path:
                target_dir = os.path.join(root_path, path)
            else:
                target_dir = root_path
            
            # 安全检查
            if not os.path.abspath(target_dir).startswith(os.path.abspath(root_path)):
                return jsonify({'success': False, 'error': '无权访问该路径'}), 403
            
            # 确保目录存在
            os.makedirs(target_dir, exist_ok=True)
            
            count = 0
            for file in files:
                if file.filename:
                    filename = file.filename
                    file_path = os.path.join(target_dir, filename)
                    
                    # 再次检查完整路径
                    if not os.path.abspath(file_path).startswith(os.path.abspath(root_path)):
                        continue
                    
                    file.save(file_path)
                    count += 1
            
            return jsonify({'success': True, 'message': f'成功上传{count}个文件', 'count': count})
            
        except Exception as e:
            return jsonify({'success': False, 'error': f'上传失败: {str(e)}'}), 500




